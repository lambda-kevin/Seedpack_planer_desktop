import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import statistics
import random

random.seed(42)

# --- Leer productos del historico ---
wb_src = openpyxl.load_workbook('Archivos Historicos/ventasXproducto.xlsx')
ws_src = wb_src.active

products = {}
ventas_qty = defaultdict(float)

for row in ws_src.iter_rows(min_row=2, values_only=True):
    linea = row[0]
    codigo = row[1]
    producto = row[6]
    costo = row[10]
    precio = row[9]
    cantidad = row[8]
    if codigo and producto:
        if codigo not in products:
            products[codigo] = {'nombre': producto, 'costos': [], 'precios': [], 'linea': linea or 'SEEDPACK'}
        if costo:
            products[codigo]['costos'].append(float(costo))
        if precio:
            products[codigo]['precios'].append(float(precio))
        if cantidad:
            ventas_qty[codigo] += float(cantidad)

# --- Crear libro de Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Stock Bodega"

COLOR_HEADER   = "1F4E79"
COLOR_SUBHEADER = "2E75B6"
COLOR_ROW_ALT  = "D6E4F0"
COLOR_ROW_NORM = "FFFFFF"
COLOR_VERDE    = "70AD47"
COLOR_AMARILLO = "FFD966"
COLOR_ROJO     = "FF5050"

def hfill(color):
    return PatternFill(fill_type='solid', fgColor=color)

thin = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

# Fila titulo
ws.merge_cells('A1:L1')
ws['A1'] = 'SEEDPACK - INVENTARIO DE BODEGA'
ws['A1'].font = Font(name='Calibri', bold=True, color="FFFFFF", size=14)
ws['A1'].fill = hfill(COLOR_HEADER)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:L2')
ws['A2'] = 'Fecha de corte: Abril 2026'
ws['A2'].font = Font(name='Calibri', italic=True, color="FFFFFF", size=10)
ws['A2'].fill = hfill(COLOR_SUBHEADER)
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18

# Encabezados
headers = [
    'Codigo', 'Linea', 'Nombre Producto',
    'Costo Unitario (COP)', 'Precio Venta (COP)', 'Margen (%)',
    'Stock Actual (und)', 'Stock Minimo (und)', 'Stock Maximo (und)',
    'Valor Inventario (COP)', 'Rotacion Historica', 'Estado'
]
col_widths = [16, 14, 45, 18, 18, 10, 16, 16, 16, 22, 16, 12]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=ci, value=h)
    cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    cell.fill = hfill(COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[3].height = 36

# Palabras clave a excluir (no son productos fisicos en bodega)
skip_kw = ['flete', 'comisi', 'retal', 'servicio', 'export']

row_num = 4
for codigo, data in sorted(products.items()):
    linea  = str(data['linea'] or '')
    nombre = str(data['nombre'] or '')

    if any(kw in linea.lower() or kw in nombre.lower() for kw in skip_kw):
        continue

    avg_costo  = statistics.mean(data['costos'])  if data['costos']  else 0
    avg_precio = statistics.mean(data['precios']) if data['precios'] else 0
    margen = ((avg_precio - avg_costo) / avg_precio * 100) if avg_precio > 0 else 0

    total_vnd = ventas_qty[codigo]

    # Stock inventado segun volumen historico de ventas
    if total_vnd > 500000:
        stock = random.randint(15000, 50000); rot = 'Alta'
    elif total_vnd > 100000:
        stock = random.randint(5000, 20000);  rot = 'Alta'
    elif total_vnd > 20000:
        stock = random.randint(1000, 8000);   rot = 'Media'
    elif total_vnd > 5000:
        stock = random.randint(200, 2000);    rot = 'Media'
    elif total_vnd > 1000:
        stock = random.randint(50, 500);      rot = 'Baja'
    else:
        stock = random.randint(10, 200);      rot = 'Baja'

    stock_min = max(int(stock * 0.2), 10)
    stock_max = int(stock * 2.5)
    valor_inv = stock * avg_costo

    if stock <= stock_min:
        estado = 'AGOTANDO'; ec = COLOR_ROJO;     fc = "FFFFFF"
    elif stock <= stock_min * 1.5:
        estado = 'BAJO';     ec = COLOR_AMARILLO; fc = "000000"
    else:
        estado = 'OK';       ec = COLOR_VERDE;    fc = "000000"

    bg = COLOR_ROW_ALT if (row_num % 2 == 0) else COLOR_ROW_NORM

    fila = [
        codigo, linea, nombre,
        round(avg_costo, 2), round(avg_precio, 2), round(margen, 1),
        stock, stock_min, stock_max,
        round(valor_inv, 2), rot, estado
    ]

    for ci, val in enumerate(fila, 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.border = thin
        if ci == 12:
            cell.font = Font(name='Calibri', bold=True, size=10, color=fc)
            cell.fill = hfill(ec)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.font = Font(name='Calibri', size=10)
            cell.fill = hfill(bg)
            if ci in [4, 5, 10]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif ci == 6:
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci in [7, 8, 9]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    row_num += 1

ws.freeze_panes = 'A4'
ws.auto_filter.ref = f'A3:L{row_num - 1}'

# --- Hoja Resumen ---
ws2 = wb.create_sheet("Resumen")
ws2.merge_cells('A1:B1')
ws2['A1'] = 'RESUMEN - INVENTARIO BODEGA SEEDPACK'
ws2['A1'].font = Font(name='Calibri', bold=True, size=13, color="FFFFFF")
ws2['A1'].fill = hfill(COLOR_HEADER)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

for ci, h in enumerate(['Indicador', 'Valor'], 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    c.fill = hfill(COLOR_SUBHEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 22

last = row_num - 1
resumen = [
    ('Total referencias en bodega',  row_num - 4),
    ('Valor total inventario (COP)',  f"=SUM('Stock Bodega'!J4:J{last})"),
    ('Productos en estado OK',        f"=COUNTIF('Stock Bodega'!L4:L{last},\"OK\")"),
    ('Productos en estado BAJO',      f"=COUNTIF('Stock Bodega'!L4:L{last},\"BAJO\")"),
    ('Productos AGOTANDO',            f"=COUNTIF('Stock Bodega'!L4:L{last},\"AGOTANDO\")"),
]

for i, (label, val) in enumerate(resumen, 3):
    bg = COLOR_ROW_ALT if i % 2 == 0 else COLOR_ROW_NORM
    c1 = ws2.cell(row=i, column=1, value=label)
    c2 = ws2.cell(row=i, column=2, value=val)
    for c in [c1, c2]:
        c.font = Font(name='Calibri', size=10)
        c.fill = hfill(bg)
        c.border = thin
        c.alignment = Alignment(horizontal='left', vertical='center')
    if i == 4:
        c2.number_format = '#,##0.00'

# --- Guardar ---
output = 'Archivos Historicos/stock_bodega.xlsx'
wb.save(output)
print(f"Archivo creado: {output}")
print(f"Total productos incluidos: {row_num - 4}")
