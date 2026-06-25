# -*- coding: utf-8 -*-
"""
SeedPack Planner — 4 pestanas: Ordenes de Produccion + Dashboard + Plan Diario + Tablas Excel
"""
import os, sys, threading, shutil, time, json
from datetime import datetime
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk

_SRC_DIR = os.path.dirname(os.path.abspath(__file__))
if getattr(sys, 'frozen', False):
    DATA_DIR = os.path.dirname(sys.executable)
else:
    DATA_DIR = os.path.join(os.path.dirname(_SRC_DIR), "app")

# Si Archivos Historicos no existe en DATA_DIR (ej: exe corriendo desde dist/),
# buscar en la carpeta app/ adyacente como fallback
if not os.path.isdir(os.path.join(DATA_DIR, "Archivos Historicos")):
    _alt = os.path.join(os.path.dirname(DATA_DIR), "app")
    if os.path.isdir(os.path.join(_alt, "Archivos Historicos")):
        DATA_DIR = os.path.normpath(_alt)

sys.path.insert(0, os.path.join(_SRC_DIR, "pasos"))

ARCHIVOS = [
    {"key": "arch_ventas",   "titulo": "Historial de ventas",               "numero": "01",
     "descripcion": "Ventas historicas por producto. Base de entrenamiento de los modelos ML.",
     "default_file": "ventasXproducto.xlsx",                 "preloaded": True},
    {"key": "arch_ops_hist", "titulo": "Historico de ordenes de produccion", "numero": "02",
     "descripcion": "Mapeo codigo vs producto y tamanos de lote historicos por referencia.",
     "default_file": "INEDITTO_OP.xlsx",                     "preloaded": True},
    {"key": "arch_pedidos",  "titulo": "Listado de pedidos pendientes",      "numero": "03",
     "descripcion": "Pedidos sin despachar en el periodo. Para planificar referencias ocasionales.",
     "default_file": "INEDITTO_Listado_Pedidos.xlsx",        "preloaded": False},
    {"key": "arch_ops_proc", "titulo": "OPs en proceso",                     "numero": "04",
     "descripcion": "Ordenes activas cuya produccion llega en los proximos meses del MRP.",
     "default_file": "ordenes de produccion en proceso.xlsx","preloaded": False},
    {"key": "arch_bodega",   "titulo": "Existencias en bodega (saldo actual)", "numero": "05",
     "descripcion": "Existencias por lote con saldo real. Usar ExistBodega (no entradasInventario).",
     "default_file": "Bodega.xlsx",                          "preloaded": False},
    {"key": "arch_lote_min",  "titulo": "Lotes minimos de produccion",       "numero": "06",
     "descripcion": "Lote minimo por codigo PT. Produccion redondeada al multiplo del lote.",
     "default_file": "Codigos lote minimo.xlsx",             "preloaded": True},
    {"key": "arch_entradas",  "titulo": "Entradas de inventario (OC)",       "numero": "07",
     "descripcion": "Registro de entradas a bodega con columna OC. Permite cruzar con OPs en proceso para calcular unidades pendientes por OP.",
     "default_file": "INEDITTO_entradasInventario.xlsx",     "preloaded": False, "optional": True},
    {"key": "arch_plan_comercial", "titulo": "Plan comercial",               "numero": "08",
     "descripcion": "Proyeccion del equipo comercial por producto y mes. Genera hoja Comparativa_Comercial para validar coherencia con el modelo ML.",
     "default_file": "plan_comercial.xlsx",                  "preloaded": True,  "optional": True},
]

# ── Paleta visual ──────────────────────────────────────────────────────────────
C_HEADER   = "#1B3D5F"; C_SIDEBAR  = "#162B45"; C_MAIN     = "#F1F5F9"
C_CARD     = "#FFFFFF";  C_ACENTO   = "#2563EB"; C_ACENTO_H = "#1D4ED8"
C_TEXTO    = "#1E293B";  C_GRIS     = "#64748B"; C_DIVIDER  = "#CBD5E1"
C_CANCEL   = "#475569";  C_CANCEL_H = "#334155"; C_SID_FG   = "#E2E8F0"
C_SID_HINT = "#94A3B8";  C_SID_LBL  = "#93C5FD"; C_SID_ENT  = "#243F60"
C_SID_BORD = "#1A3550";  C_NUM_BG   = "#DBEAFE"; C_NUM_FG   = "#1E40AF"
C_SUCCESS  = "#16A34A";  C_SUCCESS_H= "#15803D"
C_DASH_BG  = "#0F1E2E";  C_DASH_CARD= "#162B45"; C_DASH_BORD= "#1E3A58"
C_DASH_FG  = "#E2E8F0";  C_PDF_BTN  = "#7C3AED"; C_PDF_BTN_H= "#6D28D9"
SIDEBAR_W  = 272
FIG_BG     = "#0F1E2E"
C_OK_BG    = "#F0FDF4"; C_OK_BORD   = "#16A34A"
C_OK_NUM_BG= "#DCFCE7"; C_OK_NUM_FG = "#166534"
C_PN_BG    = "#FFFBEB"; C_PN_BORD   = "#D97706"
C_PN_NUM_BG= "#FEF3C7"; C_PN_NUM_FG = "#92400E"

SHEETS_CONFIG = {
    "Plan_Diario":        {"header": 1, "cols": ["Fecha","Codigo PT","Referencia","Cantidad a Producir","Tipo","Fecha Entrega"]},
    "Calendario_Semanal": {"header": 1, "cols": None},
    "OPs_En_Proceso":     {"header": 1, "cols": ["OP","Tipo de Trabajo","Cod. Producto","Referencia","Cliente","Cant. Aprobada","Fecha Programada","Compromiso Cliente"]},
    "Pedidos_Ocasionales":{"header": 1, "cols": ["Pedido","Codigo PT","Referencia","Cliente","Fecha Entrega","Cant. Pedida","Stock Disp.","OPs Proceso","Neto s/Ajuste","Lote Minimo","Neto a Producir"]},
    "Inventario_Actual":  {"header": 1, "cols": ["Cod. PT","Referencia","# Bodegas","Saldo","Costo Unit.","Costo Total"]},
    "Inventario_Detalle": {"header": 1, "cols": ["Bodega","Cod. PT","Referencia","Cant. Ingresada","Saldo","Costo","Total","OP/OC"]},
    "Gestion_Pedidos":    {"header": 0, "cols": ["Codigo","Descripcion","Pedidos","Saldo Pendiente","Stock Bodega","OPs Proceso","Proyeccion ML","Disponible","Demanda Total","Deficit","Lote Minimo","A Producir","Estado","Justificacion"]},
    "Saldo_OPs":          {"header": 1, "cols": ["OP","Tipo","Cod. Producto","Referencia","Cant. Aprobada","Cant. Producida","Cant. Pendiente","Estado","Fecha Prog.","Compromiso"]},
    "Comparativa_Comercial": {"header": 2, "cols": None},
}

_MAPEOS_PATH   = os.path.join(DATA_DIR, "column_mappings.json")
_SETTINGS_PATH = os.path.join(DATA_DIR, "app_settings.json")

# Esquema de columnas esperadas por archivo — usado por el modal de configuracion
_COL_SCHEMA = {
    "arch_ventas": {
        "titulo": "Historial de Ventas",
        "campos": [
            ("Fecha Fra.",         "Fecha de la factura",                True),
            ("Producto Terminado", "Nombre/descripcion del producto",    True),
            ("Cantidad",           "Unidades vendidas",                  True),
            ("Código",             "Codigo del producto (SKU)",          True),
            ("Factura",            "Numero de factura",                  False),
        ]
    },
    "arch_bodega": {
        "titulo": "Bodega / Existencias",
        "campos": [
            ("Bodega",           "Nombre de la bodega",                  True),
            ("Cód. PT",          "Codigo de producto terminado",        True),
            ("Saldo",            "Unidades en inventario",               True),
            ("Referencia",       "Descripcion del producto",             False),
            ("Costo",            "Costo unitario",                       False),
            ("Cant. Ingresada",  "Cantidad ingresada originalmente",     False),
            ("OP/OC",            "Numero de OP o de OC",                 False),
        ]
    },
    "arch_pedidos": {
        "titulo": "Listado de Pedidos",
        "campos": [
            ("Pedido",           "Numero del pedido",                    True),
            ("Fecha de entrega", "Fecha de entrega pactada",             True),
            ("Código PT",        "Codigo del producto pedido",           True),
            ("Referencia",       "Descripcion del producto",             True),
            ("Cantidad",         "Cantidad pendiente / saldo",           True),
            ("Cliente",          "Nombre del cliente",                   True),
            ("Línea",            "Linea de producto",                    True),
            ("Remisión",         "N. remision (si fue despachado)",      False),
            ("Factura",          "N. factura (si fue facturado)",        False),
        ]
    },
    "arch_ops_proc": {
        "titulo": "OPs en Proceso",
        "campos": [
            ("OP",               "Numero de la orden",                   True),
            ("Tipo de Trabajo",  "Tipo de trabajo",                      True),
            ("Cód. Producto",    "Codigo del producto",                  True),
            ("Referencia",       "Descripcion del producto",             True),
            ("Cant. Aprobada",   "Cantidad aprobada en la OP",           True),
            ("Fecha Programada", "Fecha programada de entrega",          True),
            ("Cliente",          "Cliente de la orden",                  False),
            ("Compromiso Cliente","Fecha compromiso con el cliente",     False),
        ]
    },
    "arch_ops_hist": {
        "titulo": "Historico de OPs",
        "campos": [
            ("Cód. Producto",    "Codigo del producto",                  True),
            ("Referencia",       "Descripcion del producto",             True),
            ("Cant. Aprobada",   "Cantidad aprobada",                    True),
        ]
    },
    "arch_entradas": {
        "titulo": "Entradas de Inventario",
        "campos": [
            ("OC",               "Numero de OC / OP de compra",          True),
            ("Cantidad",         "Cantidad ingresada",                   True),
            ("Prod. Terminado",  "Producto terminado",                   False),
            ("Bodega Sale",      "Bodega destino",                       False),
        ]
    },
}


# Aliases que el pipeline ya maneja internamente — si alguno coincide, el campo esta cubierto
_ALIAS_VALIDACION = {
    "arch_ventas": {
        "Fecha Fra.":         ["fecha fra.", "fecha factura", "fecha fac.", "fecha de factura"],
        "Producto Terminado": ["producto terminado", "producto"],
        "Cantidad":           ["cantidad", "cant.", "unidades"],
        "Código":             ["código", "codigo", "cód.", "cod."],
        "Factura":            ["factura"],
    },
    "arch_bodega": {
        "Bodega":          ["bodega", "bodega sale", "bodega_sale", "almacen"],
        "Cód. PT":         ["cód. pt", "cod. pt", "prod. terminado", "prod.terminado",
                            "producto terminado"],
        "Saldo":           ["saldo", "saldo en inventario", "saldo inventario",
                            "cantidad", "cant.", "stock"],
        "Referencia":      ["referencia", "referenciapt", "descripcion", "descripción",
                            "producto", "nombre"],
        "Costo":           ["costo", "vr. unitario", "vr unitario", "precio unitario"],
        "Cant. Ingresada": ["cant. ingresada", "cant ingresada", "cantidad ingresada"],
        "OP/OC":           ["op/oc", "o.p. / o.c.", "op / oc", "oc"],
    },
    "arch_pedidos": {
        "Pedido":           ["pedido"],
        "Fecha de entrega": ["fecha de entrega", "fecha de entrega lead time",
                             "fecha entrega comercial"],
        "Código PT":        ["código pt", "codigo pt", "cod pt", "codigo", "cód. pt"],
        "Referencia":       ["referencia", "descripcion ", "descripcion", "descripción"],
        "Cantidad":         ["cantidad", "saldo pendiente", "cant. pendiente"],
        "Cliente":          ["cliente"],
        "Línea":            ["línea", "linea"],
        "Remisión":         ["remisión", "remision"],
        "Factura":          ["factura"],
    },
    "arch_ops_proc": {
        "OP":                 ["op"],
        "Tipo de Trabajo":    ["tipo de trabajo", "tipo trabajo"],
        "Cód. Producto":      ["cód. producto", "cod. producto", "código producto",
                               "codigo producto"],
        "Referencia":         ["referencia", "descripcion"],
        "Cant. Aprobada":     ["cant. aprobada", "cantidad aprobada", "cant aprobada"],
        "Fecha Programada":   ["fecha programada", "fecha prog."],
        "Cliente":            ["cliente"],
        "Compromiso Cliente": ["compromiso cliente", "compromiso", "fecha compromiso"],
    },
    "arch_ops_hist": {
        "Cód. Producto":  ["cód. producto", "cod. producto", "código producto",
                           "codigo producto"],
        "Referencia":     ["referencia", "descripcion"],
        "Cant. Aprobada": ["cant. aprobada", "cantidad aprobada", "cant aprobada"],
    },
    "arch_entradas": {
        "OC":              ["oc"],
        "Cantidad":        ["cantidad", "cant.", "cant. ingresada"],
        "Prod. Terminado": ["prod. terminado", "prod.terminado", "producto terminado"],
        "Bodega Sale":     ["bodega sale", "bodega_sale", "bodega"],
    },
}


class SeedPackPlanner:

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("SeedPack Planner")
        self.root.configure(bg=C_HEADER)
        self.root.resizable(True, True)
        self.root.minsize(1024, 640)

        _hoy = datetime.today()
        self.fi_var  = tk.StringVar(value=_hoy.replace(day=1).strftime("%Y-%m-%d"))
        self.ff_var  = tk.StringVar(value=datetime(_hoy.year, 12, 31).strftime("%Y-%m-%d"))
        self.file_vars = {cfg["key"]: tk.StringVar() for cfg in ARCHIVOS}
        _ah = os.path.join(DATA_DIR, "Archivos Historicos")
        # 1) Cargar defaults de Archivos Historicos (preloaded o con default_file)
        for _cfg in ARCHIVOS:
            if _cfg.get("default_file"):
                _p = os.path.join(_ah, _cfg["default_file"])
                if os.path.isfile(_p):
                    self.file_vars[_cfg["key"]].set(_p)
        # 2) Sobrescribir con rutas guardadas por el usuario en sesiones anteriores
        try:
            if os.path.isfile(_SETTINGS_PATH):
                with open(_SETTINGS_PATH, encoding="utf-8") as _sf:
                    _saved = json.load(_sf).get("file_paths", {})
                for _k, _p in _saved.items():
                    if _k in self.file_vars and _p and os.path.isfile(_p):
                        self.file_vars[_k].set(_p)
        except Exception:
            pass

        self._last_excel  = None
        self._dash_loaded = None
        self._plan_data   = []
        self._plan_sort_col = None
        self._plan_sort_rev = False
        self._tablas_loaded = None
        self._gestion_data   = []
        self._gestion_loaded = None
        self._saldo_ops_data = []

        self._build_ui()
        self._centrar(1280, 780)
        self.root.mainloop()

    # ── UI principal ──────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()
        self._build_notebook()

    def _build_header(self):
        h = tk.Frame(self.root, bg=C_HEADER)
        h.pack(fill="x")
        inner = tk.Frame(h, bg=C_HEADER)
        inner.pack(fill="x", padx=26, pady=12)
        left = tk.Frame(inner, bg=C_HEADER)
        left.pack(side="left")
        tk.Label(left, text="SeedPack Planner", font=("Segoe UI", 17, "bold"),
                 bg=C_HEADER, fg="white").pack(anchor="w")
        tk.Label(left, text="Planificacion de produccion basada en machine learning",
                 font=("Segoe UI", 9), bg=C_HEADER, fg="#7EB8E0").pack(anchor="w", pady=(2,0))
        right = tk.Frame(inner, bg=C_HEADER)
        right.pack(side="right", anchor="e")
        badge = tk.Frame(right, bg="#1E3A5F", padx=12, pady=6)
        badge.pack()
        tk.Label(badge, text="Pipeline  ML", font=("Segoe UI", 8, "bold"),
                 bg="#1E3A5F", fg="#BFDBFE").pack()

    def _build_notebook(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("SP.TNotebook", background=C_HEADER, borderwidth=0, tabmargins=[0,0,0,0])
        style.configure("SP.TNotebook.Tab", background="#1E3A5F", foreground="#93C5FD",
                         font=("Segoe UI", 9, "bold"), padding=[18, 9], borderwidth=0)
        style.map("SP.TNotebook.Tab",
                  background=[("selected", C_ACENTO), ("active", "#1E4ED8")],
                  foreground=[("selected", "white"),  ("active", "white")])

        self.notebook = ttk.Notebook(self.root, style="SP.TNotebook")
        self.notebook.pack(fill="both", expand=True)

        self.tab_op      = tk.Frame(self.notebook, bg=C_MAIN)
        self.tab_dash    = tk.Frame(self.notebook, bg=C_DASH_BG)
        self.tab_plan    = tk.Frame(self.notebook, bg=C_MAIN)
        self.tab_tablas  = tk.Frame(self.notebook, bg=C_MAIN)
        self.tab_gestion = tk.Frame(self.notebook, bg=C_MAIN)
        self.tab_saldo   = tk.Frame(self.notebook, bg=C_MAIN)

        self.notebook.add(self.tab_op,      text="  Ordenes de Produccion  ")
        self.notebook.add(self.tab_dash,    text="  Dashboard Ejecutivo  ")
        self.notebook.add(self.tab_plan,    text="  Plan Diario  ")
        self.notebook.add(self.tab_tablas,  text="  Tablas del Excel  ")
        self.notebook.add(self.tab_gestion, text="  Gestion de Pedidos  ")
        self.notebook.add(self.tab_saldo,   text="  Saldo de OPs  ")

        self._build_tab_op()
        self._build_tab_dashboard()
        self._build_tab_plan()
        self._build_tab_tablas()
        self._build_tab_gestion()
        self._build_tab_saldo_ops()
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)
        # Scroll inicial: la primera pestaña activa es Ordenes
        self.root.after(100, lambda: self._rebind_scroll(self._canvas))

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 1 — Ordenes de Produccion
    # ═════════════════════════════════════════════════════════════════════════

    def _build_tab_op(self):
        body = tk.Frame(self.tab_op, bg=C_MAIN)
        body.pack(fill="both", expand=True)
        sidebar = tk.Frame(body, bg=C_SIDEBAR, width=SIDEBAR_W)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        self._build_sidebar(sidebar)
        self._build_main_area(body)

    def _build_sidebar(self, parent):
        tk.Frame(parent, bg=C_SIDEBAR, height=22).pack(fill="x")
        tk.Label(parent, text="PARAMETROS DEL PERIODO", font=("Segoe UI", 7, "bold"),
                 bg=C_SIDEBAR, fg=C_SID_LBL, anchor="w").pack(fill="x", padx=20, pady=(0,14))

        self._sidebar_field(parent, "Fecha inicial",          self.fi_var,  "Formato YYYY-MM-DD")
        self._sidebar_field(parent, "Fecha final",            self.ff_var,  "Formato YYYY-MM-DD")

        tk.Frame(parent, bg="#1E3A58", height=1).pack(fill="x", padx=20, pady=(6,20))

        btn_wrap = tk.Frame(parent, bg=C_SIDEBAR)
        btn_wrap.pack(fill="x", padx=20)

        self._make_btn(btn_wrap, "  Generar Plan de Produccion  ",
                       C_ACENTO, C_ACENTO_H, self._ejecutar).pack(fill="x", pady=(0,8), ipady=6)

        self._btn_exportar_xl = self._make_btn(btn_wrap, "  Exportar Excel  ",
                                               C_SUCCESS, C_SUCCESS_H, self._exportar_excel)
        self._btn_exportar_xl.pack(fill="x", pady=(0,8), ipady=6)
        self._btn_exportar_xl.configure(state="disabled",
                                        bg="#1E4033", disabledforeground="#4ADE80")

        self._btn_exportar_pdf = self._make_btn(btn_wrap, "  Exportar PDF  ",
                                                C_PDF_BTN, C_PDF_BTN_H, self._exportar_pdf)
        self._btn_exportar_pdf.pack(fill="x", pady=(0,8), ipady=6)
        self._btn_exportar_pdf.configure(state="disabled",
                                         bg="#3B2068", disabledforeground="#C4B5FD")

        self._make_btn(btn_wrap, "  Cancelar  ",
                       C_CANCEL, C_CANCEL_H, self.root.destroy).pack(fill="x", ipady=6)

        tk.Frame(parent, bg=C_SIDEBAR, height=18).pack(fill="x")
        tk.Label(parent, text="El proceso puede tardar\nvarios minutos.",
                 font=("Segoe UI", 7), bg=C_SIDEBAR, fg=C_SID_HINT,
                 justify="center").pack()

    def _sidebar_field(self, parent, label, var, hint):
        wrap = tk.Frame(parent, bg=C_SIDEBAR)
        wrap.pack(fill="x", padx=20, pady=(0,14))
        tk.Label(wrap, text=label, font=("Segoe UI", 8, "bold"),
                 bg=C_SIDEBAR, fg=C_SID_LBL, anchor="w").pack(fill="x")
        e = tk.Entry(wrap, textvariable=var, font=("Segoe UI", 10),
                     bg=C_SID_ENT, fg=C_SID_FG, insertbackground=C_SID_FG,
                     relief="flat", bd=0, highlightthickness=1,
                     highlightbackground=C_SID_BORD, highlightcolor=C_ACENTO)
        e.pack(fill="x", pady=(5,0), ipady=6)
        tk.Label(wrap, text=hint, font=("Segoe UI", 7),
                 bg=C_SIDEBAR, fg=C_SID_HINT, anchor="w").pack(fill="x", pady=(3,0))

    def _make_btn(self, parent, text, bg, hover_bg, command):
        btn = tk.Button(parent, text=text, font=("Segoe UI", 9, "bold"),
                        bg=bg, fg="white", activebackground=hover_bg, activeforeground="white",
                        relief="flat", padx=10, cursor="hand2", bd=0, command=command)
        btn.bind("<Enter>", lambda e: btn.configure(bg=hover_bg) if btn["state"] != "disabled" else None)
        btn.bind("<Leave>", lambda e: btn.configure(bg=bg)       if btn["state"] != "disabled" else None)
        return btn

    def _build_main_area(self, parent):
        outer = tk.Frame(parent, bg=C_MAIN)
        outer.pack(side="left", fill="both", expand=True)
        canvas    = tk.Canvas(outer, bg=C_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._canvas = canvas
        contenido = tk.Frame(canvas, bg=C_MAIN)
        self._win_id = canvas.create_window((0,0), window=contenido, anchor="nw")
        contenido.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(self._win_id, width=e.width))
        self._build_files_grid(contenido)

    def _build_files_grid(self, parent):
        tk.Frame(parent, bg=C_MAIN, height=20).pack(fill="x")
        hdr = tk.Frame(parent, bg=C_MAIN)
        hdr.pack(fill="x", padx=24)
        tk.Label(hdr, text="ARCHIVOS DE ENTRADA", font=("Segoe UI", 8, "bold"),
                 bg=C_MAIN, fg=C_GRIS, anchor="w").pack(side="left")
        tk.Label(hdr, text="6 requeridos + 1 opcional  •  formato .xlsx",
                 font=("Segoe UI", 8), bg=C_MAIN, fg=C_GRIS, anchor="e").pack(side="right")
        tk.Frame(parent, bg=C_DIVIDER, height=1).pack(fill="x", padx=24, pady=(6,16))
        grid = tk.Frame(parent, bg=C_MAIN)
        grid.pack(fill="both", expand=True, padx=24, pady=(0,24))
        grid.columnconfigure(0, weight=1, uniform="col")
        grid.columnconfigure(1, weight=1, uniform="col")
        for idx, cfg in enumerate(ARCHIVOS):
            row, col = divmod(idx, 2)
            self._file_card(grid, cfg, row, col)

    def _file_card(self, grid, cfg, row, col):
        preloaded = bool(cfg.get("preloaded"))
        optional  = bool(cfg.get("optional"))
        pad = (0, 8) if col == 0 else (8, 0)
        var = self.file_vars[cfg["key"]]

        def _colors(loaded):
            if loaded:
                return C_OK_BG, C_OK_BORD, C_OK_NUM_BG, C_OK_NUM_FG, "#F0FDF4"
            if optional:
                return "#F8FAFC", C_DIVIDER, "#F1F5F9", C_GRIS, "#FFFFFF"
            return C_PN_BG, C_PN_BORD, C_PN_NUM_BG, C_PN_NUM_FG, "#FFFBEB"

        def _tag_text(loaded):
            if loaded:
                return "Pre-cargado ✓" if preloaded else "Cargado ✓"
            return "Opcional" if optional else "Requerido"

        is_loaded = bool(var.get().strip()) and os.path.isfile(var.get().strip())
        card_bg, bord_col, num_bg, num_fg, ent_bg = _colors(is_loaded)

        card = tk.Frame(grid, bg=card_bg, highlightbackground=bord_col, highlightthickness=1)
        card.grid(row=row, column=col, padx=pad, pady=(0, 10), sticky="nsew")
        inner = tk.Frame(card, bg=card_bg)
        inner.pack(fill="x", padx=14, pady=13)

        title_row = tk.Frame(inner, bg=card_bg)
        title_row.pack(fill="x", pady=(0, 4))
        badge_lbl = tk.Label(title_row, text=cfg["numero"], font=("Segoe UI", 7, "bold"),
                             bg=num_bg, fg=num_fg, padx=7, pady=2)
        badge_lbl.pack(side="left", padx=(0, 9))
        title_lbl = tk.Label(title_row, text=cfg["titulo"], font=("Segoe UI", 9, "bold"),
                             bg=card_bg, fg=C_TEXTO, anchor="w")
        title_lbl.pack(side="left")
        tag_lbl = tk.Label(title_row, text=_tag_text(is_loaded), font=("Segoe UI", 7, "bold"),
                           bg=card_bg, fg=bord_col)
        tag_lbl.pack(side="right")

        desc_lbl = tk.Label(inner, text=cfg["descripcion"], font=("Segoe UI", 8),
                            bg=card_bg, fg=C_GRIS, anchor="w", justify="left", wraplength=340)
        desc_lbl.pack(fill="x", pady=(0, 9))

        fila = tk.Frame(inner, bg=card_bg)
        fila.pack(fill="x")
        ent = tk.Entry(fila, textvariable=var, font=("Segoe UI", 9),
                       bg=ent_bg, fg=C_TEXTO, relief="flat", bd=0,
                       highlightthickness=1, highlightbackground=bord_col,
                       highlightcolor=C_ACENTO)
        ent.pack(side="left", fill="x", expand=True, padx=(0, 8), ipady=5)

        btn_txt = "Actualizar  ↺" if preloaded else "Examinar  ↗"
        btn = tk.Button(fila, text=btn_txt, font=("Segoe UI", 8, "bold"),
                        bg=C_ACENTO, fg="white", activebackground=C_ACENTO_H, activeforeground="white",
                        relief="flat", padx=12, pady=6, cursor="hand2", bd=0,
                        command=lambda v=var, c=cfg: self._elegir_archivo(v, c))
        btn.pack(side="right")
        btn.bind("<Enter>", lambda e: btn.configure(bg=C_ACENTO_H))
        btn.bind("<Leave>", lambda e: btn.configure(bg=C_ACENTO))

        def _refresh(*_):
            p = var.get().strip()
            loaded = bool(p) and os.path.isfile(p)
            cg, bc, nbg, nfg, ebg = _colors(loaded)
            card.configure(bg=cg, highlightbackground=bc)
            for w in (inner, title_row, fila):
                w.configure(bg=cg)
            for w in (title_lbl, desc_lbl):
                w.configure(bg=cg)
            badge_lbl.configure(bg=nbg, fg=nfg)
            tag_lbl.configure(text=_tag_text(loaded), fg=bc, bg=cg)
            ent.configure(bg=ebg, highlightbackground=bc)
        var.trace_add("write", _refresh)


    # ═════════════════════════════════════════════════════════════════════════
    # TAB 2 — Dashboard Ejecutivo
    # ═════════════════════════════════════════════════════════════════════════

    def _build_tab_dashboard(self):
        topbar = tk.Frame(self.tab_dash, bg=C_DASH_CARD, height=54)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        tk.Label(topbar, text="DASHBOARD EJECUTIVO", font=("Segoe UI", 13, "bold"),
                 bg=C_DASH_CARD, fg=C_DASH_FG).pack(side="left", padx=24, pady=14)
        self._lbl_dash_ruta = tk.Label(topbar, text="Sin datos cargados",
                                        font=("Segoe UI", 8), bg=C_DASH_CARD, fg=C_SID_HINT)
        self._lbl_dash_ruta.pack(side="left", padx=6)

        btn_frame = tk.Frame(topbar, bg=C_DASH_CARD)
        btn_frame.pack(side="right", padx=24)
        self._make_btn(btn_frame, "  PDF  ", C_PDF_BTN, C_PDF_BTN_H,
                       self._exportar_pdf).pack(side="right", pady=10, ipady=4, padx=(4,0))
        self._make_btn(btn_frame, "  Actualizar  ", C_ACENTO, C_ACENTO_H,
                       self._actualizar_dashboard).pack(side="right", pady=10, ipady=4)

        outer = tk.Frame(self.tab_dash, bg=C_DASH_BG)
        outer.pack(fill="both", expand=True)
        self._dash_canvas = tk.Canvas(outer, bg=C_DASH_BG, highlightthickness=0)
        dash_scroll = ttk.Scrollbar(outer, orient="vertical", command=self._dash_canvas.yview)
        self._dash_canvas.configure(yscrollcommand=dash_scroll.set)
        dash_scroll.pack(side="right", fill="y")
        self._dash_canvas.pack(side="left", fill="both", expand=True)
        self._dash_content = tk.Frame(self._dash_canvas, bg=C_DASH_BG)
        self._dash_win_id  = self._dash_canvas.create_window((0,0), window=self._dash_content, anchor="nw")
        self._dash_content.bind("<Configure>", lambda e: self._dash_canvas.configure(
            scrollregion=self._dash_canvas.bbox("all")))
        def _on_canvas_cfg(e):
            self._dash_canvas.itemconfig(self._dash_win_id, width=e.width)
            self._on_dash_resize()
        self._dash_canvas.bind("<Configure>", _on_canvas_cfg)
        self._dash_placeholder()

    def _dash_placeholder(self):
        for w in self._dash_content.winfo_children():
            w.destroy()
        wrapper = tk.Frame(self._dash_content, bg=C_DASH_BG)
        wrapper.pack(expand=True, fill="both", pady=120)
        tk.Label(wrapper, text="Sin datos", font=("Segoe UI", 22, "bold"),
                 bg=C_DASH_BG, fg=C_DASH_BORD).pack()
        tk.Label(wrapper,
                 text="Genera un plan de produccion o carga un Excel\ny presiona Actualizar para ver el dashboard.",
                 font=("Segoe UI", 10), bg=C_DASH_BG, fg=C_SID_HINT, justify="center").pack(pady=(10,0))

    def _on_dash_resize(self):
        if not self._dash_loaded:
            return
        cw = self._dash_canvas.winfo_width()
        if abs(cw - getattr(self, '_dash_last_w', 0)) < 40:
            return
        job = getattr(self, '_dash_resize_job', None)
        if job:
            self.root.after_cancel(job)
        ruta = self._dash_loaded
        def _do():
            self._dash_resize_job = None
            self._dash_loaded = None
            self._renderizar_dashboard(ruta)
        self._dash_resize_job = self.root.after(400, _do)

    def _pedir_excel_dashboard(self):
        ruta = filedialog.askopenfilename(title="Seleccionar Excel de resultado",
                                          filetypes=[("Excel","*.xlsx *.xls")])
        if ruta:
            self._last_excel  = ruta
            self._dash_loaded = None
            self._renderizar_dashboard(ruta)

    def _actualizar_dashboard(self):
        ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
        ruta = (self._last_excel if self._last_excel and os.path.isfile(self._last_excel)
                else ruta_auto if os.path.isfile(ruta_auto) else None)
        if not ruta:
            messagebox.showinfo("Sin archivo", "No se encontro el archivo de resultados.\nGenera primero un plan de produccion.")
            return
        self._dash_loaded = None
        self._renderizar_dashboard(ruta)

    def _renderizar_dashboard(self, ruta):
        if self._dash_loaded == ruta:
            return
        if not ruta or not os.path.isfile(ruta):
            messagebox.showerror("Archivo no encontrado",
                                 f"No se encontro el archivo:\n{ruta}")
            return
        try:
            import pandas as pd
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            import matplotlib.ticker as mticker
            import matplotlib.patches as mpatches
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except Exception as exc:
            messagebox.showerror("Dependencia faltante", str(exc)); return
        try:
            xls = pd.ExcelFile(ruta)
        except Exception as exc:
            messagebox.showerror("Error leyendo Excel", str(exc)); return

        def _hoja(nombre, hdr=1):
            try:
                return pd.read_excel(xls, nombre, header=hdr) if nombre in xls.sheet_names else pd.DataFrame()
            except Exception:
                return pd.DataFrame()

        try:
            plan_df = _hoja("Plan_Diario")
            ops_df  = _hoja("OPs_En_Proceso")
            inv_df  = _hoja("Inventario_Actual")
            ocas_df = _hoja("Pedidos_Ocasionales")
            gest_df = _hoja("Gestion_Pedidos", hdr=0)
        except Exception as exc:
            messagebox.showerror("Error cargando hojas", str(exc)); return

        try:
            for w in self._dash_content.winfo_children():
                w.destroy()
            # Mostrar periodo del plan derivado del Excel
            try:
                import pandas as _pd2
                _fc = next((c for c in plan_df.columns
                            if "fecha" in str(c).lower() and "entrega" not in str(c).lower()), None)
                if _fc and not plan_df.empty:
                    _dates = _pd2.to_datetime(plan_df[_fc], errors="coerce").dropna()
                    if not _dates.empty:
                        _periodo = (f"  |  Periodo: {_dates.min().strftime('%d/%m/%Y')}"
                                    f" → {_dates.max().strftime('%d/%m/%Y')}")
                    else:
                        _periodo = ""
                else:
                    _periodo = ""
            except Exception:
                _periodo = ""
            self._lbl_dash_ruta.config(text=os.path.basename(ruta) + _periodo)
        except Exception as exc:
            messagebox.showerror("Error de interfaz", str(exc)); return

        PAD = 14; BG = C_DASH_CARD

        # Dimensiones responsivas basadas en el ancho del canvas
        self._dash_canvas.update_idletasks()
        _cw = max(self._dash_canvas.winfo_width() - 2 * PAD, 800)
        self._dash_last_w = self._dash_canvas.winfo_width()
        _DPI = 96
        _fh  = max(3.8, min(5.5, _cw / 280))             # alto charts (más alto)
        _fw3 = max(3.0, (_cw - 12) / 3 / _DPI)           # ancho fila 3 col
        _fwd = max(2.2, min(3.8, _cw * 0.20 / _DPI))     # ancho donut
        _fw1 = max(3.5, (_cw * 0.80 - 12) / 2 / _DPI)   # ancho fila 1 (2 + donut)
        _fw2 = max(4.0, min(9.0, (_cw - 6) / 2 / _DPI)) # ancho fila 2 col
        _kv  = max(13, min(22, int(_cw / 80)))            # fuente valor KPI
        _kl  = max(7,  min(10, int(_cw / 180)))           # fuente etiqueta KPI

        # Formatter compacto: usa K o M para ejes con numeros grandes
        def _fmt_units(x, _):
            if abs(x) >= 1_000_000: return f"{x/1_000_000:.1f}M"
            if abs(x) >= 1_000:     return f"{int(x/1_000)}K"
            return f"{int(x):,}"

        # Estilo global matplotlib para este dashboard
        import matplotlib as _mpl
        _mpl.rcParams.update({
            "font.family": "sans-serif",
            "axes.titlepad": 10,
        })

        # ── KPIs calculados desde hojas fuente ───────────────────────────────
        try:
            if not plan_df.empty:
                fecha_col = next((c for c in plan_df.columns if "fecha" in str(c).lower() and "entrega" not in str(c).lower()), None)
                cod_col   = next((c for c in plan_df.columns if "codigo" in str(c).lower() or "cód" in str(c).lower()), None)
                cant_col  = next((c for c in plan_df.columns if "cantidad" in str(c).lower()), None)
                tipo_col  = next((c for c in plan_df.columns if str(c).lower().strip() == "tipo"), None)

                dias  = int(plan_df[fecha_col].nunique())         if fecha_col else 0
                refs  = int(plan_df[cod_col].nunique())           if cod_col   else 0
                unds  = int(pd.to_numeric(plan_df[cant_col], errors="coerce").fillna(0).sum()) if cant_col else 0

                if tipo_col:
                    proy_mask = plan_df[tipo_col].astype(str).str.lower() == "proyeccion"
                    ocas_mask = plan_df[tipo_col].astype(str).str.lower() == "ocasional"
                    r_proy = int(plan_df.loc[proy_mask, cod_col].nunique()) if cod_col else 0
                    r_ocas = int(plan_df.loc[ocas_mask, cod_col].nunique()) if cod_col else 0
                    u_ocas = int(pd.to_numeric(plan_df.loc[ocas_mask, cant_col], errors="coerce").fillna(0).sum()) if cant_col else 0
                else:
                    r_proy = refs; r_ocas = 0; u_ocas = 0
            else:
                dias = refs = unds = r_proy = r_ocas = u_ocas = 0

            if not gest_df.empty:
                est_col = next((c for c in gest_df.columns if "estado" in str(c).lower()), None)
                stock_cub = int((gest_df[est_col].astype(str).str.strip() == "OK").sum()) if est_col else 0
            else:
                stock_cub = 0

            n_ops     = len(ops_df) if not ops_df.empty else 0
            inv_saldo = int(pd.to_numeric(inv_df["Saldo"], errors="coerce").fillna(0).sum()) if (not inv_df.empty and "Saldo" in inv_df.columns) else 0
            inv_costo = float(pd.to_numeric(inv_df["Costo Total"], errors="coerce").fillna(0).sum()) if (not inv_df.empty and "Costo Total" in inv_df.columns) else 0.0
        except Exception as exc:
            dias = refs = unds = r_proy = r_ocas = u_ocas = stock_cub = n_ops = inv_saldo = 0
            inv_costo = 0.0

        kpis = [
            ("Dias planificados",    f"{dias:,}",            "#2563EB"),
            ("Referencias en plan",  f"{refs:,}",            "#7C3AED"),
            ("Unidades a producir",  f"{unds:,}",            "#0891B2"),
            ("Refs. proyeccion",     f"{r_proy:,}",          "#16A34A"),
            ("Refs. ocasionales",    f"{r_ocas:,}",          "#EA580C"),
            ("Uds. ocasionales",     f"{u_ocas:,}",          "#DC2626"),
            ("Cubiertos por stock",  f"{stock_cub:,}",       "#CA8A04"),
            ("OPs en proceso",       f"{n_ops:,}",           "#0891B2"),
            ("Inventario (uds)",     f"{inv_saldo:,}",       "#16A34A"),
            ("Costo inventario",     f"${inv_costo/1e6:.1f}M","#7C3AED"),
        ]

        # ── Fila KPIs (2 filas x 5) ──────────────────────────────────────────
        for fila_kpi in range(2):
            kpi_row = tk.Frame(self._dash_content, bg=C_DASH_BG)
            kpi_row.pack(fill="x", padx=PAD, pady=(PAD if fila_kpi==0 else 6, 0))
            for i, (label, val, color) in enumerate(kpis[fila_kpi*5:(fila_kpi+1)*5]):
                card = tk.Frame(kpi_row, bg=BG, padx=14, pady=12)
                card.pack(side="left", fill="both", expand=True, padx=(0 if i==0 else 6, 0))
                tk.Frame(card, bg=color, width=4).pack(side="left", fill="y", padx=(0,10))
                inner = tk.Frame(card, bg=BG); inner.pack(side="left")
                tk.Label(inner, text=val, font=("Segoe UI", _kv, "bold"),
                         bg=BG, fg=color).pack(anchor="w")
                tk.Label(inner, text=label, font=("Segoe UI", _kl),
                         bg=BG, fg=C_SID_HINT).pack(anchor="w")

        def _embed(fig, parent):
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            c = FigureCanvasTkAgg(fig, master=parent)
            c.draw(); c.get_tk_widget().pack(fill="both", expand=True)
            plt.close(fig)

        # ── Fila 1 de graficas ────────────────────────────────────────────────
        charts_row1 = tk.Frame(self._dash_content, bg=C_DASH_BG)
        charts_row1.pack(fill="x", padx=PAD, pady=(10, 0))
        # Layout responsivo: 3 columnas con proporcion fija 2:2:1 (chart, chart, donut)
        # que escalan juntas en cualquier resolucion/escalado de pantalla.
        charts_row1.columnconfigure(0, weight=2, uniform="r1")
        charts_row1.columnconfigure(1, weight=2, uniform="r1")
        charts_row1.columnconfigure(2, weight=1, uniform="r1")
        charts_row1.rowconfigure(0, weight=1)

        # Grafica 1 — Top 10 referencias por unidades
        if not plan_df.empty and "Cantidad a Producir" in plan_df.columns and "Referencia" in plan_df.columns:
            try:
                top10 = (plan_df.groupby("Referencia")["Cantidad a Producir"]
                         .sum().sort_values(ascending=False).head(10))
                fig, ax = plt.subplots(figsize=(_fw1, _fh), facecolor=FIG_BG,
                                       layout="constrained")
                ax.set_facecolor(FIG_BG)
                colors_b = plt.cm.Blues_r([0.3 + 0.5*i/max(len(top10)-1,1) for i in range(len(top10))])
                labels = [str(r)[:26] for r in top10.index[::-1]]
                ax.barh(labels, top10.values[::-1], color=colors_b, height=0.65)
                ax.set_title("Top 10 Referencias por Unidades", color=C_DASH_FG,
                             fontsize=9, fontweight="bold")
                ax.tick_params(colors=C_DASH_FG, labelsize=7)
                ax.yaxis.tick_left()
                for sp in ax.spines.values(): sp.set_color(C_DASH_BORD)
                ax.set_xlabel("Unidades", color=C_SID_HINT, fontsize=7)
                ax.xaxis.set_major_locator(mticker.MaxNLocator(5, integer=True))
                ax.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_units))
                ax.tick_params(axis="x", labelsize=7, colors=C_DASH_FG)
                card = tk.Frame(charts_row1, bg=BG, padx=6, pady=6)
                card.grid(row=0, column=0, sticky="nsew", padx=(0,6))
                _embed(fig, card)
            except: pass

        # Grafica 2 — Produccion semanal
        if not plan_df.empty and "Cantidad a Producir" in plan_df.columns and "Fecha" in plan_df.columns:
            try:
                pc = plan_df.copy()
                pc["Fecha"] = pd.to_datetime(pc["Fecha"], errors="coerce")
                pc["sem"] = pc["Fecha"].dt.to_period("W").astype(str)
                agg = pc.groupby("sem")["Cantidad a Producir"].sum().reset_index().sort_values("sem")
                fig, ax = plt.subplots(figsize=(_fw1, _fh), facecolor=FIG_BG,
                                       layout="constrained")
                ax.set_facecolor(FIG_BG)
                xs = range(len(agg))
                ax.fill_between(xs, agg["Cantidad a Producir"], alpha=0.25, color="#3B82F6")
                ax.plot(xs, agg["Cantidad a Producir"], color="#3B82F6", linewidth=2,
                        marker="o", markersize=4, markerfacecolor="white")
                ax.set_title("Produccion Semanal (Unidades)", color=C_DASH_FG,
                             fontsize=9, fontweight="bold")
                # Maximo 8 etiquetas en eje X
                n_ticks = min(8, len(agg))
                step = max(1, len(agg) // n_ticks)
                tick_pos = list(range(0, len(agg), step))
                tick_lbl = [str(agg["sem"].iloc[i])[-8:] for i in tick_pos]
                ax.set_xticks(tick_pos)
                ax.set_xticklabels(tick_lbl, rotation=40, ha="right",
                                   color=C_DASH_FG, fontsize=7)
                ax.tick_params(colors=C_DASH_FG, labelsize=7)
                for sp in ax.spines.values(): sp.set_color(C_DASH_BORD)
                ax.yaxis.set_major_locator(mticker.MaxNLocator(5, integer=True))
                ax.yaxis.set_major_formatter(mticker.FuncFormatter(_fmt_units))
                card = tk.Frame(charts_row1, bg=BG, padx=6, pady=6)
                card.grid(row=0, column=1, sticky="nsew", padx=(0,6))
                _embed(fig, card)
            except: pass

        # Grafica 3 — Donut Proyeccion vs Ocasional
        if not plan_df.empty and "Tipo" in plan_df.columns:
            try:
                n_p = int((plan_df["Tipo"]=="Proyeccion").sum())
                n_o = int((plan_df["Tipo"]=="Ocasional").sum())
                if n_p + n_o > 0:
                    # figsize cuadrado: el donut se mantiene circular y centrado al
                    # escalar; layout="constrained" evita recortes de titulo/leyenda.
                    fig, ax = plt.subplots(figsize=(_fwd, _fh), facecolor=FIG_BG,
                                           layout="constrained")
                    ax.set_facecolor(FIG_BG)
                    ax.set_aspect("equal")
                    wedges, _, autotexts = ax.pie([n_p, n_o], colors=["#3B82F6","#EA580C"],
                        autopct="%1.0f%%", startangle=90, radius=1.0,
                        wedgeprops=dict(width=0.55, edgecolor=FIG_BG, linewidth=2), pctdistance=0.75)
                    for at in autotexts:
                        at.set_color("white"); at.set_fontsize(9); at.set_fontweight("bold")
                    ax.set_title("Tipo de Orden", color=C_DASH_FG, fontsize=9, fontweight="bold", pad=8)
                    patches = [mpatches.Patch(color=c, label=l)
                               for c, l in zip(["#3B82F6","#EA580C"], ["Proyeccion","Ocasional"])]
                    ax.legend(handles=patches, loc="upper center", bbox_to_anchor=(0.5, -0.02),
                              fontsize=7, frameon=False, labelcolor=C_DASH_FG, ncol=2)
                    card = tk.Frame(charts_row1, bg=BG, padx=6, pady=6)
                    card.grid(row=0, column=2, sticky="nsew")
                    _embed(fig, card)
            except: pass

        # ── Fila 2 de graficas ────────────────────────────────────────────────
        charts_row2 = tk.Frame(self._dash_content, bg=C_DASH_BG)
        charts_row2.pack(fill="x", padx=PAD, pady=(10, 0))

        # Grafica 4 — Top 15 inventario (horizontal para mejor legibilidad)
        if not inv_df.empty and "Saldo" in inv_df.columns:
            try:
                inv_df["Saldo"] = pd.to_numeric(inv_df["Saldo"], errors="coerce").fillna(0)
                if "Referencia" in inv_df.columns and inv_df["Referencia"].notna().any():
                    ncol = "Referencia"
                elif "Cod. PT" in inv_df.columns:
                    ncol = "Cod. PT"
                else:
                    ncol = inv_df.columns[1]
                top_inv = inv_df.groupby(ncol)["Saldo"].sum().sort_values(ascending=False).head(15)
                fig, ax = plt.subplots(figsize=(_fw3, _fh), facecolor=FIG_BG,
                                       layout="constrained")
                ax.set_facecolor(FIG_BG)
                labels_inv = [str(n)[:24] for n in top_inv.index[::-1]]
                ax.barh(labels_inv, top_inv.values[::-1], color="#10B981", alpha=0.85, height=0.65)
                ax.set_title("Top 15 Inventario en Bodega (Unidades)", color=C_DASH_FG,
                             fontsize=9, fontweight="bold")
                ax.tick_params(colors=C_DASH_FG, labelsize=7)
                for sp in ax.spines.values(): sp.set_color(C_DASH_BORD)
                ax.xaxis.set_major_locator(mticker.MaxNLocator(5, integer=True))
                ax.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_units))
                card = tk.Frame(charts_row2, bg=BG, padx=6, pady=6)
                card.pack(side="left", fill="both", expand=True, padx=(0,6))
                _embed(fig, card)
            except: pass

        # Grafica 5 — OPs por tipo de trabajo
        if not ops_df.empty and "Tipo de Trabajo" in ops_df.columns:
            try:
                tipo_counts = ops_df["Tipo de Trabajo"].value_counts()
                fig, ax = plt.subplots(figsize=(_fw3, _fh), facecolor=FIG_BG)
                ax.set_facecolor(FIG_BG)
                colors_ops = ["#F59E0B","#EF4444","#8B5CF6","#06B6D4","#10B981"]
                ax.pie(tipo_counts.values, labels=None, colors=colors_ops[:len(tipo_counts)],
                       startangle=90, wedgeprops=dict(edgecolor=FIG_BG, linewidth=2))
                ax.set_title("OPs por Tipo de Trabajo", color=C_DASH_FG, fontsize=9, fontweight="bold", pad=8)
                patches = [mpatches.Patch(color=colors_ops[i], label=l)
                           for i, l in enumerate(tipo_counts.index[:5])]
                ax.legend(handles=patches, loc="lower center", fontsize=7,
                          frameon=False, labelcolor=C_DASH_FG, ncol=2, bbox_to_anchor=(0.5,-0.05))
                plt.tight_layout(pad=1.0)
                card = tk.Frame(charts_row2, bg=BG, padx=6, pady=6)
                card.pack(side="left", fill="both", expand=True, padx=(0,6))
                _embed(fig, card)
            except: pass

        # Grafica 6 — Pedidos Ocasionales: Neto a Producir top 10
        if not ocas_df.empty and "Neto a Producir" in ocas_df.columns:
            try:
                ocas_df["Neto a Producir"] = pd.to_numeric(ocas_df["Neto a Producir"], errors="coerce").fillna(0)
                top_oc = ocas_df[ocas_df["Neto a Producir"]>0].nlargest(10, "Neto a Producir")
                if not top_oc.empty:
                    ref_col = "Referencia" if "Referencia" in top_oc.columns else top_oc.columns[2]
                    fig, ax = plt.subplots(figsize=(_fw3, _fh), facecolor=FIG_BG,
                                           layout="constrained")
                    ax.set_facecolor(FIG_BG)
                    ax.barh([str(r)[:24] for r in top_oc[ref_col]][::-1],
                            top_oc["Neto a Producir"].values[::-1], color="#F59E0B", height=0.65)
                    ax.set_title("Top 10 Pedidos Ocasionales (Neto)", color=C_DASH_FG,
                                 fontsize=9, fontweight="bold")
                    ax.tick_params(colors=C_DASH_FG, labelsize=7)
                    for sp in ax.spines.values(): sp.set_color(C_DASH_BORD)
                    ax.xaxis.set_major_locator(mticker.MaxNLocator(5, integer=True))
                    ax.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_units))
                    card = tk.Frame(charts_row2, bg=BG, padx=6, pady=6)
                    card.pack(side="left", fill="both", expand=True)
                    _embed(fig, card)
            except: pass

        # ── Fila 3 de graficas ────────────────────────────────────────────────
        charts_row3 = tk.Frame(self._dash_content, bg=C_DASH_BG)
        charts_row3.pack(fill="x", padx=PAD, pady=(10, 0))

        # Grafica 7 — Tabla: Resumen mensual de produccion (cifras accionables)
        if not plan_df.empty and "Cantidad a Producir" in plan_df.columns and "Fecha" in plan_df.columns:
            try:
                cod_c = next((c for c in plan_df.columns
                              if "codigo" in str(c).lower() or "cód" in str(c).lower()), None)
                pc3 = plan_df.copy()
                pc3["Fecha"]   = pd.to_datetime(pc3["Fecha"], errors="coerce")
                pc3 = pc3.dropna(subset=["Fecha"])
                pc3["_cant"]   = pd.to_numeric(pc3["Cantidad a Producir"], errors="coerce").fillna(0)
                pc3["mes_key"] = pc3["Fecha"].dt.to_period("M").astype(str)
                pc3["mes_lbl"] = pc3["Fecha"].dt.strftime("%b %Y")
                _MES_ES = {"Jan": "Ene", "Apr": "Abr", "Aug": "Ago", "Dec": "Dic"}
                t_serie = (pc3["Tipo"].astype(str).str.lower()
                           if "Tipo" in pc3.columns else pd.Series("", index=pc3.index))

                filas = []
                for mk, g in pc3.groupby("mes_key"):
                    lbl = g["mes_lbl"].iloc[0]
                    for en, es in _MES_ES.items():
                        lbl = lbl.replace(en, es)
                    t     = t_serie.loc[g.index]
                    proy  = int(g.loc[t == "proyeccion", "_cant"].sum())
                    ocas  = int(g.loc[t == "ocasional",  "_cant"].sum())
                    total = int(g["_cant"].sum())
                    refs  = int(g[cod_c].nunique()) if cod_c else 0
                    ndias = int(g["Fecha"].dt.normalize().nunique())
                    uxd   = int(round(total / ndias)) if ndias else 0
                    filas.append((mk, lbl, proy, ocas, total, refs, ndias, uxd))
                filas.sort(key=lambda x: x[0])

                tot_tot  = sum(f[4] for f in filas)
                tot_dias = int(pc3["Fecha"].dt.normalize().nunique())
                fila_tot = ("TOTAL",
                            sum(f[2] for f in filas), sum(f[3] for f in filas), tot_tot,
                            int(pc3[cod_c].nunique()) if cod_c else 0, tot_dias,
                            int(round(tot_tot / tot_dias)) if tot_dias else 0)

                encabezado = ["Mes", "Proyeccion", "Ocasional", "Total", "Refs", "Dias", "Uds/dia"]
                cuerpo = [[f[1]] + [f"{v:,}" for v in f[2:]] for f in filas]
                cuerpo.append([fila_tot[0]] + [f"{v:,}" for v in fila_tot[1:]])

                fig, ax = plt.subplots(figsize=(_fw2, _fh), facecolor=FIG_BG,
                                       layout="constrained")
                ax.set_facecolor(FIG_BG); ax.axis("off")
                ax.set_title("Resumen Mensual de Produccion", color=C_DASH_FG,
                             fontsize=9, fontweight="bold", pad=6)
                tbl = ax.table(cellText=cuerpo, colLabels=encabezado,
                               cellLoc="center", loc="center")
                tbl.auto_set_font_size(False); tbl.set_fontsize(8)
                tbl.scale(1, 1.45)
                n_rows = len(cuerpo)
                for (r, c), cell in tbl.get_celld().items():
                    cell.set_edgecolor(C_DASH_BORD)
                    if r == 0:                       # encabezado
                        cell.set_facecolor("#1E3A58")
                        cell.set_text_props(color="#FFFFFF", fontweight="bold")
                    elif r == n_rows:                # fila TOTAL
                        cell.set_facecolor("#24405E")
                        cell.set_text_props(color="#FFFFFF", fontweight="bold")
                    else:
                        cell.set_facecolor("#162B45" if r % 2 else "#13243A")
                        if   c == 1: cell.set_text_props(color="#7EB8E0")  # Proyeccion
                        elif c == 2: cell.set_text_props(color="#F4B266")  # Ocasional
                        else:        cell.set_text_props(color=C_DASH_FG)
                    if c == 0:
                        cell.set_text_props(ha="left")
                        cell.PAD = 0.04
                card = tk.Frame(charts_row3, bg=BG, padx=6, pady=6)
                card.pack(side="left", fill="both", expand=True, padx=(0, 6))
                _embed(fig, card)
            except: pass

        # Grafica 8 — Top 10 inventario por costo
        if not inv_df.empty and "Costo Total" in inv_df.columns:
            try:
                ncol2 = "Referencia" if "Referencia" in inv_df.columns else inv_df.columns[2]
                inv2 = inv_df.copy()
                inv2["Costo Total"] = pd.to_numeric(inv2["Costo Total"], errors="coerce").fillna(0)
                top_costo = inv2.groupby(ncol2)["Costo Total"].sum().sort_values(ascending=False).head(10)
                if not top_costo.empty and top_costo.sum() > 0:
                    fig, ax = plt.subplots(figsize=(_fw2, _fh), facecolor=FIG_BG,
                                           layout="constrained")
                    ax.set_facecolor(FIG_BG)
                    ax.barh([str(r)[:24] for r in top_costo.index[::-1]],
                            top_costo.values[::-1] / 1e6, color="#8B5CF6", height=0.65, alpha=0.9)
                    ax.set_title("Top 10 Inventario por Costo ($M)", color=C_DASH_FG,
                                 fontsize=9, fontweight="bold")
                    ax.tick_params(colors=C_DASH_FG, labelsize=7)
                    for sp in ax.spines.values(): sp.set_color(C_DASH_BORD)
                    ax.xaxis.set_major_locator(mticker.MaxNLocator(5))
                    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.1f}M"))
                    card = tk.Frame(charts_row3, bg=BG, padx=6, pady=6)
                    card.pack(side="left", fill="both", expand=True)
                    _embed(fig, card)
            except: pass

        try:
            tk.Frame(self._dash_content, bg=C_DASH_BG, height=24).pack(fill="x")
            self._dash_canvas.yview_moveto(0)
        except Exception:
            pass
        self._dash_loaded = ruta


    # ═════════════════════════════════════════════════════════════════════════
    # TAB 3 — Plan Diario
    # ═════════════════════════════════════════════════════════════════════════

    def _build_tab_plan(self):
        topbar = tk.Frame(self.tab_plan, bg=C_HEADER, height=48)
        topbar.pack(fill="x"); topbar.pack_propagate(False)
        tk.Label(topbar, text="PLAN DIARIO DE PRODUCCION", font=("Segoe UI", 11, "bold"),
                 bg=C_HEADER, fg="white").pack(side="left", padx=24, pady=12)

        # Filtros
        fbar = tk.Frame(self.tab_plan, bg=C_MAIN)
        fbar.pack(fill="x", padx=16, pady=(10,0))
        tk.Label(fbar, text="Buscar:", font=("Segoe UI", 9),
                 bg=C_MAIN, fg=C_TEXTO).pack(side="left", padx=(0,8))
        self._plan_search = tk.StringVar()
        self._plan_search.trace_add("write", self._filtrar_plan)
        tk.Entry(fbar, textvariable=self._plan_search, font=("Segoe UI", 9), width=28,
                 bg="white", fg=C_TEXTO, relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=C_DIVIDER,
                 highlightcolor=C_ACENTO).pack(side="left", ipady=5, padx=(0,16))

        # Filtro tipo
        tk.Label(fbar, text="Tipo:", font=("Segoe UI", 9),
                 bg=C_MAIN, fg=C_TEXTO).pack(side="left", padx=(0,6))
        self._plan_tipo = tk.StringVar(value="Todos")
        tipo_cb = ttk.Combobox(fbar, textvariable=self._plan_tipo, width=14,
                                values=["Todos","Proyeccion","Ocasional"], state="readonly")
        tipo_cb.pack(side="left", ipady=3)
        tipo_cb.bind("<<ComboboxSelected>>", self._filtrar_plan)

        self._plan_lbl = tk.Label(fbar, text="Sin datos — carga un Excel primero",
                                   font=("Segoe UI", 8), bg=C_MAIN, fg=C_GRIS)
        self._plan_lbl.pack(side="left", padx=16)

        # Leyenda tipos
        ley = tk.Frame(self.tab_plan, bg=C_MAIN)
        ley.pack(fill="x", padx=16, pady=(4,0))
        for label, color in [("Proyeccion","#DBEAFE"), ("Ocasional","#FEF3C7")]:
            f = tk.Frame(ley, bg=color, padx=8, pady=3)
            f.pack(side="left", padx=(0,8))
            tk.Label(f, text=label, font=("Segoe UI", 8, "bold"),
                     bg=color, fg=C_TEXTO).pack()

        # Tabla
        tbl = tk.Frame(self.tab_plan, bg=C_MAIN)
        tbl.pack(fill="both", expand=True, padx=16, pady=10)
        sty = ttk.Style()
        sty.configure("Plan.Treeview", background="white", foreground=C_TEXTO,
                       fieldbackground="white", rowheight=24, font=("Segoe UI", 9))
        sty.configure("Plan.Treeview.Heading", background=C_HEADER, foreground="white",
                       font=("Segoe UI", 9, "bold"))
        sty.map("Plan.Treeview", background=[("selected", C_ACENTO)],
                foreground=[("selected","white")])
        COLS = ["Fecha","Codigo PT","Referencia","Cantidad a Producir","Tipo","Fecha Entrega"]
        self._plan_tree = ttk.Treeview(tbl, columns=COLS, show="headings", style="Plan.Treeview")
        WIDTHS = {"Fecha":100,"Codigo PT":130,"Referencia":330,"Cantidad a Producir":130,"Tipo":110,"Fecha Entrega":120}
        CENTERS = {"Fecha","Codigo PT","Cantidad a Producir","Tipo","Fecha Entrega"}
        for col in COLS:
            self._plan_tree.heading(col, text=col, command=lambda c=col: self._sort_plan(c))
            self._plan_tree.column(col, width=WIDTHS.get(col,100),
                                   anchor="center" if col in CENTERS else "w")
        vsb = ttk.Scrollbar(tbl, orient="vertical",   command=self._plan_tree.yview)
        hsb = ttk.Scrollbar(tbl, orient="horizontal", command=self._plan_tree.xview)
        self._plan_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._plan_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tbl.rowconfigure(0, weight=1); tbl.columnconfigure(0, weight=1)

    def _cargar_plan_diario(self):
        ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
        if self._last_excel and os.path.isfile(self._last_excel):
            ruta = self._last_excel
        elif os.path.isfile(ruta_auto):
            ruta = ruta_auto
        else:
            ruta = filedialog.askopenfilename(title="Seleccionar Excel de resultado",
                                              filetypes=[("Excel","*.xlsx *.xls")])
        if not ruta: return
        try:
            import pandas as pd
            df = pd.read_excel(ruta, sheet_name="Plan_Diario", header=1)
        except Exception as exc:
            messagebox.showerror("Error", str(exc)); return
        self._plan_data = []
        cols_use = ["Fecha","Codigo PT","Referencia","Cantidad a Producir","Tipo","Fecha Entrega"]
        for _, row in df.iterrows():
            vals = []
            for c in cols_use:
                v = row.get(c, "")
                if hasattr(v, "date"): v = str(v.date())
                elif v != v: v = ""
                else: v = str(v)
                vals.append(v)
            self._plan_data.append(vals)
        self._refrescar_plan_tree(self._plan_data)
        date_strs = [r[0] for r in self._plan_data if r[0] and r[0] not in ("", "nan", "NaT", "None")]
        date_range = f"  •  {min(date_strs)} → {max(date_strs)}" if date_strs else ""
        self._plan_lbl.config(text=f"{len(self._plan_data):,} filas{date_range}  •  {os.path.basename(ruta)}")

    def _refrescar_plan_tree(self, data):
        self._plan_tree.delete(*self._plan_tree.get_children())
        for i, row in enumerate(data):
            tipo = row[4] if len(row) > 4 else ""
            tag = "ocasional" if "casional" in tipo else ("proy" if i%2==0 else "proy2")
            self._plan_tree.insert("", "end", values=row, tags=(tag,))
        self._plan_tree.tag_configure("ocasional", background="#FEF3C7")
        self._plan_tree.tag_configure("proy",  background="white")
        self._plan_tree.tag_configure("proy2", background="#EFF6FF")

    def _filtrar_plan(self, *_):
        q    = self._plan_search.get().lower()
        tipo = self._plan_tipo.get()
        def _match(r):
            if tipo != "Todos" and len(r) > 4 and r[4] != tipo:
                return False
            if q and not any(q in str(v).lower() for v in r):
                return False
            return True
        filtered = [r for r in self._plan_data if _match(r)]
        self._refrescar_plan_tree(filtered)
        self._plan_lbl.config(
            text=f"{len(filtered):,} filas" +
                 (f"  (de {len(self._plan_data):,})" if len(filtered) != len(self._plan_data) else ""))

    def _sort_plan(self, col):
        COLS = ["Fecha","Codigo PT","Referencia","Cantidad a Producir","Tipo","Fecha Entrega"]
        idx  = COLS.index(col)
        rev  = (self._plan_sort_col == col) and not self._plan_sort_rev
        self._plan_sort_col = col; self._plan_sort_rev = rev
        try: self._plan_data.sort(key=lambda r: r[idx], reverse=rev)
        except Exception: pass
        self._refrescar_plan_tree(self._plan_data)

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 4 — Tablas del Excel (sub-tabs por hoja)
    # ═════════════════════════════════════════════════════════════════════════

    def _build_tab_tablas(self):
        topbar = tk.Frame(self.tab_tablas, bg=C_HEADER, height=48)
        topbar.pack(fill="x"); topbar.pack_propagate(False)
        tk.Label(topbar, text="TABLAS DEL EXCEL", font=("Segoe UI", 11, "bold"),
                 bg=C_HEADER, fg="white").pack(side="left", padx=24, pady=12)
        self._make_btn(topbar, "  Exportar Excel  ", C_SUCCESS, C_SUCCESS_H,
                       self._exportar_excel).pack(side="right", pady=8, ipady=4)
        self._make_btn(topbar, "  Exportar PDF  ", C_PDF_BTN, C_PDF_BTN_H,
                       self._exportar_pdf).pack(side="right", padx=8, pady=8, ipady=4)

        self._tablas_container = tk.Frame(self.tab_tablas, bg=C_MAIN)
        self._tablas_container.pack(fill="both", expand=True)

        self._tablas_placeholder()

    def _tablas_placeholder(self):
        for w in self._tablas_container.winfo_children():
            w.destroy()
        wrapper = tk.Frame(self._tablas_container, bg=C_MAIN)
        wrapper.pack(expand=True, fill="both", pady=120)
        tk.Label(wrapper, text="Sin datos", font=("Segoe UI", 22, "bold"),
                 bg=C_MAIN, fg=C_GRIS).pack()
        tk.Label(wrapper, text="Carga un Excel de resultado para ver las tablas.",
                 font=("Segoe UI", 10), bg=C_MAIN, fg=C_GRIS, justify="center").pack(pady=(10,0))

    def _render_kpi_dashboard_tab(self, parent, xls, pd):
        BG = C_MAIN

        try:
            def _hoja(nombre, hdr=1):
                try:
                    return pd.read_excel(xls, nombre, header=hdr) if nombre in xls.sheet_names else pd.DataFrame()
                except: return pd.DataFrame()

            plan_df = _hoja("Plan_Diario")
            ops_df  = _hoja("OPs_En_Proceso")
            inv_df  = _hoja("Inventario_Actual")
            gest_df = _hoja("Gestion_Pedidos", hdr=0)

            # KPIs del plan — calculados directamente desde Plan_Diario
            if not plan_df.empty:
                fecha_col = next((c for c in plan_df.columns if "fecha" in str(c).lower() and "entrega" not in str(c).lower()), None)
                cod_col   = next((c for c in plan_df.columns if "codigo" in str(c).lower() or "cód" in str(c).lower()), None)
                cant_col  = next((c for c in plan_df.columns if "cantidad" in str(c).lower()), None)
                tipo_col  = next((c for c in plan_df.columns if str(c).lower().strip() == "tipo"), None)

                dias  = int(plan_df[fecha_col].nunique())         if fecha_col else 0
                refs  = int(plan_df[cod_col].nunique())           if cod_col   else 0
                unds  = int(pd.to_numeric(plan_df[cant_col], errors="coerce").fillna(0).sum()) if cant_col else 0

                if tipo_col:
                    proy_mask = plan_df[tipo_col].astype(str).str.lower() == "proyeccion"
                    ocas_mask = plan_df[tipo_col].astype(str).str.lower() == "ocasional"
                    r_proy = int(plan_df.loc[proy_mask, cod_col].nunique()) if cod_col else 0
                    r_ocas = int(plan_df.loc[ocas_mask, cod_col].nunique()) if cod_col else 0
                    u_ocas = int(pd.to_numeric(plan_df.loc[ocas_mask, cant_col], errors="coerce").fillna(0).sum()) if cant_col else 0
                else:
                    r_proy = refs; r_ocas = 0; u_ocas = 0
            else:
                dias = refs = unds = r_proy = r_ocas = u_ocas = 0

            if not gest_df.empty:
                est_col = next((c for c in gest_df.columns if "estado" in str(c).lower()), None)
                stock_cub = int((gest_df[est_col].astype(str).str.strip() == "OK").sum()) if est_col else 0
            else:
                stock_cub = 0

            n_ops     = len(ops_df) if not ops_df.empty else 0
            inv_saldo = int(pd.to_numeric(inv_df["Saldo"], errors="coerce").fillna(0).sum()) if (not inv_df.empty and "Saldo" in inv_df.columns) else 0
            inv_costo = float(pd.to_numeric(inv_df["Costo Total"], errors="coerce").fillna(0).sum()) if (not inv_df.empty and "Costo Total" in inv_df.columns) else 0.0

        except Exception as exc:
            tk.Label(parent, text=f"Error cargando Dashboard KPIs:\n{exc}",
                     font=("Segoe UI", 9), bg=BG, fg="#DC2626",
                     justify="left", wraplength=600).pack(padx=20, pady=20, anchor="nw")
            return

        kpis = [
            ("Dias planificados",   f"{dias:,}",              "#2563EB"),
            ("Referencias en plan", f"{refs:,}",              "#7C3AED"),
            ("Unidades a producir", f"{unds:,}",              "#0891B2"),
            ("Refs. proyeccion",    f"{r_proy:,}",            "#16A34A"),
            ("Refs. ocasionales",   f"{r_ocas:,}",            "#EA580C"),
            ("Uds. ocasionales",    f"{u_ocas:,}",            "#DC2626"),
            ("Cubiertos por stock", f"{stock_cub:,}",         "#CA8A04"),
            ("OPs en proceso",      f"{n_ops:,}",             "#0891B2"),
            ("Inventario (uds)",    f"{inv_saldo:,}",         "#16A34A"),
            ("Costo inventario",    f"${inv_costo/1e6:.1f}M", "#7C3AED"),
        ]

        body = tk.Frame(parent, bg=BG)
        body.pack(fill="both", expand=True)

        # Titulo
        title_f = tk.Frame(body, bg=BG)
        title_f.pack(fill="x", padx=24, pady=(20, 0))
        tk.Label(title_f, text="RESUMEN DEL PLAN DE PRODUCCION",
                 font=("Segoe UI", 11, "bold"), bg=BG, fg=C_TEXTO).pack(side="left")
        tk.Frame(body, bg=C_DIVIDER, height=1).pack(fill="x", padx=24, pady=(10, 18))

        # 10 KPIs en 2 filas de 5
        for fila in range(2):
            row_f = tk.Frame(body, bg=BG)
            row_f.pack(fill="x", padx=24, pady=(0, 12))
            for i, (lbl, val, col) in enumerate(kpis[fila*5:(fila+1)*5]):
                card = tk.Frame(row_f, bg="white",
                                highlightbackground="#E2E8F0", highlightthickness=1)
                card.pack(side="left", fill="both", expand=True,
                          padx=(0 if i == 0 else 10, 0))
                tk.Frame(card, bg=col, height=4).pack(fill="x")
                cb = tk.Frame(card, bg="white")
                cb.pack(fill="both", expand=True, padx=16, pady=(12, 14))
                tk.Label(cb, text=val, font=("Segoe UI", 22, "bold"),
                         bg="white", fg=col, anchor="w").pack(anchor="w")
                tk.Label(cb, text=lbl, font=("Segoe UI", 8),
                         bg="white", fg=C_GRIS, anchor="w").pack(anchor="w", pady=(4, 0))

        # Desglose por tipo
        tk.Frame(body, bg=C_DIVIDER, height=1).pack(fill="x", padx=24, pady=(6, 16))
        tk.Label(body, text="DESGLOSE POR TIPO DE ORDEN",
                 font=("Segoe UI", 9, "bold"), bg=BG, fg=C_TEXTO).pack(anchor="w", padx=24, pady=(0, 10))

        desglose_row = tk.Frame(body, bg=BG)
        desglose_row.pack(fill="x", padx=24, pady=(0, 28))

        for titulo, acento, metricas in [
            ("Plan de Proyeccion",  "#3B82F6",
             [("Referencias",  f"{r_proy:,}"),
              ("Unidades",     f"{max(0, unds - u_ocas):,}")]),
            ("Pedidos Ocasionales", "#F59E0B",
             [("Referencias",  f"{r_ocas:,}"),
              ("Unidades",     f"{u_ocas:,}")]),
            ("Stock y Cobertura",   "#10B981",
             [("Refs. cubiertas",   f"{stock_cub:,}"),
              ("Inventario (uds)",  f"{inv_saldo:,}")]),
            ("Ordenes en Proceso",  "#8B5CF6",
             [("OPs activas",  f"{n_ops:,}"),
              ("Costo inv.",   f"${inv_costo/1e6:.1f}M")]),
        ]:
            sc = tk.Frame(desglose_row, bg="white",
                          highlightbackground="#E2E8F0", highlightthickness=1)
            sc.pack(side="left", fill="both", expand=True, padx=(0, 10))
            tk.Frame(sc, bg=acento, height=4).pack(fill="x")
            sb2 = tk.Frame(sc, bg="white")
            sb2.pack(fill="both", expand=True, padx=14, pady=(10, 12))
            tk.Label(sb2, text=titulo, font=("Segoe UI", 8, "bold"),
                     bg="white", fg=C_TEXTO).pack(anchor="w")
            tk.Frame(sb2, bg="#F1F5F9", height=1).pack(fill="x", pady=(6, 4))
            for m_lbl, m_val in metricas:
                mrow = tk.Frame(sb2, bg="white")
                mrow.pack(fill="x", pady=1)
                tk.Label(mrow, text=m_lbl, font=("Segoe UI", 8),
                         bg="white", fg=C_GRIS).pack(side="left")
                tk.Label(mrow, text=m_val, font=("Segoe UI", 8, "bold"),
                         bg="white", fg=acento).pack(side="right")

    def _cargar_tablas_excel(self, ruta=None):
        if ruta is None:
            ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
            ruta = (self._last_excel if self._last_excel and os.path.isfile(self._last_excel)
                    else ruta_auto if os.path.isfile(ruta_auto) else None)
        if ruta is None:
            ruta = filedialog.askopenfilename(title="Seleccionar Excel",
                                              filetypes=[("Excel","*.xlsx *.xls")])
        if not ruta or not os.path.isfile(ruta):
            return
        if self._tablas_loaded == ruta:
            return
        try:
            import pandas as pd
        except ImportError as exc:
            messagebox.showerror("Falta pandas", str(exc)); return
        try:
            xls = pd.ExcelFile(ruta)
        except Exception as exc:
            messagebox.showerror("Error leyendo Excel", str(exc)); return

        for w in self._tablas_container.winfo_children():
            w.destroy()

        SHEET_ICONS = {
            "Plan_Diario":        "Plan Diario",
            "Plan_Mensual":       "Plan Mensual",
            "Calendario_Semanal": "Calendario Semanal",
            "OPs_En_Proceso":     "OPs en Proceso",
            "Pedidos_Ocasionales":"Pedidos Ocasionales",
            "Inventario_Actual":  "Inventario Consolidado",
            "Inventario_Detalle": "Inventario por Bodega",
            "Gestion_Pedidos":    "Gestion Pedidos",
            "Saldo_OPs":          "Saldo OPs",
            "Pedidos_Sin_Cubrir": "Pedidos Sin Cubrir",
            "Comparativa_Comercial": "Comparativa Comercial",
            "Dashboard":          "Dashboard KPIs",
        }
        hojas_mostrar = [s for s in SHEET_ICONS if s in xls.sheet_names]

        # Custom tab bar profesional (underline style)
        _tabs_content   = {}
        _tab_btns       = {}
        _tab_indicators = {}
        _cur_tab        = [None]

        tab_bar = tk.Frame(self._tablas_container, bg=C_CARD, padx=8, pady=0)
        tab_bar.pack(fill="x")
        tk.Frame(self._tablas_container, bg=C_DIVIDER, height=1).pack(fill="x")
        content_area = tk.Frame(self._tablas_container, bg=C_MAIN)
        content_area.pack(fill="both", expand=True)

        def _select_tab(name):
            _cur_tab[0] = name
            for n, f in _tabs_content.items():
                f.pack_forget()
            _tabs_content[name].pack(fill="both", expand=True)
            for n, btn in _tab_btns.items():
                sel = (n == name)
                btn.configure(
                    fg=C_ACENTO if sel else C_GRIS,
                    font=("Segoe UI", 9, "bold") if sel else ("Segoe UI", 9),
                    bg=C_CARD, activeforeground=C_ACENTO, activebackground="#EFF6FF",
                )
                _tab_indicators[n].configure(bg=C_ACENTO if sel else C_CARD)

        for sheet in hojas_mostrar:
            cfg = SHEETS_CONFIG.get(sheet, {"header": 1})
            try:
                df = pd.read_excel(xls, sheet, header=cfg.get("header", 1))
                df = df.dropna(how="all")
            except:
                df = pd.DataFrame()

            label = SHEET_ICONS.get(sheet, sheet)
            tab_col = tk.Frame(tab_bar, bg=C_CARD)
            tab_col.pack(side="left", fill="y")
            btn = tk.Button(tab_col, text=label,
                            font=("Segoe UI", 9), bg=C_CARD, fg=C_GRIS,
                            activebackground="#EFF6FF", activeforeground=C_ACENTO,
                            bd=0, relief="flat", padx=14, pady=10,
                            cursor="hand2", command=lambda s=sheet: _select_tab(s))
            btn.pack(side="top", fill="x")
            indicator = tk.Frame(tab_col, bg=C_CARD, height=3)
            indicator.pack(side="bottom", fill="x")
            btn.bind("<Enter>", lambda e, b=btn, s=sheet: b.configure(
                bg="#EFF6FF", fg=C_ACENTO) if _cur_tab[0] != s else b.configure(bg="#EFF6FF"))
            btn.bind("<Leave>", lambda e, b=btn, s=sheet: b.configure(
                bg=C_CARD, fg=C_ACENTO if _cur_tab[0] == s else C_GRIS))
            _tab_btns[sheet]       = btn
            _tab_indicators[sheet] = indicator

            frame = tk.Frame(content_area, bg=C_MAIN)
            _tabs_content[sheet] = frame

            if sheet == "Dashboard":
                self._render_kpi_dashboard_tab(frame, xls, pd)
                continue

            # Barra de busqueda por hoja
            sbar = tk.Frame(frame, bg=C_MAIN)
            sbar.pack(fill="x", padx=12, pady=(8,4))
            tk.Label(sbar, text="Buscar:", font=("Segoe UI", 9),
                     bg=C_MAIN, fg=C_TEXTO).pack(side="left", padx=(0,6))
            search_var = tk.StringVar()
            tk.Entry(sbar, textvariable=search_var, font=("Segoe UI", 9), width=30,
                     bg="white", fg=C_TEXTO, relief="flat", bd=0,
                     highlightthickness=1, highlightbackground=C_DIVIDER,
                     highlightcolor=C_ACENTO).pack(side="left", ipady=4)
            n_lbl = tk.Label(sbar, text=f"{len(df):,} registros", font=("Segoe UI", 8),
                             bg=C_MAIN, fg=C_GRIS)
            n_lbl.pack(side="left", padx=12)

            # Treeview
            tbl_fr = tk.Frame(frame, bg=C_MAIN)
            tbl_fr.pack(fill="both", expand=True, padx=12, pady=(0,12))

            cols = [str(c) for c in df.columns if not str(c).startswith("Unnamed")]
            if not cols:
                cols = [str(c) for c in df.columns]
            tree = ttk.Treeview(tbl_fr, columns=cols, show="headings",
                                 height=20, style="Plan.Treeview")
            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, width=max(80, min(200, len(str(col))*11+20)), anchor="w")
            vsb2 = ttk.Scrollbar(tbl_fr, orient="vertical",   command=tree.yview)
            hsb2 = ttk.Scrollbar(tbl_fr, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
            tree.grid(row=0, column=0, sticky="nsew")
            vsb2.grid(row=0, column=1, sticky="ns")
            hsb2.grid(row=1, column=0, sticky="ew")
            tbl_fr.rowconfigure(0, weight=1); tbl_fr.columnconfigure(0, weight=1)

            # Cargar datos con colores alternos
            # _cols capturado por valor para evitar el bug de closure en for-loop
            def _fill_tree(tr, dataframe, search="", _cols=cols):
                tr.delete(*tr.get_children())
                q = search.lower()
                for i, (_, row) in enumerate(dataframe.iterrows()):
                    vals = []
                    for c in _cols:
                        v = row.get(c, "")
                        if hasattr(v, "date"): v = str(v.date())
                        elif v != v: v = ""
                        else: v = str(v)
                        vals.append(v)
                    if q and not any(q in str(x).lower() for x in vals):
                        continue
                    tag = "alt" if i%2 else "norm"
                    tr.insert("", "end", values=vals, tags=(tag,))
                tr.tag_configure("alt",  background="#EFF6FF")
                tr.tag_configure("norm", background="white")

            _fill_tree(tree, df)

            # ft=_fill_tree captura la funcion correcta de esta iteracion del loop
            def _on_search(*_, sv=search_var, tr=tree, df_=df, nl=n_lbl, ft=_fill_tree):
                ft(tr, df_, sv.get())
                nl.config(text=f"{len(tr.get_children()):,} registros")
            search_var.trace_add("write", _on_search)

        if hojas_mostrar:
            _select_tab(hojas_mostrar[0])

        self._tablas_loaded = ruta


    # ═════════════════════════════════════════════════════════════════════════
    # MODAL MAPEO DE COLUMNAS
    # ═════════════════════════════════════════════════════════════════════════

    def _cargar_mapeos_columnas(self):
        if os.path.isfile(_MAPEOS_PATH):
            try:
                with open(_MAPEOS_PATH, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _guardar_mapeos_columnas(self, mapeos):
        try:
            with open(_MAPEOS_PATH, "w", encoding="utf-8") as f:
                json.dump(mapeos, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))

    def _abrir_modal_columnas(self):
        """Modal de configuracion de mapeo de columnas: nombre en Excel -> nombre interno."""
        import pandas as _pd

        modal = tk.Toplevel(self.root)
        modal.title("Configuracion de Columnas")
        modal.configure(bg=C_MAIN)
        modal.resizable(True, True)
        modal.minsize(840, 480)
        modal.transient(self.root)
        modal.grab_set()
        self.root.update_idletasks()
        mw, mh = 980, 700
        x = self.root.winfo_rootx() + (self.root.winfo_width()  - mw) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - mh) // 2
        modal.geometry(f"{mw}x{mh}+{max(0, x)}+{max(0, y)}")

        # ── Header ────────────────────────────────────────────────────────────
        hdr_f = tk.Frame(modal, bg=C_HEADER)
        hdr_f.pack(fill="x")
        tk.Label(hdr_f, text="  Configuracion de Columnas",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white").pack(side="left", padx=20, pady=12)
        tk.Label(hdr_f,
                 text="Si INEDITTO cambia el nombre de una columna, ajusta el mapeo aqui para que el pipeline siga funcionando.",
                 font=("Segoe UI", 8), bg=C_HEADER, fg="#7EB8E0").pack(side="left", pady=12)

        # ── Leer columnas reales de cada archivo cargado ───────────────────────
        cols_por_archivo = {}
        for cfg in ARCHIVOS:
            key  = cfg["key"]
            ruta = self.file_vars[key].get().strip()
            cols = []
            if ruta and os.path.isfile(ruta):
                try:
                    xls = _pd.ExcelFile(ruta)
                    for sheet in xls.sheet_names[:4]:
                        for hdr_row in [0, 1, 2, 3]:
                            try:
                                tmp = _pd.read_excel(ruta, sheet_name=sheet,
                                                      header=hdr_row, nrows=1)
                                real = [str(c).strip() for c in tmp.columns
                                        if "Unnamed" not in str(c) and str(c).strip()]
                                if len(real) >= 3:
                                    cols = real; break
                            except Exception:
                                pass
                        if cols:
                            break
                except Exception:
                    pass
            cols_por_archivo[key] = cols

        mapeos = self._cargar_mapeos_columnas()

        # ── Tab bar ────────────────────────────────────────────────────────────
        keys_schema  = [k for k in _COL_SCHEMA if k in {c["key"] for c in ARCHIVOS}]
        _tab_btns    = {}
        _tab_inds    = {}
        _tab_frames  = {}
        _cur         = [None]

        tab_bar = tk.Frame(modal, bg=C_CARD, padx=8)
        tab_bar.pack(fill="x")
        tk.Frame(modal, bg=C_DIVIDER, height=1).pack(fill="x")
        content_wrap = tk.Frame(modal, bg=C_MAIN)
        content_wrap.pack(fill="both", expand=True)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = tk.Frame(modal, bg=C_CARD, pady=10)
        footer.pack(fill="x", side="bottom")
        tk.Frame(modal, bg=C_DIVIDER, height=1).pack(fill="x", side="bottom")

        all_vars = {}   # {key: {campo_interno: StringVar}}

        def _sel_tab(key):
            _cur[0] = key
            for k, f in _tab_frames.items():
                f.pack_forget()
            _tab_frames[key].pack(fill="both", expand=True)
            for k, b in _tab_btns.items():
                sel = (k == key)
                b.configure(fg=C_ACENTO if sel else C_GRIS,
                             font=("Segoe UI", 9, "bold") if sel else ("Segoe UI", 9),
                             bg=C_CARD)
                _tab_inds[k].configure(bg=C_ACENTO if sel else C_CARD)

        for key in keys_schema:
            schema    = _COL_SCHEMA[key]
            file_cols = cols_por_archivo.get(key, [])
            mapeo_key = mapeos.get(key, {})
            loaded    = bool(file_cols)

            # Boton de tab
            col_wrap = tk.Frame(tab_bar, bg=C_CARD)
            col_wrap.pack(side="left", fill="y")
            dot = "●" if loaded else "○"
            dot_color = C_SUCCESS if loaded else C_GRIS
            btn_lbl = f" {schema['titulo']} "
            btn = tk.Button(col_wrap, text=btn_lbl,
                             font=("Segoe UI", 9), bg=C_CARD, fg=C_GRIS,
                             relief="flat", padx=8, cursor="hand2", bd=0,
                             activeforeground=C_ACENTO, activebackground="#EFF6FF",
                             command=lambda k=key: _sel_tab(k))
            btn.pack(side="top", pady=(6, 0))
            # Indicador de archivo cargado
            dot_lbl = tk.Label(col_wrap, text=dot, font=("Segoe UI", 7),
                                bg=C_CARD, fg=dot_color)
            dot_lbl.pack(side="top")
            ind = tk.Frame(col_wrap, bg=C_CARD, height=3)
            ind.pack(fill="x")
            _tab_btns[key] = btn
            _tab_inds[key] = ind

            # Frame del tab
            tf = tk.Frame(content_wrap, bg=C_MAIN)
            _tab_frames[key] = tf
            all_vars[key] = {}

            # Barra de estado del archivo
            ruta  = self.file_vars[key].get().strip()
            fname = os.path.basename(ruta) if ruta else "No cargado"
            sf    = tk.Frame(tf, bg=C_CARD, padx=20, pady=8)
            sf.pack(fill="x")
            tk.Label(sf, text=f"Archivo:  {fname}",
                      font=("Segoe UI", 9, "bold"), bg=C_CARD, fg=C_TEXTO).pack(side="left")
            estado_txt = ("  Cargado — columnas detectadas correctamente ✓" if loaded
                           else "  Sin archivo cargado — escribe el nombre exacto de la columna")
            tk.Label(sf, text=estado_txt, font=("Segoe UI", 8), bg=C_CARD,
                      fg=C_SUCCESS if loaded else C_GRIS).pack(side="left", padx=6)

            # Encabezados de la tabla
            th = tk.Frame(tf, bg=C_HEADER)
            th.pack(fill="x", padx=20, pady=(10, 0))
            tk.Label(th, text="     Campo que espera el sistema   (* = requerido)",
                      font=("Segoe UI", 9, "bold"), bg=C_HEADER, fg="white",
                      anchor="w").pack(side="left", pady=8, padx=(0, 0))
            tk.Label(th, text="Columna en tu archivo Excel",
                      font=("Segoe UI", 9, "bold"), bg=C_HEADER, fg="white",
                      anchor="w").pack(side="right", pady=8, padx=(0, 24))

            # Canvas con scroll para las filas
            canvas_wrap = tk.Frame(tf, bg="white")
            canvas_wrap.pack(fill="both", expand=True, padx=20, pady=(0, 8))
            canvas  = tk.Canvas(canvas_wrap, bg="white", highlightthickness=0)
            vsb     = ttk.Scrollbar(canvas_wrap, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            rows_f = tk.Frame(canvas, bg="white")
            _win   = canvas.create_window((0, 0), window=rows_f, anchor="nw")
            rows_f.bind("<Configure>", lambda e, c=canvas: c.configure(
                scrollregion=c.bbox("all")))
            canvas.bind("<Configure>", lambda e, c=canvas, w=_win: c.itemconfig(
                w, width=e.width))
            canvas.bind("<Enter>", lambda e, c=canvas: c.bind_all(
                "<MouseWheel>", lambda ev, cv=c: cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
            canvas.bind("<Leave>", lambda e, c=canvas: c.unbind_all("<MouseWheel>"))

            opciones = ["(Sin cambio)"] + file_cols

            for i, (campo, descr, req) in enumerate(schema["campos"]):
                row = tk.Frame(rows_f, bg="white")
                row.pack(fill="x")

                # Barra de acento lateral (azul=requerido, gris=opcional)
                tk.Frame(row, bg=C_ACENTO if req else "#CBD5E1",
                          width=4).pack(side="left", fill="y")

                # Columna izquierda — espaciador fuerza ancho minimo 300px
                lf = tk.Frame(row, bg="white")
                lf.pack(side="left", fill="y")
                tk.Frame(lf, bg="white", width=300, height=1).pack()
                lpad = tk.Frame(lf, bg="white")
                lpad.pack(anchor="w", padx=18, pady=16)
                mark = " *" if req else ""
                tk.Label(lpad, text=campo + mark,
                          font=("Segoe UI", 10, "bold"), bg="white",
                          fg=C_TEXTO if req else C_GRIS,
                          anchor="w", wraplength=275).pack(anchor="w")
                tk.Label(lpad, text=descr,
                          font=("Segoe UI", 9), bg="white", fg="#64748B",
                          anchor="w", wraplength=275,
                          justify="left").pack(anchor="w", pady=(5, 0))

                # Separador vertical
                tk.Frame(row, bg="#F1F5F9", width=1).pack(
                    side="left", fill="y", pady=10)

                # Columna derecha — selector ocupa todo el ancho restante
                rf = tk.Frame(row, bg="white")
                rf.pack(side="left", fill="both", expand=True, padx=20, pady=16)

                var = tk.StringVar()
                all_vars[key][campo] = var
                val_guardado = mapeo_key.get(campo, "")
                var.set(val_guardado if val_guardado else "(Sin cambio)")

                if loaded:
                    cb = ttk.Combobox(rf, textvariable=var, values=opciones,
                                       font=("Segoe UI", 10), state="readonly")
                    cb.pack(fill="x", ipady=5)
                    # Bloquear scroll sobre el combobox para que no cambie su valor accidentalmente
                    cb.bind("<MouseWheel>", lambda e: "break")
                    if val_guardado and val_guardado != campo and val_guardado in file_cols:
                        cb.configure(foreground=C_ACENTO)
                    elif val_guardado and val_guardado not in file_cols:
                        cb.configure(foreground="#DC2626")
                else:
                    ent = tk.Entry(rf, textvariable=var, font=("Segoe UI", 10),
                                    bg="#F8FAFC", fg=C_TEXTO, relief="flat", bd=0,
                                    highlightthickness=1, highlightbackground=C_DIVIDER,
                                    highlightcolor=C_ACENTO)
                    ent.pack(fill="x", ipady=6)
                    tk.Label(rf, text="Escribe el nombre exacto como aparece en el Excel",
                              font=("Segoe UI", 7), bg="white", fg=C_GRIS
                              ).pack(anchor="w", pady=(4, 0))

                # Separador horizontal entre filas
                tk.Frame(rows_f, bg="#E8EDF3", height=1).pack(fill="x")

            tk.Label(tf, text="   * Campos requeridos — un mapeo incorrecto puede causar errores en el pipeline",
                      font=("Segoe UI", 8, "italic"), bg="white", fg="#94A3B8").pack(
                anchor="w", padx=20, pady=(8, 6))

        # ── Acciones del footer ────────────────────────────────────────────────
        def _guardar():
            nuevo = {}
            for k, campos_vars in all_vars.items():
                m = {campo: var.get().strip()
                     for campo, var in campos_vars.items()
                     if var.get().strip() not in ("", "(Sin cambio)")}
                if m:
                    nuevo[k] = m
            self._guardar_mapeos_columnas(nuevo)
            messagebox.showinfo("Guardado",
                                 "Mapeo de columnas guardado correctamente.\n\n"
                                 "Se aplicara en el proximo Generar Plan de Produccion.")
            modal.destroy()

        def _limpiar():
            if messagebox.askyesno("Limpiar mapeos",
                                    "Se eliminaran todos los mapeos personalizados\n"
                                    "y se usaran los nombres originales de siempre.\n\n"
                                    "Continuar?"):
                self._guardar_mapeos_columnas({})
                modal.destroy()
                self._abrir_modal_columnas()

        self._make_btn(footer, "  Guardar cambios  ", C_ACENTO, C_ACENTO_H, _guardar).pack(side="right", ipady=6, padx=(8, 20))
        self._make_btn(footer, "  Limpiar todo  ", "#DC2626", "#B91C1C", _limpiar).pack(side="right", ipady=6)
        self._make_btn(footer, "  Cancelar  ", C_CANCEL, C_CANCEL_H, modal.destroy).pack(side="right", ipady=6, padx=(0, 8))

        # Seleccionar primer tab
        if keys_schema:
            _sel_tab(keys_schema[0])

    # ═════════════════════════════════════════════════════════════════════════
    # VALIDACIÓN AUTOMÁTICA DE COLUMNAS AL CARGAR ARCHIVO
    # ═════════════════════════════════════════════════════════════════════════

    def _leer_columnas_archivo(self, ruta, clave=None):
        """
        Lee las columnas reales de un Excel probando todas las hojas y headers.
        Cuando se proporciona clave, puntua cada combinacion segun cuantos aliases
        conocidos coinciden y retorna la de mayor puntaje (evita leer hojas auxiliares
        como 'LISTAS DESPLEGABLES' antes que la hoja de datos real).
        """
        import pandas as _pd
        # Construir set de aliases para puntuar candidatos
        _aliases = set()
        if clave and clave in _ALIAS_VALIDACION:
            for _lst in _ALIAS_VALIDACION[clave].values():
                _aliases.update(a.lower().strip() for a in _lst)

        best_cols  = []
        best_score = -1

        try:
            xls = _pd.ExcelFile(ruta)
            for sheet in xls.sheet_names[:6]:
                for hdr_row in [0, 1, 2, 3]:
                    try:
                        tmp = _pd.read_excel(ruta, sheet_name=sheet,
                                              header=hdr_row, nrows=1)
                        real = [str(c).strip() for c in tmp.columns
                                if "Unnamed" not in str(c) and str(c).strip()]
                        if len(real) < 3:
                            continue
                        score = (sum(1 for c in real if c.lower().strip() in _aliases)
                                 if _aliases else 0)
                        if score > best_score:
                            best_score = score
                            best_cols  = real
                        if score >= 3:   # coincidencias suficientes — no seguir buscando
                            return real
                    except Exception:
                        pass
        except Exception:
            pass
        return best_cols

    def _campo_cubierto(self, campo, clave, file_cols_lower, mapeos):
        """True si el campo está cubierto: por nombre exacto, alias, patron o mapeo guardado."""
        # Mapeo guardado por el usuario
        mapeo = mapeos.get(clave, {})
        if campo in mapeo:
            return mapeo[campo].lower().strip() in file_cols_lower
        # Aliases del pipeline
        aliases = _ALIAS_VALIDACION.get(clave, {}).get(campo, [campo.lower().strip()])
        for alias in aliases:
            if alias.lower().strip() in file_cols_lower:
                return True
        # Patron especial: Cód. PT en bodega (cualquier col con "pt" y "cod")
        if clave == "arch_bodega" and campo == "Cód. PT":
            for col_l in file_cols_lower:
                if "pt" in col_l and ("cód" in col_l or "cod" in col_l):
                    return True
        # OC en entradas: case-insensitive
        if clave == "arch_entradas" and campo == "OC":
            return "oc" in file_cols_lower
        return False

    def _sugerir_columna(self, campo, file_cols):
        """Devuelve la columna del archivo más parecida al campo esperado, o None."""
        if not file_cols:
            return None
        campo_norm  = campo.lower().replace(".", "").replace(" ", "")
        campo_words = set(campo.lower().replace(".", "").split())
        best_col, best_score = None, 0
        for col in file_cols:
            col_lower = col.lower().strip()
            col_norm  = col_lower.replace(".", "").replace(" ", "")
            if campo_norm in col_norm or col_norm in campo_norm:
                score = 10 + len(min(campo_norm, col_norm, key=len))
            else:
                score = len(campo_words & set(col_lower.replace(".", "").split()))
            if score > best_score:
                best_score, best_col = score, col
        return best_col if best_score > 0 else None

    def _detectar_problemas(self, ruta, clave):
        """Retorna lista de (campo, descr, sugerencia) para campos requeridos no encontrados."""
        if clave not in _COL_SCHEMA:
            return []
        file_cols = self._leer_columnas_archivo(ruta, clave)
        if not file_cols:
            return []
        file_cols_lower = {c.lower().strip(): c for c in file_cols}
        mapeos   = self._cargar_mapeos_columnas()
        problemas = []
        for campo, descr, req in _COL_SCHEMA[clave]["campos"]:
            if not req:
                continue
            if not self._campo_cubierto(campo, clave, file_cols_lower, mapeos):
                problemas.append((campo, descr,
                                   self._sugerir_columna(campo, file_cols)))
        return problemas

    def _abrir_modal_validacion(self, ruta, clave, problemas, file_cols):
        """Modal contextual que aparece solo cuando hay columnas requeridas sin mapear."""
        schema  = _COL_SCHEMA.get(clave, {})
        titulo  = schema.get("titulo", clave)
        fname   = os.path.basename(ruta)
        n       = len(problemas)
        opciones = ["(Sin cambio)"] + file_cols

        modal = tk.Toplevel(self.root)
        modal.title(f"Revision de columnas — {fname}")
        modal.configure(bg=C_MAIN)
        modal.resizable(True, True)
        modal.transient(self.root)
        modal.grab_set()
        mw = 780
        mh = min(200 + n * 100 + 90, 620)
        self.root.update_idletasks()
        mx = self.root.winfo_rootx() + (self.root.winfo_width()  - mw) // 2
        my = self.root.winfo_rooty() + (self.root.winfo_height() - mh) // 2
        modal.geometry(f"{mw}x{mh}+{max(0, mx)}+{max(0, my)}")
        modal.minsize(mw, 300)

        # Header
        hdr = tk.Frame(modal, bg=C_HEADER)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"  Revision de columnas — {titulo}",
                  font=("Segoe UI", 12, "bold"), bg=C_HEADER, fg="white"
                  ).pack(side="left", padx=20, pady=12)

        # Banner de advertencia
        C_WARN    = "#92400E"; C_WARN_BG = "#FFFBEB"; C_WARN_BD = "#F59E0B"
        banner = tk.Frame(modal, bg=C_WARN_BG)
        banner.pack(fill="x")
        tk.Frame(banner, bg=C_WARN_BD, width=5).pack(side="left", fill="y")
        msg_f = tk.Frame(banner, bg=C_WARN_BG)
        msg_f.pack(side="left", padx=14, pady=12)
        plural_c = "s" if n > 1 else ""
        plural_v = "n" if n > 1 else ""
        tk.Label(msg_f,
                  text=f"Se encontraron {n} campo{plural_c} requerido{plural_c} que no coincide{plural_v} con las columnas de  {fname}.",
                  font=("Segoe UI", 9, "bold"), bg=C_WARN_BG, fg=C_WARN,
                  anchor="w", justify="left").pack(anchor="w")
        tk.Label(msg_f,
                  text="Indica a qué columna del archivo corresponde cada campo, o continúa sin cambios.",
                  font=("Segoe UI", 8), bg=C_WARN_BG, fg="#78350F",
                  anchor="w").pack(anchor="w", pady=(3, 0))

        # Encabezado de tabla
        th = tk.Frame(modal, bg="#1E3A5F")
        th.pack(fill="x", padx=20, pady=(14, 0))
        tk.Label(th, text="  Campo que espera el sistema  (* = requerido)",
                  font=("Segoe UI", 9, "bold"), bg="#1E3A5F", fg="white",
                  anchor="w").pack(side="left", pady=7, padx=(0, 0))
        tk.Label(th, text="Columna en tu archivo Excel",
                  font=("Segoe UI", 9, "bold"), bg="#1E3A5F", fg="white",
                  anchor="w").pack(side="right", pady=7, padx=(0, 16))

        # Canvas con scroll para las filas
        canvas_w = tk.Frame(modal, bg="white")
        canvas_w.pack(fill="both", expand=True, padx=20, pady=(0, 4))
        canvas = tk.Canvas(canvas_w, bg="white", highlightthickness=0)
        vsb    = ttk.Scrollbar(canvas_w, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        rows_f = tk.Frame(canvas, bg="white")
        _wid   = canvas.create_window((0, 0), window=rows_f, anchor="nw")
        rows_f.bind("<Configure>", lambda e, c=canvas: c.configure(
            scrollregion=c.bbox("all")))
        canvas.bind("<Configure>", lambda e, c=canvas, w=_wid: c.itemconfig(
            w, width=e.width))
        canvas.bind("<Enter>", lambda e, c=canvas: c.bind_all(
            "<MouseWheel>", lambda ev, cv=c: cv.yview_scroll(int(-1*(ev.delta/120)), "units")))
        canvas.bind("<Leave>", lambda e, c=canvas: c.unbind_all("<MouseWheel>"))

        vars_campos = {}

        for i, (campo, descr, sugerencia) in enumerate(problemas):
            rbg = "white" if i % 2 == 0 else "#FAFAFA"
            row = tk.Frame(rows_f, bg=rbg)
            row.pack(fill="x")

            tk.Frame(row, bg=C_ACENTO, width=4).pack(side="left", fill="y")

            lf = tk.Frame(row, bg=rbg)
            lf.pack(side="left", fill="y")
            tk.Frame(lf, bg=rbg, width=260, height=1).pack()
            lpad = tk.Frame(lf, bg=rbg)
            lpad.pack(anchor="w", padx=14, pady=14)
            tk.Label(lpad, text=campo + " *",
                      font=("Segoe UI", 10, "bold"), bg=rbg, fg=C_TEXTO,
                      anchor="w", wraplength=240).pack(anchor="w")
            tk.Label(lpad, text=descr,
                      font=("Segoe UI", 8), bg=rbg, fg="#64748B",
                      anchor="w", wraplength=240, justify="left").pack(anchor="w", pady=(4, 0))

            tk.Frame(row, bg="#F1F5F9", width=1).pack(side="left", fill="y", pady=10)

            rf = tk.Frame(row, bg=rbg)
            rf.pack(side="left", fill="both", expand=True, padx=16, pady=14)

            var = tk.StringVar()
            vars_campos[campo] = var
            var.set(sugerencia if sugerencia else "(Sin cambio)")

            cb_row = tk.Frame(rf, bg=rbg)
            cb_row.pack(fill="x")
            cb = ttk.Combobox(cb_row, textvariable=var, values=opciones,
                               font=("Segoe UI", 9), state="readonly")
            cb.pack(side="left", fill="x", expand=True, ipady=5)
            cb.bind("<MouseWheel>", lambda e: "break")

            if sugerencia and sugerencia in file_cols:
                tk.Label(cb_row, text=" sugerido ",
                          font=("Segoe UI", 7, "bold"),
                          bg="#DBEAFE", fg=C_ACENTO,
                          padx=5, pady=2).pack(side="left", padx=(8, 0))

            tk.Frame(rows_f, bg="#E8EDF3", height=1).pack(fill="x")

        # Footer
        tk.Frame(modal, bg=C_DIVIDER, height=1).pack(fill="x", side="bottom")
        footer = tk.Frame(modal, bg=C_CARD, pady=10)
        footer.pack(fill="x", side="bottom")

        def _continuar():
            modal.destroy()

        def _guardar():
            mapeos     = self._cargar_mapeos_columnas()
            mapeo_clave = mapeos.get(clave, {})
            for campo_k, var_k in vars_campos.items():
                val = var_k.get().strip()
                if val and val != "(Sin cambio)":
                    mapeo_clave[campo_k] = val
                elif campo_k in mapeo_clave:
                    del mapeo_clave[campo_k]
            if mapeo_clave:
                mapeos[clave] = mapeo_clave
            elif clave in mapeos:
                del mapeos[clave]
            self._guardar_mapeos_columnas(mapeos)
            modal.destroy()
            messagebox.showinfo(
                "Mapeo guardado",
                f"El mapeo para {fname} fue guardado.\n"
                "Se aplicará automáticamente al ejecutar el pipeline.")

        self._make_btn(footer, "  Guardar mapeo  ", C_ACENTO, C_ACENTO_H,
                        _guardar).pack(side="right", ipady=6, padx=(8, 20))
        self._make_btn(footer, "  Continuar sin cambios  ", C_CANCEL, C_CANCEL_H,
                        _continuar).pack(side="right", ipady=6)

    def _validar_archivo_cargado(self, ruta, clave):
        """Orquesta la validación y abre el modal solo si hay columnas requeridas faltantes."""
        if not ruta or not os.path.isfile(ruta) or clave not in _COL_SCHEMA:
            return
        try:
            file_cols = self._leer_columnas_archivo(ruta, clave)
            problemas = self._detectar_problemas(ruta, clave)
            if problemas:
                self._abrir_modal_validacion(ruta, clave, problemas, file_cols)
        except Exception:
            pass  # Nunca bloquear al usuario por un error de validacion

    # ── Exportar Excel ────────────────────────────────────────────────────────

    def _exportar_excel(self):
        ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
        ruta = (self._last_excel if self._last_excel and os.path.isfile(self._last_excel)
                else ruta_auto if os.path.isfile(ruta_auto) else None)
        if not ruta:
            messagebox.showwarning("Sin archivo", "Genera primero un plan de produccion.")
            return
        dest = filedialog.asksaveasfilename(
            title="Guardar copia del Excel", defaultextension=".xlsx",
            initialfile=os.path.basename(ruta), filetypes=[("Excel","*.xlsx")])
        if not dest: return
        try:
            shutil.copy2(ruta, dest)
            if messagebox.askyesno("Exportado", f"Guardado en:\n{dest}\n\n¿Abrir ahora?"):
                os.startfile(dest)
        except Exception as exc:
            messagebox.showerror("Error al exportar", str(exc))

    # ── Exportar PDF ──────────────────────────────────────────────────────────

    def _exportar_pdf(self):
        ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
        ruta = (self._last_excel if self._last_excel and os.path.isfile(self._last_excel)
                else ruta_auto if os.path.isfile(ruta_auto) else None)
        if not ruta:
            messagebox.showwarning("Sin archivo", "Genera primero un plan de produccion.")
            return
        dest = filedialog.asksaveasfilename(
            title="Guardar PDF del Dashboard", defaultextension=".pdf",
            initialfile="SeedPack_Dashboard.pdf", filetypes=[("PDF","*.pdf")])
        if not dest: return
        try:
            import pandas as pd
            import matplotlib; matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            import matplotlib.ticker as mticker
            import matplotlib.patches as mpatches
            from matplotlib.backends.backend_pdf import PdfPages
        except ImportError as exc:
            messagebox.showerror("Dependencia faltante", str(exc)); return
        try:
            xls = pd.ExcelFile(ruta)
            plan_df = pd.read_excel(xls, "Plan_Diario",        header=1) if "Plan_Diario"        in xls.sheet_names else pd.DataFrame()
            ops_df  = pd.read_excel(xls, "OPs_En_Proceso",     header=1) if "OPs_En_Proceso"     in xls.sheet_names else pd.DataFrame()
            inv_df  = pd.read_excel(xls, "Inventario_Actual",  header=1) if "Inventario_Actual"  in xls.sheet_names else pd.DataFrame()
            ocas_df = pd.read_excel(xls, "Pedidos_Ocasionales",header=1) if "Pedidos_Ocasionales" in xls.sheet_names else pd.DataFrame()
            dash_df = pd.read_excel(xls, "Dashboard",          header=None) if "Dashboard"       in xls.sheet_names else pd.DataFrame()
        except Exception as exc:
            messagebox.showerror("Error leyendo Excel", str(exc)); return

        def _kpi(df, row, col):
            try:
                v = df.iloc[row, col]
                return 0 if (v != v) else v
            except: return 0

        dias   = int(_kpi(dash_df, 3, 1)); refs   = int(_kpi(dash_df, 3, 3))
        unds   = int(_kpi(dash_df, 3, 5)); r_proy = int(_kpi(dash_df, 8, 1))
        r_ocas = int(_kpi(dash_df, 8, 3)); u_ocas = int(_kpi(dash_df, 8, 5))
        stock_cub = int(_kpi(dash_df, 13, 1))
        n_ops = len(ops_df) if not ops_df.empty else 0

        BG = "#0F1E2E"; FG = "#E2E8F0"; AC = "#2563EB"

        with PdfPages(dest) as pdf:
            # ── Pagina 1: Portada + KPIs ──────────────────────────────────
            fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG)
            fig.text(0.5, 0.88, "SeedPack Planner", ha="center", va="center",
                     fontsize=28, fontweight="bold", color="white")
            fig.text(0.5, 0.82, "Dashboard Ejecutivo — Plan de Produccion",
                     ha="center", fontsize=14, color="#7EB8E0")
            fig.text(0.5, 0.77, f"Archivo: {os.path.basename(ruta)}",
                     ha="center", fontsize=9, color="#64748B")

            kpis_pdf = [
                ("Dias planificados", f"{dias:,}", "#2563EB"),
                ("Referencias en plan", f"{refs:,}", "#7C3AED"),
                ("Unidades a producir", f"{unds:,}", "#0891B2"),
                ("Refs. proyeccion", f"{r_proy:,}", "#16A34A"),
                ("Refs. ocasionales", f"{r_ocas:,}", "#EA580C"),
                ("Uds. ocasionales", f"{u_ocas:,}", "#DC2626"),
                ("Cubiertos por stock", f"{stock_cub:,}", "#CA8A04"),
                ("OPs en proceso", f"{n_ops:,}", "#0891B2"),
            ]
            cols_k = 4; rows_k = 2
            for i, (label, val, color) in enumerate(kpis_pdf):
                r, c = divmod(i, cols_k)
                x = 0.08 + c * 0.23; y = 0.60 - r * 0.20
                ax_k = fig.add_axes([x, y, 0.20, 0.16])
                ax_k.set_facecolor("#162B45")
                for sp in ax_k.spines.values():
                    sp.set_edgecolor(color); sp.set_linewidth(2)
                ax_k.set_xticks([]); ax_k.set_yticks([])
                ax_k.text(0.5, 0.65, val, transform=ax_k.transAxes,
                          ha="center", va="center", fontsize=18, fontweight="bold", color=color)
                ax_k.text(0.5, 0.25, label, transform=ax_k.transAxes,
                          ha="center", va="center", fontsize=8, color=FG)
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

            # ── Pagina 2: Graficas Plan_Diario ────────────────────────────
            if not plan_df.empty:
                fig, axes = plt.subplots(1, 3, figsize=(11.69, 8.27), facecolor=BG)
                fig.suptitle("Analisis del Plan Diario", color=FG, fontsize=14, fontweight="bold", y=0.98)
                for ax in axes: ax.set_facecolor(BG)

                # Top 10 refs
                if "Referencia" in plan_df.columns and "Cantidad a Producir" in plan_df.columns:
                    try:
                        top10 = plan_df.groupby("Referencia")["Cantidad a Producir"].sum().sort_values(ascending=False).head(10)
                        colors_b = plt.cm.Blues_r([0.3+0.5*i/max(len(top10)-1,1) for i in range(len(top10))])
                        axes[0].barh([str(r)[:28] for r in top10.index[::-1]], top10.values[::-1], color=colors_b, height=0.65)
                        axes[0].set_title("Top 10 Referencias", color=FG, fontsize=10, fontweight="bold")
                        axes[0].tick_params(colors=FG, labelsize=7)
                        axes[0].xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{int(x):,}"))
                        for sp in axes[0].spines.values(): sp.set_color("#1E3A58")
                    except: pass

                # Produccion semanal
                if "Fecha" in plan_df.columns and "Cantidad a Producir" in plan_df.columns:
                    try:
                        pc = plan_df.copy()
                        pc["Fecha"] = pd.to_datetime(pc["Fecha"], errors="coerce")
                        pc["sem"]   = pc["Fecha"].dt.to_period("W").astype(str)
                        agg = pc.groupby("sem")["Cantidad a Producir"].sum().reset_index().sort_values("sem")
                        xs  = range(len(agg))
                        axes[1].fill_between(xs, agg["Cantidad a Producir"], alpha=0.25, color="#3B82F6")
                        axes[1].plot(xs, agg["Cantidad a Producir"], color="#3B82F6", linewidth=2, marker="o", markersize=3, markerfacecolor="white")
                        axes[1].set_title("Produccion Semanal", color=FG, fontsize=10, fontweight="bold")
                        step2 = max(1, len(agg) // 10)
                        tick_pos2 = list(range(0, len(agg), step2))
                        axes[1].set_xticks(tick_pos2)
                        axes[1].set_xticklabels([str(agg["sem"].iloc[i])[-8:] for i in tick_pos2], rotation=40, ha="right", color=FG, fontsize=6)
                        axes[1].tick_params(colors=FG, labelsize=7)
                        axes[1].yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{int(x):,}"))
                        for sp in axes[1].spines.values(): sp.set_color("#1E3A58")
                    except: pass

                # Donut tipo
                if "Tipo" in plan_df.columns:
                    try:
                        n_p = int((plan_df["Tipo"]=="Proyeccion").sum())
                        n_o = int((plan_df["Tipo"]=="Ocasional").sum())
                        if n_p+n_o > 0:
                            wedges, _, autotexts = axes[2].pie([n_p, n_o], colors=["#3B82F6","#EA580C"],
                                autopct="%1.0f%%", startangle=90,
                                wedgeprops=dict(width=0.55, edgecolor=BG, linewidth=2), pctdistance=0.75)
                            for at in autotexts: at.set_color("white"); at.set_fontsize(9)
                            axes[2].set_title("Tipo de Orden", color=FG, fontsize=10, fontweight="bold")
                            axes[2].legend(handles=[mpatches.Patch(color=c,label=l) for c,l in zip(["#3B82F6","#EA580C"],["Proyeccion","Ocasional"])],
                                           loc="lower center", fontsize=8, frameon=False, labelcolor=FG, ncol=2)
                    except: pass
                plt.tight_layout(pad=1.5)
                pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

            # ── Pagina 3: Inventario + OPs ────────────────────────────────
            fig, axes = plt.subplots(1, 2, figsize=(11.69, 8.27), facecolor=BG)
            fig.suptitle("Inventario y OPs en Proceso", color=FG, fontsize=14, fontweight="bold", y=0.98)
            for ax in axes: ax.set_facecolor(BG)
            if not inv_df.empty and "Saldo" in inv_df.columns:
                try:
                    ncol = "Referencia" if "Referencia" in inv_df.columns else inv_df.columns[2]
                    inv_df["Saldo"] = pd.to_numeric(inv_df["Saldo"], errors="coerce").fillna(0)
                    top_inv = inv_df.groupby(ncol)["Saldo"].sum().sort_values(ascending=False).head(15)
                    axes[0].bar(range(len(top_inv)), top_inv.values, color="#10B981", alpha=0.85, width=0.7)
                    axes[0].set_xticks(range(len(top_inv)))
                    axes[0].set_xticklabels([str(n)[:16] for n in top_inv.index], rotation=45, ha="right", color=FG, fontsize=6)
                    axes[0].set_title("Top 15 Inventario en Bodega", color=FG, fontsize=10, fontweight="bold")
                    axes[0].tick_params(colors=FG, labelsize=7)
                    axes[0].yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{int(x):,}"))
                    for sp in axes[0].spines.values(): sp.set_color("#1E3A58")
                except: pass
            if not ops_df.empty and "Tipo de Trabajo" in ops_df.columns:
                try:
                    tipo_counts = ops_df["Tipo de Trabajo"].value_counts()
                    colors_ops = ["#F59E0B","#EF4444","#8B5CF6","#06B6D4","#10B981"]
                    axes[1].pie(tipo_counts.values, labels=None, colors=colors_ops[:len(tipo_counts)],
                                startangle=90, wedgeprops=dict(edgecolor=BG, linewidth=2),
                                autopct="%1.0f%%", pctdistance=0.8)
                    axes[1].set_title("OPs por Tipo de Trabajo", color=FG, fontsize=10, fontweight="bold")
                    axes[1].legend(handles=[mpatches.Patch(color=colors_ops[i],label=l)
                                   for i,l in enumerate(tipo_counts.index[:5])],
                                   loc="lower center", fontsize=8, frameon=False, labelcolor=FG, ncol=3)
                except: pass
            plt.tight_layout(pad=1.5)
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

            # ── Pagina 4: Tabla Plan_Diario (primeras 60 filas) ───────────
            if not plan_df.empty:
                cols_show = ["Fecha","Codigo PT","Referencia","Cantidad a Producir","Tipo"]
                cols_ok   = [c for c in cols_show if c in plan_df.columns]
                sub = plan_df[cols_ok].head(60).fillna("").astype(str)
                fig, ax = plt.subplots(figsize=(11.69, 8.27), facecolor=BG)
                fig.suptitle("Plan Diario — Primeros 60 registros", color=FG, fontsize=12, fontweight="bold")
                ax.axis("off")
                tbl_data = [sub.columns.tolist()] + sub.values.tolist()
                tbl = ax.table(cellText=tbl_data[1:], colLabels=tbl_data[0],
                               cellLoc="center", loc="center")
                tbl.auto_set_font_size(False); tbl.set_fontsize(7)
                tbl.auto_set_column_width(col=list(range(len(cols_ok))))
                for (row, col), cell in tbl.get_celld().items():
                    if row == 0:
                        cell.set_facecolor(AC); cell.set_text_props(color="white", fontweight="bold")
                    else:
                        cell.set_facecolor("#162B45" if row%2==0 else "#1E3A58")
                        cell.set_text_props(color=FG)
                    cell.set_edgecolor("#0F1E2E")
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        if messagebox.askyesno("PDF Exportado", f"Guardado en:\n{dest}\n\n¿Abrir ahora?"):
            os.startfile(dest)

    # ── Cambio de pestana ─────────────────────────────────────────────────────

    def _rebind_scroll(self, canvas):
        """Redirige el scroll del mouse al canvas activo."""
        if canvas is not None:
            self.root.bind_all("<MouseWheel>",
                lambda e, c=canvas: c.yview_scroll(-1*(e.delta//120), "units"))
        else:
            self.root.unbind_all("<MouseWheel>")

    def _on_tab_change(self, event):
        selected = self.notebook.tab(self.notebook.select(), "text").strip()
        ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
        ruta = (self._last_excel if self._last_excel and os.path.isfile(self._last_excel)
                else ruta_auto if os.path.isfile(ruta_auto) else None)
        if "Dashboard" in selected:
            self._rebind_scroll(self._dash_canvas)
            if ruta and self._dash_loaded != ruta:
                self._renderizar_dashboard(ruta)
        elif "Ordenes" in selected:
            self._rebind_scroll(self._canvas)
        elif "Gestion" in selected:
            self._rebind_scroll(None)
            self._analizar_pedidos()
        elif "Saldo" in selected:
            self._rebind_scroll(None)
            self._analizar_saldo_ops()
        else:
            # Plan Diario y Tablas usan Treeview con scroll nativo
            self._rebind_scroll(None)
            if "Plan" in selected and not self._plan_data and ruta:
                self._last_excel = self._last_excel or ruta
                self._cargar_plan_diario()
            elif "Tablas" in selected and self._tablas_loaded != ruta and ruta:
                self._cargar_tablas_excel(ruta)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _centrar(self, w, h):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth(); sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _elegir_archivo(self, var, cfg=None):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel","*.xlsx *.xls"), ("Todos los archivos","*.*")])
        if ruta:
            ruta_final = ruta
            if cfg and cfg.get("preloaded") and cfg.get("default_file"):
                # Archivos precargados: copiar a Archivos Historicos para persistir al reabrir
                dest = os.path.join(DATA_DIR, "Archivos Historicos", cfg["default_file"])
                try:
                    os.makedirs(os.path.dirname(dest), exist_ok=True)
                    shutil.copy2(ruta, dest)
                    var.set(dest)
                    ruta_final = dest
                    messagebox.showinfo(
                        "Archivo actualizado",
                        f"'{cfg['titulo']}' actualizado correctamente.\n\n"
                        f"Guardado en:\n{dest}")
                except Exception as exc:
                    var.set(ruta)
                    messagebox.showwarning(
                        "No se pudo copiar",
                        f"No se pudo guardar el archivo en Archivos Historicos:\n{exc}\n\n"
                        f"Se usara desde la ubicacion original:\n{ruta}\n\n"
                        f"NOTA: al reiniciar la app se cargara el archivo anterior.")
            else:
                # Archivos manuales: usar directamente desde donde el usuario los tenga
                var.set(ruta)

            # Persistir la ruta seleccionada para la proxima sesion
            self._guardar_settings()

            # Validar columnas del archivo cargado — abre modal solo si hay problemas
            if cfg and cfg.get("key"):
                clave = cfg["key"]
                rv    = ruta_final
                self.root.after(450, lambda r=rv, k=clave:
                                self._validar_archivo_cargado(r, k))

    def _guardar_settings(self):
        """Persiste las rutas de archivos seleccionadas para recuperarlas al reabrir."""
        try:
            data = {"file_paths": {k: v.get() for k, v in self.file_vars.items() if v.get()}}
            with open(_SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # ── Validacion ────────────────────────────────────────────────────────────

    def _validar(self):
        for etiqueta, var in [("Fecha inicial", self.fi_var), ("Fecha final", self.ff_var)]:
            val = var.get().strip()
            try: datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Fecha invalida",
                    f"{etiqueta}: '{val}' no es valida.\nUsa el formato YYYY-MM-DD.")
                return False
        fi = datetime.strptime(self.fi_var.get().strip(), "%Y-%m-%d")
        ff = datetime.strptime(self.ff_var.get().strip(), "%Y-%m-%d")
        if fi >= ff:
            messagebox.showerror("Rango invalido", "La fecha inicial debe ser anterior a la final.")
            return False
        for cfg in ARCHIVOS:
            if cfg.get("optional"):
                continue
            ruta = self.file_vars[cfg["key"]].get().strip()
            if not ruta:
                messagebox.showerror("Archivo requerido", f"Debes seleccionar:\n\n{cfg['titulo']}")
                return False
            if not os.path.isfile(ruta):
                messagebox.showerror("Archivo no encontrado", f"No se encontro:\n\n{ruta}")
                return False
        return True

    # ── Pipeline ──────────────────────────────────────────────────────────────

    def _ejecutar(self):
        if not self._validar(): return
        params = {
            "fi": self.fi_var.get().strip(), "ff": self.ff_var.get().strip(),
            "arch_ventas":   self.file_vars["arch_ventas"].get().strip(),
            "arch_ops_hist": self.file_vars["arch_ops_hist"].get().strip(),
            "arch_pedidos":  self.file_vars["arch_pedidos"].get().strip(),
            "arch_ops_proc": self.file_vars["arch_ops_proc"].get().strip(),
            "arch_bodega":   self.file_vars["arch_bodega"].get().strip(),
            "arch_lote_min": self.file_vars["arch_lote_min"].get().strip(),
            "arch_entradas":       self.file_vars["arch_entradas"].get().strip() or None,
            "arch_plan_comercial": self.file_vars["arch_plan_comercial"].get().strip() or None,
        }
        self._ventana_progreso(params)

    def _ventana_progreso(self, params):
        PASOS = [
            ("Limpieza de datos",        "Preparando y filtrando registros historicos"),
            ("Clusterizacion",           "Agrupando productos por patrones de demanda"),
            ("Modelos ML de prediccion", "Entrenando y evaluando modelos de prediccion"),
            ("Proyeccion de ventas",     "Generando proyecciones para el periodo"),
            ("Ordenes de produccion",    "Calculando el plan de produccion final"),
            ("Gestion de pedidos",       "Cruzando cuadro de pedidos con stock, OPs y proyeccion ML"),
        ]
        N           = len(PASOS)
        _cancelled  = threading.Event()
        _cur        = [0]
        _step_start = [None] * N
        _step_dur   = [None] * N
        _t0         = [time.time()]

        dlg = tk.Toplevel(self.root)
        dlg.title("Generando plan de produccion...")
        dlg.configure(bg=C_CARD)
        dlg.resizable(False, False)
        dlg.transient(self.root)
        dlg.grab_set()
        W, H = 540, 590
        dlg.update_idletasks()
        sw = dlg.winfo_screenwidth(); sh = dlg.winfo_screenheight()
        dlg.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(dlg, bg=C_HEADER)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚙", font=("Segoe UI", 22),
                 bg=C_HEADER, fg="#93C5FD").pack(pady=(18, 2))
        lbl_titulo = tk.Label(hdr, text="Generando plan de produccion",
                              font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white")
        lbl_titulo.pack()
        lbl_sub = tk.Label(hdr, text="Iniciando...",
                           font=("Segoe UI", 9), bg=C_HEADER, fg="#93C5FD")
        lbl_sub.pack(pady=(2, 16))

        # ── Body ──────────────────────────────────────────────────────────────
        body = tk.Frame(dlg, bg=C_CARD)
        body.pack(fill="both", expand=True, padx=28)

        # Progress bar
        pb_row = tk.Frame(body, bg=C_CARD)
        pb_row.pack(fill="x", pady=(18, 2))
        lbl_pct = tk.Label(pb_row, text="0%", font=("Segoe UI", 9, "bold"),
                           bg=C_CARD, fg=C_ACENTO, width=4, anchor="e")
        lbl_pct.pack(side="right")
        pb_canvas = tk.Canvas(pb_row, height=12, bg="#E2E8F0",
                              highlightthickness=0, relief="flat")
        pb_canvas.pack(side="left", fill="x", expand=True, padx=(0, 8))
        pb_fill = pb_canvas.create_rectangle(0, 0, 0, 12, fill=C_ACENTO, outline="")

        def _set_progress(pct):
            pb_canvas.update_idletasks()
            cw = pb_canvas.winfo_width() or 444
            pb_canvas.coords(pb_fill, 0, 0, int(cw * pct / 100), 12)
            lbl_pct.configure(text=f"{int(pct)}%")

        lbl_paso = tk.Label(body, text="Preparando...",
                            font=("Segoe UI", 8), bg=C_CARD, fg=C_GRIS, anchor="w")
        lbl_paso.pack(fill="x", pady=(0, 10))

        # Step rows
        step_widgets = []
        for i, (nombre, _) in enumerate(PASOS):
            row = tk.Frame(body, bg=C_CARD)
            row.pack(fill="x", pady=2)
            badge = tk.Label(row, text=str(i + 1), font=("Segoe UI", 8, "bold"),
                             bg=C_NUM_BG, fg=C_NUM_FG, width=2, padx=4, pady=2, relief="flat")
            badge.pack(side="left")
            name_lbl = tk.Label(row, text=nombre, font=("Segoe UI", 9),
                                bg=C_CARD, fg=C_GRIS, anchor="w")
            name_lbl.pack(side="left", padx=(8, 0), fill="x", expand=True)
            st_lbl = tk.Label(row, text="—", font=("Segoe UI", 8),
                              bg=C_CARD, fg=C_GRIS, width=12, anchor="e")
            st_lbl.pack(side="right")
            step_widgets.append((badge, name_lbl, st_lbl))

        # Divider + timers
        tk.Frame(body, bg=C_DIVIDER, height=1).pack(fill="x", pady=(14, 8))
        time_row = tk.Frame(body, bg=C_CARD)
        time_row.pack(fill="x")
        lbl_elapsed = tk.Label(time_row, text="Transcurrido: 0:00",
                               font=("Segoe UI", 8), bg=C_CARD, fg=C_GRIS)
        lbl_elapsed.pack(side="left")
        lbl_remaining = tk.Label(time_row, text="Estimado: calculando...",
                                 font=("Segoe UI", 8), bg=C_CARD, fg=C_GRIS)
        lbl_remaining.pack(side="right")
        lbl_status = tk.Label(body, text="", font=("Segoe UI", 8, "italic"),
                              bg=C_CARD, fg=C_GRIS, wraplength=484, anchor="w", justify="left")
        lbl_status.pack(fill="x", pady=(6, 0))

        # ── Cancel ────────────────────────────────────────────────────────────
        def _cancelar():
            if _cancelled.is_set():
                return
            _cancelled.set()
            lbl_titulo.configure(text="Proceso cancelado")
            lbl_sub.configure(text="Finalizando tarea en curso...")
            btn_cancel.configure(state="disabled", text="Cancelando...", bg="#94A3B8")
            dlg.after(1800, lambda: dlg.destroy() if dlg.winfo_exists() else None)

        dlg.protocol("WM_DELETE_WINDOW", _cancelar)
        btn_cancel = tk.Button(dlg, text="✕  Cancelar proceso",
                               font=("Segoe UI", 9, "bold"),
                               bg=C_CANCEL, fg="white", relief="flat",
                               activebackground=C_CANCEL_H, activeforeground="white",
                               cursor="hand2", padx=24, pady=9, bd=0,
                               command=_cancelar)
        btn_cancel.pack(pady=(6, 20))

        # ── Helpers ───────────────────────────────────────────────────────────
        def _fmt(s):
            m, sec = divmod(int(max(0, s)), 60)
            return f"{m}:{sec:02d}"

        def _set_step(idx, state):
            badge, name_lbl, st_lbl = step_widgets[idx]
            if state == "pending":
                badge.configure(bg=C_NUM_BG, fg=C_NUM_FG)
                name_lbl.configure(fg=C_GRIS, font=("Segoe UI", 9))
                st_lbl.configure(text="—", fg=C_GRIS)
            elif state == "active":
                badge.configure(bg=C_ACENTO, fg="white")
                name_lbl.configure(fg=C_TEXTO, font=("Segoe UI", 9, "bold"))
                st_lbl.configure(text="⏳ activo", fg=C_ACENTO)
            elif state == "done":
                badge.configure(bg="#DCFCE7", fg="#166534")
                name_lbl.configure(fg=C_TEXTO, font=("Segoe UI", 9))
                dur = _step_dur[idx]
                st_lbl.configure(
                    text=f"✓ {_fmt(dur)}" if dur is not None else "✓",
                    fg=C_SUCCESS)

        def _advance(to_step):
            prev = _cur[0]
            if _step_start[prev] is not None and _step_dur[prev] is None:
                _step_dur[prev] = time.time() - _step_start[prev]
            _set_step(prev, "done")
            if to_step < N:
                _step_start[to_step] = time.time()
                _set_step(to_step, "active")
                _cur[0] = to_step
                _set_progress((to_step / N) * 100)
                lbl_paso.configure(text=f"Paso {to_step + 1} de {N}  —  {PASOS[to_step][0]}")
                lbl_sub.configure(text=PASOS[to_step][1])

        # Init step 0
        _step_start[0] = time.time()
        _set_step(0, "active")
        _set_progress(0)
        lbl_paso.configure(text=f"Paso 1 de {N}  —  {PASOS[0][0]}")
        lbl_sub.configure(text=PASOS[0][1])

        # ── Tick (timers each second) ──────────────────────────────────────────
        def _tick():
            if not dlg.winfo_exists() or _cancelled.is_set():
                return
            lbl_elapsed.configure(text=f"Transcurrido: {_fmt(time.time() - _t0[0])}")
            done_durs = [d for d in _step_dur if d is not None]
            if done_durs:
                avg     = sum(done_durs) / len(done_durs)
                cs      = _cur[0]
                s_ela   = (time.time() - _step_start[cs]) if _step_start[cs] else 0
                est     = max(0, avg - s_ela) + max(0, N - len(done_durs) - 1) * avg
                lbl_remaining.configure(text=f"Estimado: ~{_fmt(est)}")
            dlg.after(1000, _tick)

        _tick()

        # ── Log parser ────────────────────────────────────────────────────────
        def _on_log(msg):
            if _cancelled.is_set():
                return
            m = msg.upper()
            if   "PASO 2" in m and _cur[0] < 1: _advance(1)
            elif "PASO 3" in m and _cur[0] < 2: _advance(2)
            elif "PASO 4" in m and _cur[0] < 3: _advance(3)
            elif "PASO 5" in m and _cur[0] < 4: _advance(4)
            elif "PASO 6" in m and _cur[0] < 5: _advance(5)
            stripped = msg.strip().lstrip("= ")
            if stripped and "PASO" not in stripped.upper() and "SEEDPACK" not in stripped.upper():
                lbl_status.configure(text=stripped[:90])

        # ── Pipeline thread ───────────────────────────────────────────────────
        def _pipeline():
            old_out = sys.stdout
            class _Redir:
                def write(self, msg):
                    if msg.strip() and not _cancelled.is_set():
                        try:
                            dlg.after(0, _on_log, msg.rstrip())
                        except Exception:
                            pass
                def flush(self): pass
            sys.stdout = _Redir()
            ruta = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
            try:
                import main as _main
                ruta_resultado = _main.run(
                    FI=params["fi"], FF=params["ff"], REF=None,
                    arch_ventas=params["arch_ventas"],
                    arch_ops_hist=params["arch_ops_hist"],
                    arch_pedidos=params["arch_pedidos"],
                    arch_ops_proc=params["arch_ops_proc"],
                    arch_bodega=params["arch_bodega"],
                    arch_lote_min=params["arch_lote_min"],
                    arch_entradas=params["arch_entradas"],
                    arch_plan_comercial=params["arch_plan_comercial"],
                )
                if ruta_resultado and os.path.isfile(ruta_resultado):
                    ruta = ruta_resultado
                sys.stdout = old_out
                dlg.after(0, lambda r=ruta: _done(True, r))
            except Exception as exc:
                sys.stdout = old_out
                dlg.after(0, lambda: _done(False, str(exc)))

        def _done(ok, dato):
            if _cancelled.is_set():
                return
            cs = _cur[0]
            if _step_start[cs] is not None and _step_dur[cs] is None:
                _step_dur[cs] = time.time() - _step_start[cs]
            for i in range(N):
                _set_step(i, "done")
            _set_progress(100)
            lbl_paso.configure(text="Todos los pasos completados" if ok else "Error en el pipeline")
            lbl_sub.configure(text="Resultado generado correctamente" if ok else "Revisa el mensaje de error")
            lbl_status.configure(text="")
            dlg.after(700, lambda: _finalizar(ok, dato))

        def _finalizar(ok, dato):
            if _cancelled.is_set():
                return
            dlg.destroy()
            if ok:
                self._last_excel    = dato
                self._dash_loaded   = None
                self._tablas_loaded = None
                self._plan_data     = []
                self._btn_exportar_xl.configure(state="normal")
                self._btn_exportar_pdf.configure(state="normal")
                self.root.update_idletasks()
                self.notebook.select(self.tab_dash)
                self.root.update_idletasks()
                self._renderizar_dashboard(dato)
            else:
                messagebox.showerror("Error en el pipeline", f"Ocurrio un error:\n\n{dato}")

        threading.Thread(target=_pipeline, daemon=True).start()


    # ═════════════════════════════════════════════════════════════════════════
    # TAB 5 — Gestion de Pedidos
    # ═════════════════════════════════════════════════════════════════════════

    def _build_tab_gestion(self):
        topbar = tk.Frame(self.tab_gestion, bg=C_HEADER, height=48)
        topbar.pack(fill="x"); topbar.pack_propagate(False)
        tk.Label(topbar, text="GESTION DE PEDIDOS", font=("Segoe UI", 11, "bold"),
                 bg=C_HEADER, fg="white").pack(side="left", padx=24, pady=12)
        self._make_btn(topbar, "  Analizar  ", C_ACENTO, C_ACENTO_H,
                       self._analizar_pedidos).pack(side="right", padx=16, pady=8, ipady=4)
        self._make_btn(topbar, "  Exportar Excel  ", C_SUCCESS, C_SUCCESS_H,
                       self._exportar_gestion).pack(side="right", padx=(0, 4), pady=8, ipady=4)

        # ── KPI cards ────────────────────────────────────────────────────────
        kpi_row = tk.Frame(self.tab_gestion, bg=C_MAIN)
        kpi_row.pack(fill="x", padx=16, pady=(12, 0))
        self._gest_kpi = {}
        for i, (key, label, color) in enumerate([
            ("total_pedidos",  "Pedidos activos",           "#2563EB"),
            ("total_saldo",    "Saldo pendiente total",     "#7C3AED"),
            ("refs_op",        "Refs. con OP requerida",    "#DC2626"),
            ("total_producir", "Total a producir",          "#EA580C"),
        ]):
            card = tk.Frame(kpi_row, bg="white",
                            highlightbackground="#E2E8F0", highlightthickness=1)
            card.pack(side="left", fill="both", expand=True, padx=(0 if i == 0 else 8, 0))
            tk.Frame(card, bg=color, height=4).pack(fill="x")
            inner = tk.Frame(card, bg="white")
            inner.pack(fill="x", padx=14, pady=(10, 12))
            val_lbl = tk.Label(inner, text="—", font=("Segoe UI", 20, "bold"),
                               bg="white", fg=color, anchor="w")
            val_lbl.pack(anchor="w")
            tk.Label(inner, text=label, font=("Segoe UI", 8),
                     bg="white", fg=C_GRIS, anchor="w").pack(anchor="w", pady=(3, 0))
            self._gest_kpi[key] = val_lbl

        # ── Barra de filtro ───────────────────────────────────────────────────
        fbar = tk.Frame(self.tab_gestion, bg=C_MAIN)
        fbar.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(fbar, text="Buscar:", font=("Segoe UI", 9),
                 bg=C_MAIN, fg=C_TEXTO).pack(side="left", padx=(0, 8))
        self._gest_search = tk.StringVar()
        self._gest_search.trace_add("write", self._filtrar_gestion)
        tk.Entry(fbar, textvariable=self._gest_search, font=("Segoe UI", 9), width=30,
                 bg="white", fg=C_TEXTO, relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=C_DIVIDER,
                 highlightcolor=C_ACENTO).pack(side="left", ipady=5, padx=(0, 16))
        self._gest_lbl = tk.Label(fbar, text="Sin datos — carga los archivos y presiona Analizar",
                                   font=("Segoe UI", 8), bg=C_MAIN, fg=C_GRIS)
        self._gest_lbl.pack(side="left")

        # ── Barra de filtros (Estado + Origen O.C) — botones combinables ───────
        self._gest_filtro_estado = "Todos"
        self._gest_filtro_origen = "Todos"
        self._gest_chips = {}
        ley = tk.Frame(self.tab_gestion, bg=C_MAIN)
        ley.pack(fill="x", padx=16, pady=(6, 0))

        tk.Label(ley, text="Estado:", font=("Segoe UI", 8, "bold"),
                 bg=C_MAIN, fg=C_GRIS).pack(side="left", padx=(0, 6))
        for val, abg, afg in [("Todos", "#475569", "white"),
                              ("OK — Cubierto", "#16A34A", "white"),
                              ("Solicitar OP", "#D97706", "white")]:
            self._mk_gest_chip(ley, "estado", val, abg, afg)

        tk.Frame(ley, bg=C_MAIN, width=24).pack(side="left")

        tk.Label(ley, text="Origen O.C:", font=("Segoe UI", 8, "bold"),
                 bg=C_MAIN, fg=C_GRIS).pack(side="left", padx=(0, 6))
        for val, abg, afg in [("Todos", "#475569", "white"),
                              ("Abastecimiento", "#2563EB", "white"),
                              ("En firme", "#BE185D", "white")]:
            self._mk_gest_chip(ley, "origen", val, abg, afg)

        self._restyle_gest_chips()

        # ── Treeview ──────────────────────────────────────────────────────────
        tbl = tk.Frame(self.tab_gestion, bg=C_MAIN)
        tbl.pack(fill="both", expand=True, padx=16, pady=10)
        sty = ttk.Style()
        sty.configure("Gest.Treeview", background="white", foreground=C_TEXTO,
                       fieldbackground="white", rowheight=26, font=("Segoe UI", 9))
        sty.configure("Gest.Treeview.Heading", background=C_HEADER, foreground="white",
                       font=("Segoe UI", 9, "bold"))
        sty.map("Gest.Treeview", background=[("selected", C_ACENTO)],
                foreground=[("selected", "white")])

        COLS_G = ["Codigo", "Descripcion", "Pedidos", "Saldo Pend.", "Stock Bodega",
                  "OPs Proceso", "Proy. ML", "Disponible", "Dem. Total",
                  "Deficit", "Lote Min.", "A Producir", "Estado", "Justificacion"]
        WIDTHS_G = {"Codigo": 130, "Descripcion": 270, "Pedidos": 68,
                    "Saldo Pend.": 100, "Stock Bodega": 100, "OPs Proceso": 100,
                    "Proy. ML": 90, "Disponible": 100, "Dem. Total": 100,
                    "Deficit": 88, "Lote Min.": 90, "A Producir": 100,
                    "Estado": 110, "Justificacion": 420}
        CENTERS_G = {"Codigo", "Pedidos", "Saldo Pend.", "Stock Bodega", "OPs Proceso",
                     "Proy. ML", "Disponible", "Dem. Total", "Deficit",
                     "Lote Min.", "A Producir", "Estado"}

        self._gest_tree = ttk.Treeview(tbl, columns=COLS_G, show="headings",
                                        style="Gest.Treeview")
        for col in COLS_G:
            self._gest_tree.heading(col, text=col)
            self._gest_tree.column(col, width=WIDTHS_G.get(col, 100),
                                   anchor="center" if col in CENTERS_G else "w",
                                   minwidth=60)
        vsb = ttk.Scrollbar(tbl, orient="vertical",   command=self._gest_tree.yview)
        hsb = ttk.Scrollbar(tbl, orient="horizontal", command=self._gest_tree.xview)
        self._gest_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._gest_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tbl.rowconfigure(0, weight=1); tbl.columnconfigure(0, weight=1)

    def _refrescar_gest_tree(self, data):
        self._gest_tree.delete(*self._gest_tree.get_children())
        for row in data:
            estado = row[12] if len(row) > 12 else ""
            tag = "op_req" if "Solicitar" in estado else "ok_row"
            # Solo se muestran las 14 columnas visibles; row[14] (origen) es metadata
            self._gest_tree.insert("", "end", values=row[:14], tags=(tag,))
        self._gest_tree.tag_configure("op_req", background="#FFF3CD")
        self._gest_tree.tag_configure("ok_row", background="#D4EDDA")

    def _mk_gest_chip(self, parent, grupo, valor, active_bg, active_fg):
        """Crea un boton-chip de filtro (Estado u Origen) que se resalta al estar activo."""
        f = tk.Frame(parent, bg="white", padx=10, pady=4, cursor="hand2",
                     highlightbackground=C_DIVIDER, highlightthickness=1)
        f.pack(side="left", padx=(0, 6))
        lbl = tk.Label(f, text=valor, font=("Segoe UI", 8, "bold"),
                       bg="white", fg=C_TEXTO, cursor="hand2")
        lbl.pack()

        def _click(_=None):
            if grupo == "estado":
                self._gest_filtro_estado = valor
            else:
                self._gest_filtro_origen = valor
            self._restyle_gest_chips()
            self._filtrar_gestion()

        f.bind("<Button-1>", _click)
        lbl.bind("<Button-1>", _click)
        self._gest_chips[(grupo, valor)] = (f, lbl, active_bg, active_fg)

    def _restyle_gest_chips(self):
        for (grupo, valor), (f, lbl, abg, afg) in self._gest_chips.items():
            activo = valor == (self._gest_filtro_estado if grupo == "estado"
                               else self._gest_filtro_origen)
            if activo:
                f.configure(bg=abg, highlightbackground=abg)
                lbl.configure(bg=abg, fg=afg)
            else:
                f.configure(bg="white", highlightbackground=C_DIVIDER)
                lbl.configure(bg="white", fg=C_TEXTO)

    def _set_kpis_gestion(self, rows):
        total_ped = sum(int(str(r[2]).replace(",", "")) for r in rows)
        total_sal = sum(int(str(r[3]).replace(",", "")) for r in rows)
        refs_op   = sum(1 for r in rows if r[12] == "Solicitar OP")
        total_pro = sum(int(str(r[11]).replace(",", "")) for r in rows)
        self._gest_kpi["total_pedidos"].config(text=f"{total_ped:,}")
        self._gest_kpi["total_saldo"].config(text=f"{total_sal:,}")
        self._gest_kpi["refs_op"].config(text=f"{refs_op:,}")
        self._gest_kpi["total_producir"].config(text=f"{total_pro:,}")

    def _filtrar_gestion(self, *_):
        q  = self._gest_search.get().lower()
        fe = getattr(self, "_gest_filtro_estado", "Todos")
        fo = getattr(self, "_gest_filtro_origen", "Todos")
        filtered = []
        for r in self._gestion_data:
            if q and not any(q in str(v).lower() for v in r[:14]):
                continue
            estado = r[12]
            if fe == "OK — Cubierto" and estado != "OK":
                continue
            if fe == "Solicitar OP" and estado != "Solicitar OP":
                continue
            origins = r[14] if len(r) > 14 else set()
            if fo != "Todos" and fo not in origins:
                continue
            filtered.append(r)

        self._gest_filtrado = filtered
        self._refrescar_gest_tree(filtered)
        self._set_kpis_gestion(filtered)
        total = len(self._gestion_data)
        extra = f"  (de {total:,})" if len(filtered) != total else ""
        self._gest_lbl.config(
            text=f"{len(filtered):,} referencias{extra}{getattr(self, '_gest_nota', '')}")

    def _exportar_gestion(self):
        """Exporta la gestion de pedidos a Excel respetando el filtro activo."""
        rows = getattr(self, "_gest_filtrado", None) or self._gestion_data
        if not rows:
            messagebox.showwarning(
                "Sin datos", "Presiona Analizar para generar la gestion de pedidos.")
            return
        fe = getattr(self, "_gest_filtro_estado", "Todos")
        fo = getattr(self, "_gest_filtro_origen", "Todos")
        suf = [s.replace("—", "").replace(" ", "_").strip("_")
               for s in (fe, fo) if s != "Todos"]
        nombre = "Gestion_Pedidos" + ("_" + "_".join(suf) if suf else "") + ".xlsx"
        dest = filedialog.asksaveasfilename(
            title="Exportar gestion de pedidos", defaultextension=".xlsx",
            initialfile=nombre, filetypes=[("Excel", "*.xlsx")])
        if not dest:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            HEADERS = ["Codigo", "Descripcion", "Pedidos", "Saldo Pendiente",
                       "Stock Bodega", "OPs Proceso", "Proyeccion ML", "Disponible",
                       "Demanda Total", "Deficit", "Lote Minimo", "A Producir",
                       "Estado", "Origen O.C", "Justificacion"]
            WIDTHS  = [18, 40, 10, 15, 15, 15, 15, 15, 15, 12, 14, 14, 16, 18, 80]
            NUMERIC = {4, 5, 6, 7, 8, 9, 10, 11, 12}   # columnas con miles

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gestion_Pedidos"
            bord = Border(*[Side(style="thin", color="BFBFBF")] * 4)

            for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = Font(bold=True, color="FFFFFF", size=11)
                c.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = bord
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.row_dimensions[1].height = 30

            for ri, r in enumerate(rows, 2):
                is_op  = (r[12] == "Solicitar OP")
                bg     = "FFF3CD" if is_op else "D4EDDA"
                origen = " + ".join(sorted(r[14])) if len(r) > 14 and r[14] else ""
                vals   = list(r[:13]) + [origen, r[13]]
                for ci, val in enumerate(vals, 1):
                    if ci in NUMERIC and isinstance(val, str):
                        try:
                            val = int(val.replace(",", ""))
                        except ValueError:
                            pass
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                    cell.border = bord
                    cell.alignment = Alignment(
                        horizontal="center" if ci not in (2, 15) else "left",
                        vertical="center", wrap_text=(ci == 15))
                    if ci in NUMERIC and isinstance(cell.value, int):
                        cell.number_format = "#,##0"

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
            wb.save(dest)
            if messagebox.askyesno("Exportado", f"Guardado en:\n{dest}\n\n¿Abrir ahora?"):
                os.startfile(dest)
        except Exception as exc:
            messagebox.showerror("Error al exportar", str(exc))

    def _analizar_pedidos(self):
        import math

        path_cuadro = self.file_vars["arch_pedidos"].get().strip()
        path_bodega = self.file_vars["arch_bodega"].get().strip()
        path_ops    = self.file_vars["arch_ops_proc"].get().strip()
        path_lmin   = self.file_vars["arch_lote_min"].get().strip()

        if not path_cuadro or not os.path.isfile(path_cuadro):
            self._gest_lbl.config(
                text="Carga el archivo 03 (Listado de pedidos) para ver el analisis.")
            return

        try:
            import pandas as pd
        except ImportError as exc:
            messagebox.showerror("Falta pandas", str(exc)); return

        # ── Cuadro de pedidos (hoja Maestro o primera hoja disponible) ───────
        try:
            _xl_ped = pd.ExcelFile(path_cuadro)
            df_raw  = None
            if "Maestro" in _xl_ped.sheet_names:
                try:
                    _df_m = pd.read_excel(_xl_ped, sheet_name="Maestro", header=3)
                    if any(str(c).upper() == "PEDIDO" for c in _df_m.columns):
                        df_raw = _df_m
                except Exception:
                    pass
            if df_raw is None:
                df_raw = pd.read_excel(_xl_ped)   # fallback: primera hoja

            # Aplicar mapeo personalizado de columnas si existe
            try:
                import json as _json
                _mp = os.path.join(DATA_DIR, "column_mappings.json")
                if os.path.isfile(_mp):
                    _m = _json.load(open(_mp, encoding="utf-8")).get("arch_pedidos", {})
                    _rn = {v: k for k, v in _m.items() if v and v in df_raw.columns and v != k}
                    if _rn:
                        df_raw = df_raw.rename(columns=_rn)
            except Exception:
                pass

            df_raw.columns = [str(c).strip() for c in df_raw.columns]
            col_up  = {c.upper(): c for c in df_raw.columns}
            sp_col  = (col_up.get("SALDO PENDIENTE")
                       or col_up.get("CANTIDAD")
                       or col_up.get("CANT.")
                       or "SALDO PENDIENTE")
            cod_col = (col_up.get("CODIGO")
                       or col_up.get("CÓDIGO")
                       or next((v for k, v in col_up.items()
                                if "CODIGO" in k or "CÓDIGO" in k), None)
                       or "CODIGO")
            desc_col = (col_up.get("DESCRIPCION")
                        or col_up.get("DESCRIPCIÓN")
                        or col_up.get("REFERENCIA")
                        or "DESCRIPCION")
            ped_col  = col_up.get("PEDIDO", "PEDIDO")
            # Columna "O.C del cliente #": para clasificar origen del pedido
            oc_col   = (col_up.get("O.C DEL CLIENTE #")
                        or col_up.get("O.C DEL CLIENTE")
                        or next((v for k, v in col_up.items()
                                 if "O.C" in k or "OC DEL CLIENTE" in k
                                 or ("CLIENTE" in k and "#" in k)), None))

            df_raw[sp_col] = pd.to_numeric(df_raw[sp_col], errors="coerce").fillna(0)
            needed = [c for c in [ped_col, cod_col, desc_col, sp_col, oc_col]
                      if c and c in df_raw.columns]
            df_pend = df_raw[df_raw[sp_col] > 0][needed].copy()
            df_pend = df_pend[df_pend[cod_col].notna()]
            df_pend[cod_col] = df_pend[cod_col].astype(str).str.strip()

            # Origen de la O.C: "Abastecimiento" vs "En firme" (cualquier otro nombre).
            # Una referencia puede tener pedidos de ambos tipos -> se guarda el conjunto.
            # str() dentro de la funcion: las celdas vacias (NaN) quedan como "nan".
            def _clasif_origen(v):
                return "Abastecimiento" if "abastec" in str(v).strip().lower() else "En firme"
            if oc_col and oc_col in df_pend.columns:
                df_pend["_origen"] = df_pend[oc_col].apply(_clasif_origen)
            else:
                df_pend["_origen"] = "En firme"
            origen_por_cod = df_pend.groupby(cod_col)["_origen"].agg(
                lambda s: set(s)).to_dict()
        except Exception as exc:
            messagebox.showerror("Error leyendo cuadro de pedidos", str(exc)); return

        # ── Stock en bodega ───────────────────────────────────────────────────
        stock_dict = {}
        if path_bodega and os.path.isfile(path_bodega):
            try:
                df_bod = pd.read_excel(path_bodega)
                cb = {c.lower().strip(): c for c in df_bod.columns}
                if "bodega" not in cb:
                    for alias in ("bodega sale", "bodega_sale", "almacen"):
                        if alias in cb:
                            df_bod = df_bod.rename(columns={cb[alias]: "Bodega"}); break
                pt_ok = any("pt" in c.lower() and ("cód" in c.lower() or "cod" in c.lower())
                            for c in df_bod.columns)
                if not pt_ok:
                    cb2 = {c.lower().strip(): c for c in df_bod.columns}
                    for alias in ("prod. terminado", "prod.terminado", "producto terminado"):
                        if alias in cb2:
                            df_bod = df_bod.rename(columns={cb2[alias]: "Cód. PT"})
                            df_bod["Cód. PT"] = (df_bod["Cód. PT"].astype(str)
                                                  .str.split(" - ").str[0].str.strip())
                            break
                else:
                    for c in df_bod.columns:
                        if "pt" in c.lower() and ("cód" in c.lower() or "cod" in c.lower()):
                            df_bod = df_bod.rename(columns={c: "Cód. PT"}); break
                cb3 = {c.lower().strip(): c for c in df_bod.columns}
                if "saldo" not in cb3:
                    for alias in ("saldo en inventario", "cantidad", "cant.", "stock"):
                        if alias in cb3:
                            df_bod = df_bod.rename(columns={cb3[alias]: "Saldo"}); break
                if "Bodega" in df_bod.columns:
                    df_bod = df_bod[~df_bod["Bodega"].isin(["Obsoletos"])]
                if "Saldo" in df_bod.columns and "Cód. PT" in df_bod.columns:
                    df_bod["Saldo"] = pd.to_numeric(df_bod["Saldo"], errors="coerce").fillna(0)
                    stock_dict = df_bod.groupby("Cód. PT")["Saldo"].sum().to_dict()
            except Exception:
                pass

        # ── OPs en proceso ────────────────────────────────────────────────────
        ops_dict = {}
        if path_ops and os.path.isfile(path_ops):
            try:
                df_ops = pd.read_excel(path_ops)
                df_ops["Cant. Aprobada"] = pd.to_numeric(
                    df_ops.get("Cant. Aprobada", pd.Series(dtype=float)),
                    errors="coerce").fillna(0)
                cod_ops = next((c for c in df_ops.columns
                                if "producto" in c.lower() and "cód" in c.lower()), None)
                if cod_ops:
                    ops_dict = df_ops.groupby(cod_ops)["Cant. Aprobada"].sum().to_dict()
            except Exception:
                pass

        # ── Lote minimo ───────────────────────────────────────────────────────
        lmin_dict = {}
        if path_lmin and os.path.isfile(path_lmin):
            try:
                df_lm = pd.read_excel(path_lmin, skiprows=2, header=None)
                df_lm.columns = ["codigo", "referencia", "linea", "cavidades", "lote_minimo"]
                df_lm["lote_minimo"] = pd.to_numeric(
                    df_lm["lote_minimo"], errors="coerce").fillna(0).astype(int)
                lmin_dict = df_lm[df_lm["lote_minimo"] > 0].set_index(
                    "codigo")["lote_minimo"].to_dict()
            except Exception:
                pass

        # ── Proyeccion ML (si ya se ejecuto el pipeline) ──────────────────────
        proyeccion_dict = {}
        ruta_auto = os.path.join(DATA_DIR, "Resultado Final", "orden_produccion_final.xlsx")
        ruta_out  = (self._last_excel if self._last_excel and os.path.isfile(self._last_excel)
                     else ruta_auto if os.path.isfile(ruta_auto) else None)
        tiene_proy = False
        if ruta_out:
            try:
                df_plan = pd.read_excel(ruta_out, sheet_name="Plan_Diario", header=1)
                if "Tipo" in df_plan.columns:
                    df_plan = df_plan[df_plan["Tipo"] == "Proyeccion"]
                if "Codigo PT" in df_plan.columns and "Cantidad a Producir" in df_plan.columns:
                    df_plan["Cantidad a Producir"] = pd.to_numeric(
                        df_plan["Cantidad a Producir"], errors="coerce").fillna(0)
                    proyeccion_dict = (df_plan.groupby("Codigo PT")["Cantidad a Producir"]
                                       .sum().to_dict())
                    tiene_proy = bool(proyeccion_dict)
            except Exception:
                pass

        # ── Calcular por codigo ────────────────────────────────────────────────
        grouped = df_pend.groupby(cod_col).agg(
            descripcion=(desc_col, "first"),
            n_pedidos=(ped_col, "count"),
            saldo_total=(sp_col, "sum"),
        ).reset_index()

        rows = []
        for _, r in grouped.iterrows():
            cod   = str(r[cod_col]).strip()
            desc  = str(r["descripcion"]).strip()
            n_ped = int(r["n_pedidos"])
            saldo = int(r["saldo_total"])
            stock = int(stock_dict.get(cod, 0))
            ops   = int(ops_dict.get(cod, 0))
            proy  = int(proyeccion_dict.get(cod, 0))
            disp  = stock + ops
            total = saldo + proy
            deficit = max(0, total - disp)
            lmin  = int(lmin_dict.get(cod, 0))

            if deficit > 0:
                a_prod = math.ceil(deficit / lmin) * lmin if lmin > 0 else deficit
                n_lot  = a_prod // lmin if lmin > 0 else 1
                estado = "Solicitar OP"
            else:
                a_prod = 0; n_lot = 0; estado = "OK"

            parts = [f"{n_ped} pedido(s), saldo: {saldo:,} uds."]
            if proy > 0:
                parts.append(f"Proy. ML: {proy:,} uds.")
            parts.append(f"Disp: {stock:,} bod + {ops:,} OPs = {disp:,}.")
            if deficit > 0:
                if lmin > 0:
                    parts.append(f"Deficit {deficit:,} -> lote min {lmin:,} -> producir {a_prod:,} ({n_lot} lotes).")
                else:
                    parts.append(f"Deficit {deficit:,} uds. Sin lote min. definido.")
            else:
                parts.append(f"Cubierto. Excedente: {disp - total:,} uds.")

            origins = origen_por_cod.get(cod, {"En firme"})
            rows.append([cod, desc, n_ped,
                         f"{saldo:,}", f"{stock:,}", f"{ops:,}", f"{proy:,}",
                         f"{disp:,}", f"{total:,}", f"{deficit:,}",
                         f"{lmin:,}", f"{a_prod:,}", estado, " ".join(parts),
                         origins])

        self._gestion_data   = rows
        self._gestion_loaded = path_cuadro
        self._gest_nota = ("  •  Con proyeccion ML" if tiene_proy
                           else "  •  Sin proyeccion ML (ejecuta el plan primero)")
        # Pinta tabla + KPIs respetando el filtro activo (busqueda/estado/origen)
        self._filtrar_gestion()


    # ═════════════════════════════════════════════════════════════════════════
    # TAB 6 — Saldo de OPs (cruce OC vs OPs en proceso)
    # ═════════════════════════════════════════════════════════════════════════

    def _build_tab_saldo_ops(self):
        topbar = tk.Frame(self.tab_saldo, bg=C_HEADER, height=48)
        topbar.pack(fill="x"); topbar.pack_propagate(False)
        tk.Label(topbar, text="SALDO REAL DE ORDENES DE PRODUCCION", font=("Segoe UI", 11, "bold"),
                 bg=C_HEADER, fg="white").pack(side="left", padx=24, pady=12)
        self._make_btn(topbar, "  Analizar  ", C_ACENTO, C_ACENTO_H,
                       self._analizar_saldo_ops).pack(side="right", padx=16, pady=8, ipady=4)

        # ── KPI cards ────────────────────────────────────────────────────────
        kpi_row = tk.Frame(self.tab_saldo, bg=C_MAIN)
        kpi_row.pack(fill="x", padx=16, pady=(12, 0))
        self._saldo_kpi = {}
        for i, (key, label, color) in enumerate([
            ("total_ops",      "OPs en proceso",          "#2563EB"),
            ("total_aprobado", "Total aprobado (uds)",    "#7C3AED"),
            ("total_producido","Ya producido (uds)",      "#16A34A"),
            ("total_pendiente","Pendiente por producir",  "#DC2626"),
        ]):
            card = tk.Frame(kpi_row, bg="white",
                            highlightbackground="#E2E8F0", highlightthickness=1)
            card.pack(side="left", fill="both", expand=True, padx=(0 if i == 0 else 8, 0))
            tk.Frame(card, bg=color, height=4).pack(fill="x")
            inner = tk.Frame(card, bg="white")
            inner.pack(fill="x", padx=14, pady=(10, 12))
            val_lbl = tk.Label(inner, text="—", font=("Segoe UI", 20, "bold"),
                               bg="white", fg=color, anchor="w")
            val_lbl.pack(anchor="w")
            tk.Label(inner, text=label, font=("Segoe UI", 8),
                     bg="white", fg=C_GRIS, anchor="w").pack(anchor="w", pady=(3, 0))
            self._saldo_kpi[key] = val_lbl

        # ── Barra de filtro ───────────────────────────────────────────────────
        fbar = tk.Frame(self.tab_saldo, bg=C_MAIN)
        fbar.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(fbar, text="Buscar:", font=("Segoe UI", 9),
                 bg=C_MAIN, fg=C_TEXTO).pack(side="left", padx=(0, 8))
        self._saldo_search = tk.StringVar()
        self._saldo_search.trace_add("write", self._filtrar_saldo_ops)
        tk.Entry(fbar, textvariable=self._saldo_search, font=("Segoe UI", 9), width=30,
                 bg="white", fg=C_TEXTO, relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=C_DIVIDER,
                 highlightcolor=C_ACENTO).pack(side="left", ipady=5, padx=(0, 16))
        self._saldo_lbl = tk.Label(fbar, text="Carga el archivo 04 (OPs en proceso) y presiona Analizar",
                                    font=("Segoe UI", 8), bg=C_MAIN, fg=C_GRIS)
        self._saldo_lbl.pack(side="left")

        # ── Leyenda ───────────────────────────────────────────────────────────
        ley = tk.Frame(self.tab_saldo, bg=C_MAIN)
        ley.pack(fill="x", padx=16, pady=(4, 0))
        for label, bg_c in [("Completada", "#D4EDDA"), ("En Proceso", "#FFF3CD"),
                             ("Sin Iniciar", "#FFE4E4")]:
            f = tk.Frame(ley, bg=bg_c, padx=8, pady=3,
                         highlightbackground=C_DIVIDER, highlightthickness=1)
            f.pack(side="left", padx=(0, 8))
            tk.Label(f, text=label, font=("Segoe UI", 8, "bold"),
                     bg=bg_c, fg=C_TEXTO).pack()

        # ── Treeview ──────────────────────────────────────────────────────────
        tbl = tk.Frame(self.tab_saldo, bg=C_MAIN)
        tbl.pack(fill="both", expand=True, padx=16, pady=10)
        sty = ttk.Style()
        sty.configure("Saldo.Treeview", background="white", foreground=C_TEXTO,
                       fieldbackground="white", rowheight=26, font=("Segoe UI", 9))
        sty.configure("Saldo.Treeview.Heading", background=C_HEADER, foreground="white",
                       font=("Segoe UI", 9, "bold"))
        sty.map("Saldo.Treeview", background=[("selected", C_ACENTO)],
                foreground=[("selected", "white")])

        COLS_S = ["OP", "Tipo", "Cód. Producto", "Referencia",
                  "Aprobado", "Producido", "Pendiente",
                  "Estado", "Fecha Prog.", "Compromiso"]
        WIDTHS_S = {"OP": 80, "Tipo": 130, "Cód. Producto": 130, "Referencia": 300,
                    "Aprobado": 100, "Producido": 100, "Pendiente": 100,
                    "Estado": 110, "Fecha Prog.": 110, "Compromiso": 110}
        CENTERS_S = {"OP", "Tipo", "Aprobado", "Producido", "Pendiente",
                     "Estado", "Fecha Prog.", "Compromiso"}

        self._saldo_tree = ttk.Treeview(tbl, columns=COLS_S, show="headings",
                                         style="Saldo.Treeview")
        for col in COLS_S:
            self._saldo_tree.heading(col, text=col)
            self._saldo_tree.column(col, width=WIDTHS_S.get(col, 100),
                                    anchor="center" if col in CENTERS_S else "w",
                                    minwidth=60)
        vsb = ttk.Scrollbar(tbl, orient="vertical",   command=self._saldo_tree.yview)
        hsb = ttk.Scrollbar(tbl, orient="horizontal", command=self._saldo_tree.xview)
        self._saldo_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._saldo_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tbl.rowconfigure(0, weight=1); tbl.columnconfigure(0, weight=1)

    def _refrescar_saldo_tree(self, data):
        self._saldo_tree.delete(*self._saldo_tree.get_children())
        TAG_BG = {
            "Completada":  "#D4EDDA",
            "En Proceso":  "#FFF3CD",
            "Sin Iniciar": "#FFE4E4",
        }
        for row in data:
            estado = row[7] if len(row) > 7 else ""
            tag = estado.replace(" ", "_")
            self._saldo_tree.insert("", "end", values=row, tags=(tag,))
        for estado, bg in TAG_BG.items():
            self._saldo_tree.tag_configure(estado.replace(" ", "_"), background=bg)

    def _filtrar_saldo_ops(self, *_):
        q = self._saldo_search.get().lower()
        filtered = [r for r in self._saldo_ops_data
                    if not q or any(q in str(v).lower() for v in r)]
        self._refrescar_saldo_tree(filtered)
        self._saldo_lbl.config(
            text=f"{len(filtered):,} OPs" +
                 (f"  (de {len(self._saldo_ops_data):,})" if len(filtered) != len(self._saldo_ops_data) else ""))

    def _analizar_saldo_ops(self):
        path_ops      = self.file_vars["arch_ops_proc"].get().strip()
        path_entradas = self.file_vars["arch_entradas"].get().strip()

        if not path_ops or not os.path.isfile(path_ops):
            self._saldo_lbl.config(
                text="Carga el archivo 04 (OPs en proceso) para ver el analisis.")
            return

        try:
            import pandas as pd
        except ImportError as exc:
            messagebox.showerror("Falta pandas", str(exc)); return

        # ── OPs en proceso ────────────────────────────────────────────────────
        try:
            df_ops = pd.read_excel(path_ops)
            df_ops["Cant. Aprobada"] = pd.to_numeric(
                df_ops.get("Cant. Aprobada", pd.Series(dtype=float)),
                errors="coerce").fillna(0)
            df_ops["Fecha Programada"]   = pd.to_datetime(df_ops.get("Fecha Programada"),   errors="coerce")
            df_ops["Compromiso Cliente"] = pd.to_datetime(df_ops.get("Compromiso Cliente"), errors="coerce")
            cod_col = next((c for c in df_ops.columns
                            if "producto" in c.lower() and ("cód" in c.lower() or "cod" in c.lower())), None)
        except Exception as exc:
            messagebox.showerror("Error leyendo OPs", str(exc)); return

        # ── Entradas de inventario (opcional) ────────────────────────────────
        producido_dict = {}
        tiene_entradas = False
        if path_entradas and os.path.isfile(path_entradas):
            try:
                df_ent = pd.read_excel(path_entradas)
                oc_col  = next((c for c in df_ent.columns if c.upper().strip() == "OC"), None)
                qty_col = next((c for c in df_ent.columns
                                if c.lower().strip() in ("cantidad", "cant.", "cant. ingresada")), None)
                if oc_col and qty_col:
                    df_ent[oc_col]  = pd.to_numeric(df_ent[oc_col],  errors="coerce")
                    df_ent[qty_col] = pd.to_numeric(df_ent[qty_col], errors="coerce").fillna(0)
                    producido_dict = df_ent.groupby(oc_col)[qty_col].sum().to_dict()
                    tiene_entradas = bool(producido_dict)
            except Exception:
                pass

        # ── Calcular por OP ───────────────────────────────────────────────────
        rows = []
        for _, r in df_ops.iterrows():
            op_num   = r.get("OP", "")
            tipo     = str(r.get("Tipo de Trabajo", ""))
            cod      = str(r.get(cod_col, "")) if cod_col else ""
            ref      = str(r.get("Referencia", ""))
            aprobada = float(r.get("Cant. Aprobada", 0) or 0)
            if aprobada == 0:
                continue
            fp       = r.get("Fecha Programada")
            cc       = r.get("Compromiso Cliente")

            try:
                op_key = float(op_num)
            except (ValueError, TypeError):
                op_key = None
            producida = float(producido_dict.get(op_key, 0)) if op_key is not None else 0
            pendiente = max(0.0, aprobada - producida)
            pct       = round(producida / aprobada * 100, 1) if aprobada > 0 else 0.0

            if pct >= 100:
                estado = "Completada"
            elif producida > 0:
                estado = "En Proceso"
            else:
                estado = "Sin Iniciar"

            fp_str = fp.strftime("%Y-%m-%d") if pd.notna(fp) else ""
            cc_str = cc.strftime("%Y-%m-%d") if pd.notna(cc) else ""

            rows.append([op_num, tipo, cod, ref,
                         f"{aprobada:,.0f}", f"{producida:,.0f}", f"{pendiente:,.0f}",
                         estado, fp_str, cc_str])

        self._saldo_ops_data = rows
        self._refrescar_saldo_tree(rows)

        # ── KPIs ──────────────────────────────────────────────────────────────
        def _n(s): return float(str(s).replace(",", "").replace("%", "") or 0)
        total_aprobado  = sum(_n(r[4]) for r in rows)
        total_producido = sum(_n(r[5]) for r in rows)
        total_pendiente = sum(_n(r[6]) for r in rows)
        self._saldo_kpi["total_ops"].config(text=f"{len(rows):,}")
        self._saldo_kpi["total_aprobado"].config(text=f"{total_aprobado:,.0f}")
        self._saldo_kpi["total_producido"].config(text=f"{total_producido:,.0f}")
        self._saldo_kpi["total_pendiente"].config(text=f"{total_pendiente:,.0f}")

        if tiene_entradas:
            nota = "  •  Con entradas de inventario (OC cruzado)"
        else:
            nota = "  •  Sin archivo de entradas — producido = 0 para todas las OPs"
        self._saldo_lbl.config(text=f"{len(rows):,} OPs en proceso{nota}")


if __name__ == "__main__":
    SeedPackPlanner()

