# -*- coding: utf-8 -*-
"""
PASO 5 - Plan de Produccion Diario - SEEDPACK
==============================================
Genera un plan de produccion diario basado en:
  - Proyeccion de demanda mensual (paso 4 - ML)
  - Inventario actual en bodega   (Bodega.xlsx)
  - Ordenes de produccion en proceso
  - Pedidos pendientes (referencias ocasionales)

Parametros (editar en el bloque __main__):
  FI  : fecha inicio  ej. "2026-06-01"  (por defecto: primer dia del mes actual)
  FF  : fecha fin     ej. "2026-12-31"  (por defecto: ultimo dia del año actual)
  MOD : modelo       "Random Forest" | "XGBoost" | "Reg. Regularizada" | "Red Neuronal"
  REF : producto     nombre exacto, o None para todos

Salida:
  resultados/paso5_orden_produccion.xlsx
    Hojas:
      Plan_Diario          - que producir por referencia cada dia habil
      Calendario_Semanal   - pivot referencia x semana
      OPs_En_Proceso       - ordenes ya en fabricacion
      Pedidos_Ocasionales  - pedidos de referencias fuera de la proyeccion regular
      Inventario_Actual    - stock activo en bodega
      Dashboard            - KPIs resumen
"""

import math
import os
import re
import sys
import warnings
from collections import defaultdict

import holidays as _holidays_co
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Rutas ─────────────────────────────────────────────────────────────────────
_PASO5_DIR = os.path.dirname(os.path.abspath(__file__))
BASE       = ((os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
               else os.path.join(os.path.dirname(os.path.dirname(_PASO5_DIR)), "app")) + "/")
RUTA_AH   = BASE + "Archivos Historicos/"
RUTA_RES  = BASE + "Resultado Final/"

ARCH_BODEGA    = RUTA_AH + "Bodega.xlsx"
ARCH_OPS_PROC  = RUTA_AH + "ordenes de produccion en proceso.xlsx"
ARCH_PEDIDOS   = RUTA_AH + "INEDITTO_Listado_Pedidos.xlsx"
ARCH_OPS_HIST  = RUTA_AH + "INEDITTO_OP.xlsx"
ARCH_VENTAS    = RUTA_AH + "ventasXproducto.xlsx"
ARCH_LOTE_MIN  = RUTA_AH + "Codigos lote minimo.xlsx"
ARCH_ENTRADAS        = None  # opcional: entradas de inventario para cruce OC vs OPs
ARCH_PLAN_COMERCIAL  = None  # opcional: plan comercial para comparativa vs ML


def _aplicar_mapeos_usuario(df, clave):
    """Renombra columnas segun el mapeo guardado por el usuario en column_mappings.json."""
    import json as _json
    _path = os.path.normpath(BASE + "column_mappings.json")
    if not os.path.isfile(_path):
        return df
    try:
        with open(_path, encoding="utf-8") as _f:
            _m = _json.load(_f).get(clave, {})
        # Normalizar con strip para que espacios en el Excel no rompan el matching
        _cols = {str(c).strip(): c for c in df.columns}
        _rename = {_cols[str(v).strip()]: k
                   for k, v in _m.items()
                   if v and str(v).strip() in _cols and _cols[str(v).strip()] != str(k)}
        return df.rename(columns=_rename) if _rename else df
    except Exception:
        return df

# ── Configuracion ─────────────────────────────────────────────────────────────
LEAD_TIME_DIAS = 3                        # dias habiles de anticipacion para ocasionales
DIAS_SEMANA    = "Mon Tue Wed Thu Fri Sat" # dias laborables (incluye sabado)
BODEGA_EXCLUIR = ["Obsoletos"]            # bodegas que NO cuentan como stock util
SKIP_LINEAS    = {"FLETE VENTAS", "RETAL", "SERVICIO"} # lineas sin plan de produccion
# Palabras clave en el nombre del producto que lo excluyen del plan de produccion
# (servicios, fletes, ventas de retal, obsoletos marcados con "NO USAR")
SKIP_NOMBRE_KW = ("NO USAR", "FLETE", "SERVICIO", "VENTA RETAL", "VTARTL")

# ── Paleta de colores ─────────────────────────────────────────────────────────
C_AZUL_OSC = "1F4E79"
C_AZUL_MED = "2E75B6"
C_ALT      = "D6E4F0"
C_BLANCO   = "FFFFFF"
C_VERDE    = "70AD47"
C_AMARILLO = "FFD966"
C_ROJO     = "FF5050"
C_NARANJA  = "ED7D31"
C_MORADO   = "7030A0"
C_MORADO_CLR = "E2D0F0"

# ── Helpers de estilo openpyxl ────────────────────────────────────────────────
_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _fill(c):
    return PatternFill(fill_type="solid", fgColor=c)


def hdr(ws, row, col, val, bg=C_AZUL_OSC, fc="FFFFFF", sz=11, bold=True, wrap=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Calibri", bold=bold, color=fc, size=sz)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _border
    return cell


def dat(ws, row, col, val, bg=C_BLANCO, fmt=None, bold=False, fc="000000", align="left"):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Calibri", size=10, bold=bold, color=fc)
    cell.fill = _fill(bg)
    cell.border = _border
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


# ── Helpers de fechas ─────────────────────────────────────────────────────────

def _dias_habiles(start, end):
    """Lista de Timestamps de dias habiles (lun-sab) excluyendo festivos colombianos."""
    s = pd.Timestamp(start)
    e = pd.Timestamp(end)
    festivos = list(_holidays_co.Colombia(years=range(s.year, e.year + 1)).keys())
    return list(pd.bdate_range(start=s, end=e, freq="C", weekmask=DIAS_SEMANA, holidays=festivos))


_MESES_ES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
             7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

def _label_semana(ts):
    """Etiqueta legible para la semana. Cuando cruza mes o año muestra ambos."""
    d      = pd.Timestamp(ts)
    inicio = d - pd.Timedelta(days=d.weekday())  # lunes de esa semana
    fin    = inicio + pd.Timedelta(days=5)        # sabado
    if inicio.month == fin.month:
        return f"{inicio.day}-{fin.day} {_MESES_ES[inicio.month]} {inicio.year}"
    if inicio.year == fin.year:
        return f"{inicio.day} {_MESES_ES[inicio.month]} - {fin.day} {_MESES_ES[fin.month]} {inicio.year}"
    return f"{inicio.day} {_MESES_ES[inicio.month]} {inicio.year} - {fin.day} {_MESES_ES[fin.month]} {fin.year}"


# ── Carga de datos ────────────────────────────────────────────────────────────

def _normalizar_bodega(df):
    """Homogeniza columnas del archivo de bodega independientemente del formato exportado."""
    df   = _aplicar_mapeos_usuario(df, "arch_bodega")
    cols = {c.lower().strip(): c for c in df.columns}

    # Columna bodega (filtro de obsoletos)
    if "bodega" not in cols:
        for alias in ("bodega sale", "bodega_sale", "almacen"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Bodega"})
                break

    # Columna código PT
    if not any("pt" in c.lower() and ("cód" in c.lower() or "cod" in c.lower()) for c in df.columns):
        for alias in ("prod. terminado", "prod.terminado", "producto terminado"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Cód. PT"})
                df["Cód. PT"] = df["Cód. PT"].astype(str).str.split(" - ").str[0].str.strip()
                break
    else:
        for c in df.columns:
            if "pt" in c.lower() and ("cód" in c.lower() or "cod" in c.lower()):
                df = df.rename(columns={c: "Cód. PT"})
                break

    # Columna saldo (ExistBodega tiene "Saldo" real; EBPT tiene "Saldo en Inventario")
    cols = {c.lower().strip(): c for c in df.columns}
    if "saldo" not in cols:
        for alias in ("saldo en inventario", "cantidad", "cant.", "stock"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Saldo"})
                break

    # Cant. Ingresada (solo ExistBodega; entradasInventario no la tiene)
    cols = {c.lower().strip(): c for c in df.columns}
    if "cant. ingresada" not in cols and "cant ingresada" not in cols:
        for alias in ("cantidad ingresada",):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Cant. Ingresada"})
                break

    # Costo unitario
    cols = {c.lower().strip(): c for c in df.columns}
    if "costo" not in cols:
        for alias in ("vr. unitario", "vr unitario", "precio unitario"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Costo"})
                break

    # Total valor
    cols = {c.lower().strip(): c for c in df.columns}
    if "total" not in cols:
        for alias in ("vr. total", "vr total", "valor total"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Total"})
                break

    # Referencia / descripción del producto
    cols = {c.lower().strip(): c for c in df.columns}
    if "referencia" not in cols:
        for alias in ("referenciapt", "descripcion", "descripción", "producto", "nombre"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "Referencia"})
                break

    # OP/OC: ExistBodega tiene "O.P. / O.C.", entradasInventario tiene "OC"
    cols = {c.lower().strip(): c for c in df.columns}
    if "op/oc" not in cols:
        for alias in ("o.p. / o.c.", "op / oc", "op/oc", "oc"):
            if alias in cols:
                df = df.rename(columns={cols[alias]: "OP/OC"})
                break

    return df


def _cargar_bodega():
    """dict codigo_pt -> saldo (solo bodegas activas, sin Obsoletos)."""
    df = _normalizar_bodega(pd.read_excel(ARCH_BODEGA))
    df = df[~df["Bodega"].isin(BODEGA_EXCLUIR)].copy()
    df["Saldo"] = pd.to_numeric(df["Saldo"], errors="coerce").fillna(0)
    return df.groupby("Cód. PT")["Saldo"].sum().to_dict()


def _cargar_bodega_df():
    """DataFrame completo de bodega (no-obsoletos) para la hoja Inventario."""
    df = _normalizar_bodega(pd.read_excel(ARCH_BODEGA))
    return df[~df["Bodega"].isin(BODEGA_EXCLUIR)].copy()


def _cargar_ops_proceso():
    """
    Retorna (DataFrame de OPs en proceso,
             dict (codigo_pt, año, mes) -> cantidad pendiente total).

    Usa Cant. Pendiente = max(0, Cant. Aprobada - Cant. Ingresada) cuando
    la columna Cant. Ingresada existe en el archivo; de lo contrario usa
    Cant. Aprobada completa como estimado conservador.
    """
    df = pd.read_excel(ARCH_OPS_PROC)
    df = _aplicar_mapeos_usuario(df, "arch_ops_proc")
    df["Fecha Programada"]   = pd.to_datetime(df["Fecha Programada"],   errors="coerce")
    df["Compromiso Cliente"] = pd.to_datetime(df["Compromiso Cliente"], errors="coerce")
    df["Cant. Aprobada"]     = pd.to_numeric(df["Cant. Aprobada"],     errors="coerce").fillna(0)

    # Detectar columna de cantidad ya producida/ingresada
    _ing_aliases = ("cant. ingresada", "cant ingresada", "cantidad ingresada",
                    "cant. producida", "cant producida", "cantidad producida")
    _ing_col = next(
        (c for c in df.columns if c.lower().strip() in _ing_aliases), None
    )
    if _ing_col:
        df["Cant. Ingresada"] = pd.to_numeric(df[_ing_col], errors="coerce").fillna(0)
    else:
        df["Cant. Ingresada"] = 0.0

    # Pendiente real: lo que falta por producir en cada OP
    df["Cant. Pendiente"] = (df["Cant. Aprobada"] - df["Cant. Ingresada"]).clip(lower=0)

    ops_supply = {}
    for _, row in df.iterrows():
        fp     = row["Fecha Programada"]
        codigo = str(row["Cód. Producto"]).strip() if pd.notna(row["Cód. Producto"]) else ""
        if pd.isna(fp) or not codigo:
            continue
        key = (codigo, fp.year, fp.month)
        ops_supply[key] = ops_supply.get(key, 0.0) + float(row["Cant. Pendiente"])

    return df, ops_supply


def _build_name_code_map():
    """Mapeos bidireccionales nombre<->codigo desde ventas historicas + OPs."""
    _v = pd.read_excel(ARCH_VENTAS)
    _v = _aplicar_mapeos_usuario(_v, "arch_ventas")
    _prod_col = "Producto Terminado" if "Producto Terminado" in _v.columns else "Producto"
    ventas = _v[["Código", _prod_col]].dropna().rename(columns={_prod_col: "Producto"})
    freq   = (ventas.groupby(["Producto", "Código"]).size()
                    .reset_index(name="n").sort_values("n", ascending=False))
    name_to_code = freq.groupby("Producto")["Código"].first().to_dict()
    code_to_name = (ventas.drop_duplicates("Código")
                          .set_index("Código")["Producto"].to_dict())
    # Complementar con OPs historico (codigos sin nombre en ventas)
    try:
        ops_h = pd.read_excel(ARCH_OPS_HIST)
        ops_h = _aplicar_mapeos_usuario(ops_h, "arch_ops_hist")
        ops_h = ops_h[["Cód. Producto", "Referencia"]].dropna()
        for _, row in ops_h.iterrows():
            c = str(row["Cód. Producto"]).strip()
            n = str(row["Referencia"]).strip()
            if c and c not in code_to_name:
                code_to_name[c] = n
    except Exception:
        pass
    return name_to_code, code_to_name


def _infer_batch_sizes():
    """Mediana del tamano de lote por codigo_pt desde historico de OPs."""
    try:
        ops = pd.read_excel(ARCH_OPS_HIST)
        ops = _aplicar_mapeos_usuario(ops, "arch_ops_hist")
        ops["Cant. Aprobada"] = pd.to_numeric(ops["Cant. Aprobada"], errors="coerce")
        ops = ops.dropna(subset=["Cant. Aprobada", "Cód. Producto"])
        ops = ops[ops["Cant. Aprobada"] > 0]
        return (ops.groupby("Cód. Producto")["Cant. Aprobada"]
                   .median().round().astype(int).to_dict())
    except Exception:
        return {}


def _normalizar_pedidos(df):
    """Normaliza columnas del archivo de pedidos al formato estandar interno."""
    cols = {c.strip().lower(): c for c in df.columns}

    if "fecha de entrega" not in cols:
        rename = {}
        for destino, candidatos in [
            ("Fecha de entrega", ["fecha de entrega lead time", "fecha entrega comercial",
                                  "fecha de entrega"]),
            ("Código PT",        ["codigo"]),
            ("Referencia",       ["descripcion ", "descripcion"]),
            ("Cantidad",         ["saldo pendiente", "cantidad"]),
            ("Pedido",           ["pedido"]),
            ("Cliente",          ["cliente"]),
            ("Línea",            ["linea"]),
        ]:
            for cand in candidatos:
                if cand in cols:
                    rename[cols[cand]] = destino
                    break

        df = df.rename(columns=rename)

        # Si no hay Código PT usar Referencia como clave (productos ocasionales sin codigo)
        if "Código PT" not in df.columns and "Referencia" in df.columns:
            df["Código PT"] = df["Referencia"]

    # Siempre garantizar columnas de filtro (pueden faltar en formatos historicos)
    for col in ("Remisión", "Factura"):
        if col not in df.columns:
            df[col] = np.nan

    return df


def _cargar_pedidos(FI, FF):
    """Pedidos pendientes activos (saldo > 0) con entrega hasta FF.
    Incluye pedidos vencidos (fecha < FI) para que aparezcan en Pedidos_Sin_Cubrir.
    Soporta tanto el formato original como el nuevo INEDITTO_Listado_Pedidos.xlsx.
    """
    xl = pd.ExcelFile(ARCH_PEDIDOS)
    sheets = xl.sheet_names

    df = None
    # Maestro es la hoja maestra con todos los pedidos activos (header en fila 4)
    if "Maestro" in sheets:
        try:
            df_m = pd.read_excel(xl, sheet_name="Maestro", header=3)
            if "PEDIDO" in df_m.columns or "Pedido" in df_m.columns:
                df = df_m
        except Exception:
            pass

    if df is None:
        df = pd.read_excel(xl)   # fallback: primera hoja / formato original

    df = _aplicar_mapeos_usuario(df, "arch_pedidos")
    df = _normalizar_pedidos(df)
    df["Fecha de entrega"] = pd.to_datetime(df["Fecha de entrega"], errors="coerce")
    df["Cantidad"]         = pd.to_numeric(df["Cantidad"], errors="coerce").fillna(0)

    ff = pd.Timestamp(FF)
    mask = (
        df["Remisión"].isna()
        & df["Factura"].isna()
        & (df["Fecha de entrega"] <= ff)
        & ~df["Línea"].isin(SKIP_LINEAS)
        & (df["Cantidad"] > 0)
    )
    return df[mask].copy()


# ── Lotes minimos de produccion ──────────────────────────────────────────────

def _cargar_lote_minimo():
    """
    Carga Codigos lote minimo.xlsx.
    Retorna dict codigo_pt -> lote_minimo (int).
    El archivo tiene 2 filas de encabezado (1 vacia + 1 con titulos).
    """
    df = pd.read_excel(ARCH_LOTE_MIN, skiprows=2, header=None)
    df.columns = ["codigo", "referencia", "linea", "cavidades", "lote_minimo"]
    df = df.dropna(subset=["codigo"])
    df["lote_minimo"] = pd.to_numeric(df["lote_minimo"], errors="coerce").fillna(0).astype(int)
    df = df[df["lote_minimo"] > 0]
    return df.set_index("codigo")["lote_minimo"].to_dict()


_MESES_ES_INV = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}

def _cargar_plan_comercial(path):
    """
    Lee el Plan Comercial detectando dinamicamente hoja, header y columnas de meses.

    Tolerancias:
    - Hoja: prueba nombres conocidos y cae al primer sheet con datos.
    - Header: prueba filas 0-3 hasta encontrar una con columnas utiles.
    - Meses: solo toma columnas de CANTIDAD (descarta las de Valor/Precio).
    - Nombre producto: acepta "producto terminado", "producto", "referencia", "descripcion".

    Retorna (DataFrame: codigo/año/mes/cantidad_comercial, dict codigo->nombre).
    """
    xl = pd.ExcelFile(path)
    sheets = xl.sheet_names

    # Intentar hojas conocidas primero, luego cualquier hoja con datos
    _HOJAS_CONOCIDAS = ["Plan comercial UNDS", "Plan Comercial UNDS",
                        "plan comercial unds", "PLAN COMERCIAL UNDS"]
    _HOJAS_PRIORIDAD = [s for s in _HOJAS_CONOCIDAS if s in sheets]
    _HOJAS_BUSQUEDA  = _HOJAS_PRIORIDAD + [s for s in sheets if s not in _HOJAS_PRIORIDAD]

    df = None
    for sheet in _HOJAS_BUSQUEDA:
        for hdr_row in [0, 1, 2, 3]:
            try:
                tmp = pd.read_excel(xl, sheet_name=sheet, header=hdr_row)
                # Necesitamos al menos: codigo + nombre + 1 columna de mes
                cols_lower = [str(c).lower() for c in tmp.columns]
                tiene_mes  = any(mes in " ".join(cols_lower) for mes in _MESES_ES_INV)
                tiene_cod  = any(c in ("código", "codigo", "cód.", "cod.") for c in cols_lower)
                if tiene_mes and tiene_cod and len(tmp) > 0:
                    df = tmp
                    break
            except Exception:
                pass
        if df is not None:
            break

    if df is None:
        return pd.DataFrame(columns=["codigo", "año", "mes", "cantidad_comercial"]), {}

    # Renombrar primera columna a "codigo"
    df.rename(columns={df.columns[0]: "codigo"}, inplace=True)
    df["codigo"] = df["codigo"].astype(str).str.strip()
    df = df[df["codigo"].str.len() > 0].copy()

    # Mapa codigo -> nombre del producto
    _NOMBRE_ALIAS = ("producto terminado", "producto", "referencia",
                     "descripcion", "descripción", "nombre")
    nombre_col = next(
        (c for c in df.columns if str(c).lower().strip() in _NOMBRE_ALIAS), None
    )
    nombre_map = {}
    if nombre_col:
        nombre_map = (df[["codigo", nombre_col]]
                      .dropna(subset=[nombre_col])
                      .drop_duplicates("codigo")
                      .set_index("codigo")[nombre_col]
                      .astype(str).str.strip().to_dict())

    # Detectar columnas de meses — solo CANTIDAD (excluir Valor/Precio/$)
    _EXCLUIR_KW = ("valor", "vr.", "precio", "price", "$", "total", "promedio", "prom")
    month_cols = {}
    for col in df.columns:
        col_lower = str(col).lower()
        # Descartar columnas de valor o promedios
        if any(kw in col_lower for kw in _EXCLUIR_KW):
            continue
        for mes_es, mes_num in _MESES_ES_INV.items():
            if mes_es in col_lower:
                yr_match = re.search(r"\b(20\d{2})\b", str(col))
                yr = int(yr_match.group(1)) if yr_match else pd.Timestamp.today().year
                month_cols[col] = (yr, mes_num)
                break

    if not month_cols:
        return pd.DataFrame(columns=["codigo", "año", "mes", "cantidad_comercial"]), nombre_map

    rows = []
    for col, (yr, mes) in month_cols.items():
        sub = df[["codigo", col]].copy()
        sub["año"] = yr
        sub["mes"] = mes
        sub["cantidad_comercial"] = pd.to_numeric(sub[col], errors="coerce").fillna(0)
        rows.append(sub[["codigo", "año", "mes", "cantidad_comercial"]])

    result = pd.concat(rows, ignore_index=True)
    agg = (result.groupby(["codigo", "año", "mes"])["cantidad_comercial"]
                 .sum().round().astype(int).reset_index())
    return agg, nombre_map


def _cargar_codigos_stock():
    """Retorna frozenset de codigos Genericos: productos que deben mantenerse en stock."""
    df = pd.read_excel(ARCH_LOTE_MIN, skiprows=2, header=None)
    df.columns = ["codigo", "referencia", "linea", "cavidades", "lote_minimo"]
    df = df.dropna(subset=["codigo"])
    df["linea"] = df["linea"].str.strip()
    return frozenset(df.loc[df["linea"] == "Genérico", "codigo"].str.strip())


def _redondear_a_lote(cantidad, lote):
    """
    Redondea `cantidad` hacia arriba al multiplo mas cercano de `lote`.
    Si cantidad==0 o lote==0 retorna la cantidad sin cambios.
    """
    if cantidad <= 0 or lote <= 0:
        return cantidad
    return math.ceil(cantidad / lote) * lote


def _aplicar_lote_minimo(mrp_df, lote_min):
    """
    Garantiza que mrp_df tiene las columnas 'lote_minimo' y 'neto_sin_ajuste'.
    Cuando _mrp_mensual recibio lote_min (caso normal), el ajuste ya esta hecho
    y el rollforward de stock es coherente; esta funcion solo completa columnas
    faltantes para compatibilidad con llamadas externas.
    """
    df = mrp_df.copy()
    if "lote_minimo" not in df.columns:
        df["lote_minimo"] = df["codigo_pt"].map(lote_min).fillna(0).astype(int)
    if "neto_sin_ajuste" not in df.columns:
        # Columna ausente: el ajuste no se habia aplicado aun — aplicar ahora.
        df["neto_sin_ajuste"] = df["neto_producir"]
        df["neto_producir"]   = df.apply(
            lambda r: _redondear_a_lote(r["neto_producir"], r["lote_minimo"]), axis=1
        )
    return df


# ── MRP Mensual ───────────────────────────────────────────────────────────────

def _mrp_mensual(proj_df, stock, ops_supply, name_to_code, lote_min=None):
    """
    Calcula cantidad neta a producir por (codigo_pt, año, mes).
    El inventario se consume mes a mes (rollforward).

    stock_cierre(mes N) = stock_inicio(mes N+1)
    neto_sin_aj = max(0, demanda - stock_inicio - ops_que_completan_ese_mes)
    neto_producir = ceil(neto_sin_aj / lote_min) * lote_min   <- ajuste incluido
    stock_cierre  = max(0, stock_inicio + ops + neto_producir - demanda)

    El exceso de produccion por lote minimo se acredita como inventario
    disponible para el mes siguiente, evitando sobre-produccion acumulada.
    """
    _lote = lote_min or {}

    p = proj_df.copy()
    p["codigo_pt"] = p["producto"].map(name_to_code)
    p = p.dropna(subset=["codigo_pt"])

    # Agregar por codigo (varios nombres pueden mapear al mismo SKU)
    agg = (p.groupby(["codigo_pt", "año", "mes", "periodo"])
            .agg(demanda=("cantidad_proyectada", "sum"),
                 producto=("producto", "first"))
            .reset_index()
            .sort_values(["codigo_pt", "año", "mes"]))

    stock_act = dict(stock)  # copia local para el rollforward
    resultados = []

    for _, row in agg.iterrows():
        codigo  = row["codigo_pt"]
        año     = int(row["año"])
        mes     = int(row["mes"])
        demanda = float(row["demanda"])

        inv          = stock_act.get(codigo, 0.0)
        ops_mes      = ops_supply.get((codigo, año, mes), 0.0)
        neto_sin_aj  = max(0.0, demanda - inv - ops_mes)
        lm           = _lote.get(codigo, 0)
        neto_prod    = _redondear_a_lote(neto_sin_aj, lm)
        # stock_cierre incluye el exceso generado por el redondeo de lote
        cierre       = max(0.0, inv + ops_mes + neto_prod - demanda)
        stock_act[codigo] = cierre

        resultados.append({
            "codigo_pt":          codigo,
            "producto":           row["producto"],
            "año":                año,
            "mes":                mes,
            "periodo":            row["periodo"],
            "demanda_proyectada": round(demanda),
            "stock_inicio":       round(inv),
            "ops_proceso_mes":    round(ops_mes),
            "neto_sin_ajuste":    round(neto_sin_aj),
            "lote_minimo":        lm,
            "neto_producir":      round(neto_prod),
            "stock_cierre":       round(cierre),
        })

    return pd.DataFrame(resultados)


# ── Distribucion diaria ───────────────────────────────────────────────────────

def _distribuir_diario(mrp_df, FI, FF, batch_sizes):
    """
    Distribuye la produccion mensual neta en dias habiles del mes.
    Si hay tamano de lote historico, agrupa en lotes; si no, distribuye uniforme.
    """
    fi, ff = pd.Timestamp(FI), pd.Timestamp(FF)
    rows = []

    for _, row in mrp_df.iterrows():
        if row["neto_producir"] <= 0:
            continue

        codigo = row["codigo_pt"]
        neto   = int(row["neto_producir"])
        año    = row["año"]
        mes    = row["mes"]

        mes_ini = max(fi, pd.Timestamp(int(año), int(mes), 1))
        mes_fin = min(ff, pd.Timestamp(int(año), int(mes), 1) + pd.offsets.MonthEnd(0))
        dias    = _dias_habiles(mes_ini, mes_fin)

        if not dias:
            continue

        batch = batch_sizes.get(codigo, 0)

        if batch > 0 and neto > batch:
            # Distribuir en lotes espaciados
            n_lotes   = max(1, round(neto / batch))
            step      = max(1, len(dias) // n_lotes)
            dias_lote = dias[::step][:n_lotes]
            qty       = round(neto / len(dias_lote))
            for d in dias_lote:
                rows.append({"fecha": d.date(), "codigo_pt": codigo,
                             "referencia": row["producto"], "cantidad": qty,
                             "tipo": "Proyeccion", "fecha_entrega": None})
        else:
            # Distribucion uniforme con residuo en los primeros dias
            qty_base = neto // len(dias)
            residuo  = neto % len(dias)
            for i, d in enumerate(dias):
                q = qty_base + (1 if i < residuo else 0)
                if q > 0:
                    rows.append({"fecha": d.date(), "codigo_pt": codigo,
                                 "referencia": row["producto"], "cantidad": q,
                                 "tipo": "Proyeccion", "fecha_entrega": None})

    return (pd.DataFrame(rows) if rows
            else pd.DataFrame(columns=["fecha", "codigo_pt", "referencia",
                                        "cantidad", "tipo", "fecha_entrega"]))


# ── Pedidos Ocasionales ───────────────────────────────────────────────────────

def _planificar_ocasionales(pedidos_df, stock, ops_supply, proj_codigos,
                             batch_sizes, lote_min, FI, FF):
    """
    Planifica produccion para pedidos cuyo codigo NO esta en la proyeccion regular.
    Aplica lotes minimos de produccion si el codigo los tiene definidos.
    Retorna (plan_diario_df, tabla_resumen_df).
    """
    fi = pd.Timestamp(FI)

    # Stock total por codigo (suma sobre todos los meses de ops_supply)
    ops_total = defaultdict(float)
    for (cod, _a, _m), q in ops_supply.items():
        ops_total[cod] += q

    stock_ocas = dict(stock)  # copia independiente del stock

    ocas     = pedidos_df[~pedidos_df["Código PT"].isin(proj_codigos)].copy()
    ocas     = ocas.sort_values("Fecha de entrega")  # procesar por fecha de entrega

    plan_rows  = []
    tabla_rows = []

    for _, ped in ocas.iterrows():
        codigo     = str(ped["Código PT"]).strip()
        referencia = str(ped["Referencia"]).strip()
        qty_ped    = float(ped["Cantidad"])
        f_entrega  = pd.Timestamp(ped["Fecha de entrega"])

        inv     = stock_ocas.get(codigo, 0.0)
        ops_q   = ops_total.get(codigo, 0.0)
        neto    = max(0.0, qty_ped - inv - ops_q)

        # Consumir stock y OPs para este pedido
        stock_ocas[codigo] = max(0.0, inv - qty_ped)
        ops_total[codigo]  = max(0.0, ops_q - max(0.0, qty_ped - inv))

        # Aplicar lote minimo al neto
        lote      = lote_min.get(codigo, 0)
        neto_aj   = _redondear_a_lote(neto, lote)

        tabla_rows.append({
            "Pedido":              ped["Pedido"],
            "Código PT":           codigo,
            "Referencia":          referencia,
            "Cliente":             ped.get("Cliente", ""),
            "Fecha Entrega":       f_entrega.date() if pd.notna(f_entrega) else None,
            "Cantidad Pedida":     round(qty_ped),
            "Stock Disponible":    round(inv),
            "OPs Proceso":         round(ops_q),
            "Neto sin Ajuste":     round(neto),
            "Lote Minimo":         lote,
            "Neto a Producir":     round(neto_aj),
        })

        if neto_aj <= 0:
            continue

        # Fecha inicio produccion = entrega menos LEAD_TIME dias habiles
        dias_hasta_entrega = _dias_habiles(fi, f_entrega)
        if len(dias_hasta_entrega) > LEAD_TIME_DIAS:
            prod_start = dias_hasta_entrega[-LEAD_TIME_DIAS - 1]
        else:
            prod_start = dias_hasta_entrega[0] if dias_hasta_entrega else fi
        prod_start = max(prod_start, fi)

        batch = batch_sizes.get(codigo, 0)
        if batch > 0 and neto_aj > batch:
            n_lotes   = max(1, round(neto_aj / batch))
            dias_prod = _dias_habiles(fi, prod_start)
            if not dias_prod:
                dias_prod = [prod_start]
            step      = max(1, len(dias_prod) // n_lotes)
            dias_lote = dias_prod[::step][:n_lotes]
            qty       = round(neto_aj / len(dias_lote))
            for d in dias_lote:
                plan_rows.append({"fecha": d.date(), "codigo_pt": codigo,
                                  "referencia": referencia, "cantidad": qty,
                                  "tipo": "Ocasional",
                                  "fecha_entrega": f_entrega.date()})
        else:
            plan_rows.append({"fecha": prod_start.date(), "codigo_pt": codigo,
                               "referencia": referencia, "cantidad": round(neto_aj),
                               "tipo": "Ocasional",
                               "fecha_entrega": f_entrega.date()})

    plan_df  = (pd.DataFrame(plan_rows) if plan_rows
                else pd.DataFrame(columns=["fecha", "codigo_pt", "referencia",
                                            "cantidad", "tipo", "fecha_entrega"]))
    tabla_df = pd.DataFrame(tabla_rows) if tabla_rows else pd.DataFrame()
    return plan_df, tabla_df


# ══════════════════════════════════════════════════════════════════════════════
# ESCRITURA EXCEL - una funcion por hoja
# ══════════════════════════════════════════════════════════════════════════════

def _hoja_plan_diario(wb, plan_df):
    ws = wb.create_sheet("Plan_Diario")

    ws.merge_cells("A1:F1")
    ws["A1"] = "SEEDPACK - PLAN DE PRODUCCION DIARIO"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = [("Fecha", 12), ("Codigo PT", 18), ("Referencia", 42),
            ("Cantidad a Producir", 18), ("Tipo", 14), ("Fecha Entrega", 14),
            ("Categoria", 14)]
    for ci, (h, w) in enumerate(cols, 1):
        hdr(ws, 2, ci, h, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 24

    r = 3
    for _, row in plan_df.iterrows():
        bg      = C_ALT if r % 2 == 0 else C_BLANCO
        tipo_bg = C_AZUL_MED if row["tipo"] == "Proyeccion" else C_NARANJA
        cat     = row.get("categoria", "Stock")
        cat_bg  = C_MORADO if cat == "Pedido Activo" else C_VERDE

        dat(ws, r, 1, row["fecha"], bg=bg, fmt="DD/MM/YYYY", align="center")
        dat(ws, r, 2, row["codigo_pt"], bg=bg, align="center")
        dat(ws, r, 3, row["referencia"], bg=bg)
        dat(ws, r, 4, int(row["cantidad"]), bg=bg, fmt="#,##0",
            align="right", bold=True, fc=C_AZUL_OSC)
        c = ws.cell(row=r, column=5, value=row["tipo"])
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(tipo_bg)
        c.border    = _border
        c.alignment = Alignment(horizontal="center", vertical="center")
        dat(ws, r, 6, row["fecha_entrega"],
            bg=bg, fmt="DD/MM/YYYY", align="center")
        c = ws.cell(row=r, column=7, value=cat)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(cat_bg)
        c.border    = _border
        c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1

    ws.freeze_panes = "A3"
    if r > 3:
        ws.auto_filter.ref = f"A2:G{r - 1}"


def _hoja_plan_mensual(wb, plan_df):
    ws = wb.create_sheet("Plan_Mensual")

    if plan_df.empty:
        ws["A1"] = "Sin datos de produccion en el periodo"
        return

    MESES_ES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

    pt = plan_df.copy()
    pt["fecha"]     = pd.to_datetime(pt["fecha"])
    pt["mes_key"]   = pt["fecha"].apply(lambda d: (d.year, d.month))
    pt["mes_label"] = pt["fecha"].apply(lambda d: f"{MESES_ES[d.month]} {d.year}")

    orden_meses = (pt[["mes_label","mes_key"]].drop_duplicates()
                     .sort_values("mes_key")["mes_label"].tolist())

    pivot = (pt.groupby(["codigo_pt","referencia","mes_label"])["cantidad"]
               .sum()
               .unstack("mes_label")
               .fillna(0).astype(int)
               .rename_axis(None, axis=1)
               .reindex(columns=orden_meses, fill_value=0)
               .reset_index())

    COL_TOTAL  = 2 + len(orden_meses) + 1   # columna "Total Periodo"
    total_col  = COL_TOTAL
    ws.merge_cells(f"A1:{get_column_letter(total_col)}1")
    ws["A1"] = "SEEDPACK - PLAN DE PRODUCCION MENSUAL"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill      = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    hdr(ws, 2, 1, "Codigo PT",  bg=C_AZUL_OSC)
    hdr(ws, 2, 2, "Referencia", bg=C_AZUL_OSC)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    for pi, mes in enumerate(orden_meses, 3):
        hdr(ws, 2, pi, mes, bg=C_AZUL_MED, wrap=True)
        ws.column_dimensions[get_column_letter(pi)].width = 14
    hdr(ws, 2, COL_TOTAL, "Total Periodo", bg=C_VERDE, wrap=True)
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 16
    ws.row_dimensions[2].height = 30

    codigos_pa = set(plan_df[plan_df.get("categoria", pd.Series(dtype=str)) == "Pedido Activo"]["codigo_pt"]) \
        if "categoria" in plan_df.columns else set()

    r = 3
    for _, row in pivot.iterrows():
        es_pa  = row["codigo_pt"] in codigos_pa
        bg     = C_MORADO_CLR if es_pa else (C_ALT if r % 2 == 0 else C_BLANCO)
        fc_ref = C_MORADO if es_pa else "000000"
        dat(ws, r, 1, row["codigo_pt"],  bg=bg, align="center", fc=fc_ref, bold=es_pa)
        dat(ws, r, 2, row["referencia"], bg=bg, fc=fc_ref, bold=es_pa)
        fila_total = 0
        for pi, mes in enumerate(orden_meses, 3):
            val = int(row.get(mes, 0))
            fila_total += val
            if val > 0:
                c = ws.cell(row=r, column=pi, value=val)
                c.font         = Font(name="Calibri", bold=True, size=10, color=C_MORADO if es_pa else C_AZUL_OSC)
                c.fill         = _fill(C_MORADO_CLR if es_pa else C_ALT)
                c.border       = _border
                c.number_format = "#,##0"
                c.alignment    = Alignment(horizontal="right", vertical="center")
            else:
                dat(ws, r, pi, "-", bg=bg, align="center")
        # Total periodo
        ct = ws.cell(row=r, column=COL_TOTAL, value=fila_total)
        ct.font         = Font(name="Calibri", bold=True, size=10, color="166534")
        ct.fill         = _fill("D4EDDA")
        ct.border       = _border
        ct.number_format = "#,##0"
        ct.alignment    = Alignment(horizontal="right", vertical="center")
        r += 1

    dat(ws, r, 1, "TOTAL MES", bg=C_AZUL_OSC, bold=True, fc="FFFFFF", align="center")
    dat(ws, r, 2, "",          bg=C_AZUL_OSC)
    gran_total = 0
    for pi, mes in enumerate(orden_meses, 3):
        tot = int(pivot[mes].sum()) if mes in pivot.columns else 0
        gran_total += tot
        c = ws.cell(row=r, column=pi, value=tot)
        c.font         = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill         = _fill(C_AZUL_MED)
        c.border       = _border
        c.number_format = "#,##0"
        c.alignment    = Alignment(horizontal="right", vertical="center")
    ct = ws.cell(row=r, column=COL_TOTAL, value=gran_total)
    ct.font         = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ct.fill         = _fill(C_VERDE)
    ct.border       = _border
    ct.number_format = "#,##0"
    ct.alignment    = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "C3"


def _hoja_resumen_mrp(wb, mrp_df):
    """
    Hoja de trazabilidad MRP: muestra por (producto, mes) todos los componentes
    del calculo de produccion neta antes y despues del ajuste por lote minimo.

    Columnas: Codigo PT | Referencia | Periodo | Demanda ML | Stock Inicio |
              OPs Proceso | Neto s/Ajuste | Lote Min | Neto Final | Stock Cierre
    """
    ws = wb.create_sheet("Resumen_MRP")

    if mrp_df is None or mrp_df.empty:
        ws["A1"] = "Sin datos de MRP en el periodo"
        return

    COLS = [
        ("Codigo PT",       16),
        ("Referencia",      42),
        ("Periodo",         14),
        ("Demanda ML",      15),
        ("Stock Inicio",    15),
        ("OPs Proceso",     15),
        ("Neto s/Ajuste",   15),
        ("Lote Min",        12),
        ("Neto Final",      15),
        ("Stock Cierre",    15),
    ]
    n = len(COLS)

    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    ws["A1"] = "SEEDPACK - DETALLE MRP MENSUAL  (Demanda - Stock - OPs en Proceso = Neto a Producir)"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (h, w) in enumerate(COLS, 1):
        hdr(ws, 2, ci, h, bg=C_AZUL_OSC, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 32

    df = mrp_df.sort_values(["codigo_pt", "año", "mes"]).reset_index(drop=True)

    # Totales por columna para la fila de cierre
    tot_demanda = tot_stock_i = tot_ops = tot_neto_s = tot_neto_f = 0

    r = 3
    prev_codigo = None
    for _, row in df.iterrows():
        codigo   = str(row.get("codigo_pt", ""))
        cambio   = (codigo != prev_codigo)
        prev_codigo = codigo

        bg = C_ALT if r % 2 == 0 else C_BLANCO
        demanda   = int(row.get("demanda_proyectada", 0))
        stock_i   = int(row.get("stock_inicio",       0))
        ops_mes   = int(row.get("ops_proceso_mes",    0))
        neto_s    = int(row.get("neto_sin_ajuste",    row.get("neto_producir", 0)))
        lote_m    = int(row.get("lote_minimo",        0))
        neto_f    = int(row.get("neto_producir",      0))
        stock_c   = int(row.get("stock_cierre",       0))

        tot_demanda += demanda
        tot_stock_i += stock_i
        tot_ops     += ops_mes
        tot_neto_s  += neto_s
        tot_neto_f  += neto_f

        dat(ws, r, 1, codigo,                     bg=bg, align="center", bold=cambio)
        dat(ws, r, 2, str(row.get("producto","")),bg=bg, bold=cambio)
        dat(ws, r, 3, str(row.get("periodo","")), bg=bg, align="center")
        dat(ws, r, 4, demanda, bg=bg, fmt="#,##0", align="right")
        dat(ws, r, 5, stock_i, bg=bg, fmt="#,##0", align="right",
            fc=C_VERDE if stock_i >= demanda else ("000000"))
        dat(ws, r, 6, ops_mes, bg=bg, fmt="#,##0", align="right",
            fc=C_AZUL_MED if ops_mes > 0 else "000000")
        dat(ws, r, 7, neto_s,  bg=bg, fmt="#,##0", align="right")

        # Lote min: resaltar cuando hubo ajuste
        ajuste = (neto_f != neto_s and neto_s > 0)
        dat(ws, r, 8, lote_m if lote_m > 0 else "-",
            bg=C_AMARILLO if ajuste else bg, align="center")

        # Neto final: verde si es 0 (cubierto), azul oscuro si hay produccion
        c9 = ws.cell(row=r, column=9, value=neto_f)
        c9.font         = Font(name="Calibri", size=10, bold=(neto_f > 0),
                               color=C_AZUL_OSC if neto_f > 0 else "166534")
        c9.fill         = _fill(C_ALT if neto_f > 0 else "D4EDDA")
        c9.border       = _border
        c9.number_format = "#,##0"
        c9.alignment    = Alignment(horizontal="right", vertical="center")

        dat(ws, r, 10, stock_c, bg=bg, fmt="#,##0", align="right")
        r += 1

    # Fila de totales
    dat(ws, r, 1, "TOTAL", bg=C_AZUL_OSC, bold=True, fc="FFFFFF", align="center")
    dat(ws, r, 2, "",      bg=C_AZUL_OSC)
    dat(ws, r, 3, "",      bg=C_AZUL_OSC)
    for ci, val in enumerate([tot_demanda, tot_stock_i, tot_ops, tot_neto_s], 4):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = _fill(C_AZUL_MED); c.border = _border
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right", vertical="center")
    dat(ws, r, 8, "", bg=C_AZUL_OSC)
    ct = ws.cell(row=r, column=9, value=tot_neto_f)
    ct.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ct.fill = _fill(C_VERDE); ct.border = _border
    ct.number_format = "#,##0"
    ct.alignment = Alignment(horizontal="right", vertical="center")
    dat(ws, r, 10, "", bg=C_AZUL_OSC)

    ws.freeze_panes = "D3"
    if r > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(n)}{r - 1}"


def _hoja_calendario_semanal(wb, plan_df):
    ws = wb.create_sheet("Calendario_Semanal")

    if plan_df.empty:
        ws["A1"] = "Sin datos de produccion en el periodo"
        return

    pt = plan_df.copy()
    pt["fecha"]     = pd.to_datetime(pt["fecha"])
    pt["semana_ini"] = pt["fecha"].apply(lambda d: d - pd.Timedelta(days=d.weekday()))
    pt["semana"]    = pt["fecha"].apply(_label_semana)

    pivot = (pt.groupby(["codigo_pt", "referencia", "semana"])["cantidad"]
               .sum()
               .unstack("semana")
               .fillna(0)
               .astype(int)
               .rename_axis(None, axis=1)
               .reset_index())

    # Ordenar semanas cronologicamente por fecha real del lunes
    semana_orden = (pt[["semana_ini", "semana"]]
                    .drop_duplicates()
                    .sort_values("semana_ini")["semana"]
                    .tolist())
    semanas = [s for s in semana_orden if s in pivot.columns]
    total_col = 2 + len(semanas)

    ws.merge_cells(f"A1:{get_column_letter(total_col)}1")
    ws["A1"] = "SEEDPACK - CALENDARIO DE PRODUCCION SEMANAL"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    hdr(ws, 2, 1, "Codigo PT",   bg=C_AZUL_OSC)
    hdr(ws, 2, 2, "Referencia",  bg=C_AZUL_OSC)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    for pi, sem in enumerate(semanas, 3):
        hdr(ws, 2, pi, sem, bg=C_AZUL_MED, wrap=True)
        ws.column_dimensions[get_column_letter(pi)].width = 16
    ws.row_dimensions[2].height = 44

    codigos_pa_cal = set(plan_df[plan_df["categoria"] == "Pedido Activo"]["codigo_pt"]) \
        if "categoria" in plan_df.columns else set()

    r = 3
    for _, row in pivot.iterrows():
        es_pa  = row["codigo_pt"] in codigos_pa_cal
        bg     = C_MORADO_CLR if es_pa else (C_ALT if r % 2 == 0 else C_BLANCO)
        fc_ref = C_MORADO if es_pa else "000000"
        dat(ws, r, 1, row["codigo_pt"],  bg=bg, align="center", fc=fc_ref, bold=es_pa)
        dat(ws, r, 2, row["referencia"], bg=bg, fc=fc_ref, bold=es_pa)
        for pi, sem in enumerate(semanas, 3):
            val = int(row[sem])
            if val > 0:
                c = ws.cell(row=r, column=pi, value=val)
                c.font      = Font(name="Calibri", bold=True, size=10, color=C_MORADO if es_pa else C_AZUL_OSC)
                c.fill      = _fill(C_MORADO_CLR if es_pa else C_ALT)
                c.border    = _border
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                dat(ws, r, pi, "-", bg=bg, align="center")
        r += 1

    # Fila totales
    dat(ws, r, 1, "TOTAL SEMANA", bg=C_AZUL_OSC, bold=True, fc="FFFFFF", align="center")
    dat(ws, r, 2, "",             bg=C_AZUL_OSC)
    for pi, sem in enumerate(semanas, 3):
        tot = int(pivot[sem].sum())
        c = ws.cell(row=r, column=pi, value=tot)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(C_AZUL_MED)
        c.border    = _border
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "C3"


def _enriquecer_ops_con_entradas(ops_df, entradas_df):
    """Agrega Cant. Producida, Cant. Pendiente y Estado a ops_df cruzando por OC."""
    ent = _aplicar_mapeos_usuario(entradas_df.copy(), "arch_entradas")
    oc_col  = next((c for c in ent.columns if c.upper().strip() == "OC"), None)
    qty_col = next((c for c in ent.columns
                    if c.lower().strip() in ("cantidad", "cant.", "cant. ingresada")), None)
    prod_dict = {}
    if oc_col and qty_col:
        ent[oc_col]  = pd.to_numeric(ent[oc_col],  errors="coerce")
        ent[qty_col] = pd.to_numeric(ent[qty_col], errors="coerce").fillna(0)
        prod_dict = ent.groupby(oc_col)[qty_col].sum().to_dict()

    producidas, pendientes, estados = [], [], []
    for _, row in ops_df.iterrows():
        aprobada = float(row.get("Cant. Aprobada", 0) or 0)
        try:
            op_key = float(row.get("OP", ""))
        except (ValueError, TypeError):
            op_key = None
        producida = float(prod_dict.get(op_key, 0)) if op_key is not None else 0.0
        pendiente = max(0.0, aprobada - producida)
        pct = producida / aprobada * 100 if aprobada > 0 else 0.0
        if pct >= 100:
            estado = "Completada"
        elif producida > 0:
            estado = "En Proceso"
        else:
            estado = "Sin Iniciar"
        producidas.append(producida)
        pendientes.append(pendiente)
        estados.append(estado)

    ops_df = ops_df.copy()
    ops_df["Cant. Producida"] = producidas
    ops_df["Cant. Pendiente"] = pendientes
    ops_df["Estado"] = estados
    return ops_df


def _hoja_ops_proceso(wb, ops_df):
    ws = wb.create_sheet("OPs_En_Proceso")

    tiene_saldo = "Cant. Producida" in ops_df.columns

    # Color map para estado
    _ESTADO_BG = {"Completada": "D4EDDA", "En Proceso": "FFF3CD", "Sin Iniciar": "FFE4E4"}
    _ESTADO_FC = {"Completada": "166534", "En Proceso": "92400E", "Sin Iniciar": "991B1B"}

    if tiene_saldo:
        n_cols = 11
        cols = [("OP", 8), ("Tipo de Trabajo", 18), ("Cod. Producto", 18),
                ("Referencia", 42), ("Cliente", 28),
                ("Cant. Aprobada", 15), ("Cant. Producida", 15), ("Cant. Pendiente", 15),
                ("Estado", 14), ("Fecha Programada", 18), ("Compromiso Cliente", 18)]
    else:
        n_cols = 8
        cols = [("OP", 8), ("Tipo de Trabajo", 18), ("Cod. Producto", 18),
                ("Referencia", 42), ("Cliente", 28),
                ("Cant. Aprobada", 15), ("Fecha Programada", 18), ("Compromiso Cliente", 18)]

    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "ORDENES DE PRODUCCION EN PROCESO"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (h, w) in enumerate(cols, 1):
        hdr(ws, 2, ci, h, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 34

    r = 3
    for _, row in ops_df.sort_values("Fecha Programada").iterrows():
        bg   = C_ALT if r % 2 == 0 else C_BLANCO
        cant = row.get("Cant. Aprobada", 0)
        fp   = row.get("Fecha Programada")
        cc   = row.get("Compromiso Cliente")

        dat(ws, r, 1, row.get("OP", ""), bg=bg, align="center")
        dat(ws, r, 2, row.get("Tipo de Trabajo", ""), bg=bg)
        dat(ws, r, 3, row.get("Cód. Producto", ""), bg=bg, align="center")
        dat(ws, r, 4, row.get("Referencia", ""), bg=bg)
        dat(ws, r, 5, row.get("Cliente", ""), bg=bg)
        dat(ws, r, 6,
            int(cant) if cant > 0 else "Sin definir",
            bg=bg, fmt="#,##0", align="right",
            fc=C_ROJO if cant == 0 else "000000")

        if tiene_saldo:
            producida = float(row.get("Cant. Producida", 0) or 0)
            pendiente = float(row.get("Cant. Pendiente", 0) or 0)
            estado    = str(row.get("Estado", "Sin Iniciar"))
            ebg = _ESTADO_BG.get(estado, bg)
            efc = _ESTADO_FC.get(estado, "000000")
            dat(ws, r, 7, int(producida), bg=bg, fmt="#,##0", align="right")
            dat(ws, r, 8, int(pendiente),
                bg=("FFE4E4" if pendiente > 0 else "D4EDDA"),
                fmt="#,##0", align="right",
                fc=("991B1B" if pendiente > 0 else "166534"))
            c_estado = ws.cell(row=r, column=9, value=estado)
            c_estado.font      = Font(name="Calibri", size=10, bold=True, color=efc)
            c_estado.fill      = _fill(ebg)
            c_estado.border    = _border
            c_estado.alignment = Alignment(horizontal="center", vertical="center")
            dat(ws, r, 10,
                fp.date() if pd.notna(fp) else "-",
                bg=bg, fmt="DD/MM/YYYY", align="center")
            dat(ws, r, 11,
                cc.date() if pd.notna(cc) else "-",
                bg=bg, fmt="DD/MM/YYYY", align="center")
        else:
            dat(ws, r, 7,
                fp.date() if pd.notna(fp) else "-",
                bg=bg, fmt="DD/MM/YYYY", align="center")
            dat(ws, r, 8,
                cc.date() if pd.notna(cc) else "-",
                bg=bg, fmt="DD/MM/YYYY", align="center")
        r += 1

    ws.freeze_panes = "A3"
    if r > 3:
        ws.auto_filter.ref = f"A2:{last_col}{r - 1}"


def _hoja_pedidos_ocasionales(wb, tabla_df):
    ws = wb.create_sheet("Pedidos_Ocasionales")

    ws.merge_cells("A1:K1")
    ws["A1"] = "PEDIDOS PENDIENTES - REFERENCIAS OCASIONALES (fuera de la proyeccion regular)"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    cols = [("Pedido", 10), ("Codigo PT", 18), ("Referencia", 40), ("Cliente", 28),
            ("Fecha Entrega", 15), ("Cant. Pedida", 14), ("Stock Disp.", 12),
            ("OPs Proceso", 12), ("Neto s/Ajuste", 14), ("Lote Minimo", 12),
            ("Neto a Producir", 16)]
    for ci, (h, w) in enumerate(cols, 1):
        hdr(ws, 2, ci, h, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 34

    if tabla_df.empty:
        ws.cell(row=3, column=1, value="Sin pedidos ocasionales en el periodo")
        return

    tabla_sorted = tabla_df.sort_values("Fecha Entrega")
    r = 3
    for _, row in tabla_sorted.iterrows():
        bg        = C_ALT if r % 2 == 0 else C_BLANCO
        neto_orig = row.get("Neto sin Ajuste",  0)
        lote      = row.get("Lote Minimo",      0)
        neto_aj   = row.get("Neto a Producir",  0)
        ajustado  = (lote > 0 and neto_aj != neto_orig)
        nc        = C_AMARILLO if neto_aj > 0 else C_VERDE

        dat(ws, r,  1, row.get("Pedido", ""),           bg=bg, align="center")
        dat(ws, r,  2, row.get("Código PT", ""),        bg=bg, align="center")
        dat(ws, r,  3, row.get("Referencia", ""),       bg=bg)
        dat(ws, r,  4, row.get("Cliente", ""),          bg=bg)
        dat(ws, r,  5, row.get("Fecha Entrega"),        bg=bg, fmt="DD/MM/YYYY", align="center")
        dat(ws, r,  6, row.get("Cantidad Pedida",  0),  bg=bg, fmt="#,##0", align="right")
        dat(ws, r,  7, row.get("Stock Disponible", 0),  bg=bg, fmt="#,##0", align="right")
        dat(ws, r,  8, row.get("OPs Proceso",      0),  bg=bg, fmt="#,##0", align="right")
        dat(ws, r,  9, neto_orig,                       bg=bg, fmt="#,##0", align="right")
        dat(ws, r, 10, lote if lote > 0 else "-",       bg=bg, fmt="#,##0", align="right",
            fc=C_AZUL_MED if lote > 0 else "888888")
        c = ws.cell(row=r, column=11, value=neto_aj)
        c.font      = Font(name="Calibri", bold=True, size=10,
                           color=C_AZUL_OSC if ajustado else "000000")
        c.fill      = _fill(nc)
        c.border    = _border
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    ws.freeze_panes = "A3"
    if r > 3:
        ws.auto_filter.ref = f"A2:K{r - 1}"


def _hoja_inventario(wb, bodega_df):
    # ── Hoja 1: Inventario_Actual  (CONSOLIDADO por Cód. PT) ─────────────────
    ws = wb.create_sheet("Inventario_Actual")

    n_cols_cons = 6
    ws.merge_cells(f"A1:{get_column_letter(n_cols_cons)}1")
    ws["A1"] = "INVENTARIO ACTUAL — CONSOLIDADO POR PRODUCTO  (excluye Obsoletos)"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    cons_cols = [("Cod. PT", 18), ("Referencia", 44), ("# Bodegas", 12),
                 ("Saldo", 14), ("Costo Unit.", 14), ("Costo Total", 16)]
    for ci, (h, w) in enumerate(cons_cols, 1):
        hdr(ws, 2, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    has_costo = "Costo" in bodega_df.columns
    has_total = "Total" in bodega_df.columns

    agg_dict = {
        "ReferenciaP": ("ReferenciaP", "first"),
        "n_bodegas":   ("Bodega",      "nunique"),
        "Saldo":       ("Saldo",       "sum"),
    }

    bodega_df["Saldo"] = pd.to_numeric(bodega_df["Saldo"], errors="coerce").fillna(0)
    if has_costo:
        bodega_df["Costo"] = pd.to_numeric(bodega_df["Costo"], errors="coerce").fillna(0)
        bodega_df["_ct"] = bodega_df["Saldo"] * bodega_df["Costo"]
        agg_dict["_ct"] = ("_ct", "sum")

    cons = bodega_df.groupby("Cód. PT").agg(**agg_dict).reset_index()
    cons = cons.sort_values("Saldo", ascending=False)

    r = 3
    for _, row in cons.iterrows():
        bg = C_ALT if r % 2 == 0 else C_BLANCO
        saldo = float(row.get("Saldo", 0))
        if has_costo:
            total = float(row.get("_ct", 0))
            costo = total / saldo if saldo > 0 else 0.0
        else:
            costo = 0.0
            total = 0.0
        dat(ws, r, 1, str(row["Cód. PT"]),        bg=bg, align="center")
        dat(ws, r, 2, str(row.get("ReferenciaP", "")), bg=bg)
        dat(ws, r, 3, int(row.get("n_bodegas", 1)),    bg=bg, align="center")
        dat(ws, r, 4, saldo,  bg=bg, fmt="#,##0",   align="right")
        dat(ws, r, 5, costo,  bg=bg, fmt="#,##0.00",align="right")
        dat(ws, r, 6, total,  bg=bg, fmt="#,##0",   align="right")
        r += 1

    ws.freeze_panes = "A3"
    if r > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(n_cols_cons)}{r - 1}"

    # ── Hoja 2: Inventario_Detalle  (detalle por Bodega) ─────────────────────
    ws2 = wb.create_sheet("Inventario_Detalle")

    has_cant_ing = "Cant. Ingresada" in bodega_df.columns
    has_opoc     = "OP/OC" in bodega_df.columns
    det_col_defs = [("Bodega", 24), ("Cod. PT", 18), ("Referencia", 44)]
    if has_cant_ing:
        det_col_defs.append(("Cant. Ingresada", 16))
    det_col_defs += [("Saldo", 14)]
    if has_costo:
        det_col_defs.append(("Costo", 14))
        det_col_defs.append(("Total", 14))
    if has_opoc:
        det_col_defs.append(("OP/OC", 18))
    n_cols_det = len(det_col_defs)

    ws2.merge_cells(f"A1:{get_column_letter(n_cols_det)}1")
    ws2["A1"] = "INVENTARIO ACTUAL — DETALLE POR BODEGA  (excluye Obsoletos)"
    ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws2["A1"].fill = _fill(C_AZUL_OSC)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for ci, (h, w) in enumerate(det_col_defs, 1):
        hdr(ws2, 2, ci, h)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 20

    r2 = 3
    for _, row in bodega_df.sort_values(["Bodega", "Saldo"], ascending=[True, False]).iterrows():
        bg = C_ALT if r2 % 2 == 0 else C_BLANCO
        ci = 1
        dat(ws2, r2, ci, str(row.get("Bodega", "")),       bg=bg);              ci += 1
        dat(ws2, r2, ci, str(row.get("Cód. PT", "")),      bg=bg, align="center"); ci += 1
        dat(ws2, r2, ci, str(row.get("ReferenciaP", "")),  bg=bg);              ci += 1
        if has_cant_ing:
            dat(ws2, r2, ci, float(row.get("Cant. Ingresada", 0)), bg=bg, fmt="#,##0", align="right"); ci += 1
        dat(ws2, r2, ci, float(row.get("Saldo", 0)), bg=bg, fmt="#,##0", align="right"); ci += 1
        if has_costo:
            _saldo_det = float(row.get("Saldo", 0))
            _costo_det = float(row.get("Costo", 0))
            dat(ws2, r2, ci, _costo_det, bg=bg, fmt="#,##0.00", align="right"); ci += 1
            dat(ws2, r2, ci, _saldo_det * _costo_det, bg=bg, fmt="#,##0", align="right"); ci += 1
        if has_opoc:
            dat(ws2, r2, ci, str(row.get("OP/OC", "")), bg=bg, align="center"); ci += 1
        r2 += 1

    ws2.freeze_panes = "A3"
    if r2 > 3:
        ws2.auto_filter.ref = f"A2:{get_column_letter(n_cols_det)}{r2 - 1}"


def _hoja_dashboard(wb, plan_df, mrp_df, tabla_ocas,
                    stock, ops_df, pedidos_df, lote_min, codigos_stock, pc_df=None):
    ws = wb.create_sheet("Dashboard")

    # ── Helpers de layout ─────────────────────────────────────────────────────
    def card(row, col, label, val, bg, fc="FFFFFF", fmt="#,##0"):
        lc = get_column_letter(col); lc1 = get_column_letter(col + 1)
        ws.merge_cells(f"{lc}{row}:{lc1}{row}")
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.fill = _fill(C_AZUL_MED); c.border = _border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 26
        ws.merge_cells(f"{lc}{row+1}:{lc1}{row+1}")
        c2 = ws.cell(row=row+1, column=col, value=val)
        c2.font = Font(name="Calibri", bold=True, size=17, color=fc)
        c2.fill = _fill(bg); c2.border = _border
        c2.number_format = fmt
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row+1].height = 40

    def sec(row, label):
        ws.merge_cells(f"B{row}:I{row}")
        c = ws.cell(row=row, column=2, value=f"   {label}")
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill = _fill(C_AZUL_OSC); c.border = _border
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24

    def blank(row):
        ws.row_dimensions[row].height = 8

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = "SEEDPACK  —  DASHBOARD EJECUTIVO"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 44

    for ci, w in enumerate([2,16,16,16,16,16,16,16,16,2], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ══ SECCIÓN 1 — PLAN DE PRODUCCIÓN ═══════════════════════════════════════
    n_dias  = plan_df["fecha"].nunique()     if not plan_df.empty else 0
    n_und   = int(plan_df["cantidad"].sum()) if not plan_df.empty else 0
    n_stock = (plan_df[plan_df["categoria"] == "Stock"]["codigo_pt"].nunique()
               if "categoria" in plan_df.columns and not plan_df.empty else 0)
    n_pa    = (plan_df[plan_df["categoria"] == "Pedido Activo"]["codigo_pt"].nunique()
               if "categoria" in plan_df.columns and not plan_df.empty else 0)

    blank(2); sec(3, "PLAN DE PRODUCCION")
    card(4, 2, "Dias planificados",          n_dias,  C_AZUL_MED)
    card(4, 4, "Unidades totales",           n_und,   C_AZUL_MED)
    card(4, 6, "Referencias stock",          n_stock, C_VERDE,   "000000")
    card(4, 8, "Referencias pedido activo",  n_pa,    C_NARANJA)

    # ══ SECCIÓN 2 — INVENTARIO ACTUAL ════════════════════════════════════════
    stock_total = int(sum(stock.values()))
    gen_ok   = sum(1 for c in codigos_stock
                   if lote_min.get(c, 0) > 0 and stock.get(c, 0) >= lote_min.get(c, 0))
    gen_crit = sum(1 for c in codigos_stock
                   if lote_min.get(c, 0) > 0 and 0 < stock.get(c, 0) < lote_min.get(c, 0))
    gen_cero = sum(1 for c in codigos_stock if stock.get(c, 0) == 0)

    blank(7); sec(8, "INVENTARIO ACTUAL")
    card(9,  2, "Unidades en bodega",                          stock_total, C_AZUL_MED)
    card(9,  4, f"Genericos OK  (/{len(codigos_stock)})",      gen_ok,      C_VERDE,    "000000")
    card(9,  6, "Genericos bajo lote minimo",                  gen_crit,    C_AMARILLO, "000000")
    card(9,  8, "Genericos en cero",                           gen_cero,    C_ROJO if gen_cero else C_VERDE, "000000" if not gen_cero else "FFFFFF")

    # ══ SECCIÓN 3 — ÓRDENES DE PRODUCCIÓN EN PROCESO ═════════════════════════
    ops_activas = len(ops_df)
    if not ops_df.empty and "Cant. Pendiente" in ops_df.columns:
        ops_und     = int(pd.to_numeric(ops_df["Cant. Pendiente"], errors="coerce").fillna(0).sum())
        ops_und_lbl = "Uds. pendientes por producir"
    else:
        ops_und     = int(pd.to_numeric(ops_df.get("Cant. Aprobada", pd.Series(dtype=float)),
                                        errors="coerce").fillna(0).sum())
        ops_und_lbl = "Unidades en proceso"
    hoy      = pd.Timestamp.today().normalize()
    ops_prox = 0
    if not ops_df.empty and "Fecha Programada" in ops_df.columns:
        fp = pd.to_datetime(ops_df["Fecha Programada"], errors="coerce")
        ops_prox = int(((fp >= hoy) & (fp <= hoy + pd.Timedelta(days=15))).sum())

    blank(12); sec(13, "ORDENES DE PRODUCCION EN PROCESO")
    card(14, 2, "OPs activas",              ops_activas, C_AZUL_MED)
    card(14, 4, ops_und_lbl,               ops_und,     C_AZUL_MED)
    card(14, 6, "OPs entregan proximos 15d",ops_prox,    C_NARANJA if ops_prox > 0 else C_AZUL_MED)

    # ══ SECCIÓN 4 — PEDIDOS ══════════════════════════════════════════════════
    ped_total    = len(pedidos_df)
    ped_prod     = int((tabla_ocas["Neto a Producir"] > 0).sum()) if not tabla_ocas.empty else 0
    und_ped_prod = int(tabla_ocas["Neto a Producir"].sum())       if not tabla_ocas.empty else 0

    blank(17); sec(18, "PEDIDOS PENDIENTES")
    card(19, 2, "Total pedidos pendientes",  ped_total,    C_AZUL_MED)
    card(19, 4, "Requieren produccion",      ped_prod,     C_NARANJA if ped_prod > 0 else C_VERDE, "000000" if not ped_prod else "FFFFFF")
    card(19, 6, "Unidades por producir",     und_ped_prod, C_AZUL_MED)

    # ══ SECCIÓN 5 — PLAN COMERCIAL (solo si está disponible) ═════════════════
    if pc_df is not None and not pc_df.empty and not mrp_df.empty:
        meses_pc = (pc_df[["año","mes"]].drop_duplicates().sort_values(["año","mes"]))
        n_meses  = len(meses_pc)
        n_prods  = pc_df["codigo"].nunique()

        blank(22); sec(23, "PLAN COMERCIAL")
        card(24, 2, "Meses en plan comercial", n_meses, C_AZUL_MED)
        card(24, 4, "Productos en plan",       n_prods, C_AZUL_MED)


def _hoja_saldo_ops(wb, ops_df, entradas_df):
    """Cruza OPs en proceso con entradas de inventario (por OC) para calcular pendiente real."""
    ws = wb.create_sheet("Saldo_OPs")

    COLS = [("OP", 10), ("Tipo", 18), ("Cód. Producto", 18), ("Referencia", 40),
            ("Cant. Aprobada", 16), ("Cant. Producida", 16), ("Cant. Pendiente", 16),
            ("Estado", 14),
            ("Fecha Prog.", 18), ("Compromiso", 18)]
    n = len(COLS)

    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    ws["A1"] = "SALDO REAL DE ORDENES DE PRODUCCION EN PROCESO  (OPs vs Entradas Inventario)"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (h, w) in enumerate(COLS, 1):
        hdr(ws, 2, ci, h, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # Mapa OC → cant. producida (suma de Cantidad por OC)
    entradas_df = entradas_df.copy()
    entradas_df = _aplicar_mapeos_usuario(entradas_df, "arch_entradas")
    oc_col  = next((c for c in entradas_df.columns if c.upper().strip() == "OC"), None)
    qty_col = next((c for c in entradas_df.columns
                    if c.lower().strip() in ("cantidad", "cant.", "cant. ingresada")), None)
    producido_dict = {}
    if oc_col and qty_col:
        entradas_df[oc_col]  = pd.to_numeric(entradas_df[oc_col],  errors="coerce")
        entradas_df[qty_col] = pd.to_numeric(entradas_df[qty_col], errors="coerce").fillna(0)
        producido_dict = entradas_df.groupby(oc_col)[qty_col].sum().to_dict()

    cod_col = next((c for c in ops_df.columns
                    if "producto" in c.lower() and ("cód" in c.lower() or "cod" in c.lower())), None)

    r = 3
    for _, row in ops_df.iterrows():
        op_num    = row.get("OP", "")
        tipo      = str(row.get("Tipo de Trabajo", ""))
        cod       = str(row.get(cod_col, "")) if cod_col else ""
        ref       = str(row.get("Referencia", ""))
        aprobada  = float(row.get("Cant. Aprobada", 0) or 0)
        if aprobada == 0:
            continue
        fp        = row.get("Fecha Programada")
        cc        = row.get("Compromiso Cliente")

        try:
            op_key = float(op_num)
        except (ValueError, TypeError):
            op_key = None
        producida = float(producido_dict.get(op_key, 0)) if op_key is not None else 0
        pendiente = max(0.0, aprobada - producida)
        pct       = round(producida / aprobada * 100, 1) if aprobada > 0 else 0.0

        if pct >= 100:
            estado = "Completada";   bg = "D4EDDA"; fc = "166534"
        elif producida > 0:
            estado = "En Proceso";   bg = "FFF3CD"; fc = "92400E"
        else:
            estado = "Sin Iniciar";  bg = "FFE4E4"; fc = "991B1B"

        vals = [op_num, tipo, cod, ref,
                aprobada, producida, pendiente,
                estado,
                fp.strftime("%Y-%m-%d") if pd.notna(fp) else "",
                cc.strftime("%Y-%m-%d") if pd.notna(cc) else ""]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font      = Font(name="Calibri", size=10,
                                  color=(fc if ci == 8 else "000000"),
                                  bold=(ci == 8))
            cell.fill      = _fill(bg)
            cell.border    = _border
            num_cols = {5, 6, 7}
            cell.alignment = Alignment(
                horizontal="right" if ci in num_cols else
                           "center" if ci in {1, 2, 8, 9, 10} else "left",
                vertical="center")
            if ci in num_cols:
                cell.number_format = "#,##0"
        r += 1

    ws.freeze_panes = "A3"
    if r > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(n)}{r - 1}"


def _hoja_pedidos_sin_cubrir(wb, pedidos_df, stock, ops_supply, plan_df):
    ws = wb.create_sheet("Pedidos_Sin_Cubrir")

    ws.merge_cells("A1:L1")
    ws["A1"] = "PEDIDOS SIN CUBRIR — deficit tras stock + OPs + plan de produccion hasta fecha de entrega"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = _fill(C_ROJO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    col_def = [("Pedido",12),("Codigo PT",18),("Referencia",40),("Cliente",28),
               ("Fecha Entrega",15),("Cant. Pedida",14),("Stock",12),
               ("OPs Proceso",12),("Plan hasta Entrega",18),("Total Disp.",14),
               ("Deficit",12),("Estado",14)]
    for ci, (h, w) in enumerate(col_def, 1):
        hdr(ws, 2, ci, h, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 34

    if pedidos_df.empty:
        ws.cell(row=3, column=1, value="Sin pedidos en el periodo.")
        return

    # OPs totales por codigo
    ops_total = defaultdict(float)
    for (cod, _a, _m), q in ops_supply.items():
        ops_total[cod] += q

    # Produccion acumulada por codigo hasta cada fecha (para lookup rapido)
    plan_cum = {}
    if not plan_df.empty:
        tmp = plan_df.copy()
        tmp["fecha"] = pd.to_datetime(tmp["fecha"])
        for cod, grp in tmp.groupby("codigo_pt"):
            daily = grp.groupby("fecha")["cantidad"].sum().sort_index()
            plan_cum[cod] = daily.cumsum()

    def _plan_hasta(codigo, fecha):
        if codigo not in plan_cum:
            return 0.0
        ser = plan_cum[codigo]
        sel = ser[ser.index <= pd.Timestamp(fecha)]
        return float(sel.iloc[-1]) if len(sel) > 0 else 0.0

    sin_cubrir = []
    for _, ped in pedidos_df.sort_values("Fecha de entrega").iterrows():
        codigo   = str(ped["Código PT"]).strip()
        qty      = float(ped.get("Cantidad", 0))
        f_ent    = ped["Fecha de entrega"]
        inv      = float(stock.get(codigo, 0))
        ops      = float(ops_total.get(codigo, 0))
        plan_q   = _plan_hasta(codigo, f_ent)
        total    = inv + ops + plan_q
        deficit  = max(0.0, qty - total)
        if deficit <= 0:
            continue
        sin_cubrir.append({
            "Pedido":            ped.get("Pedido", ""),
            "Código PT":         codigo,
            "Referencia":        str(ped.get("Referencia", "")),
            "Cliente":           str(ped.get("Cliente", "")),
            "Fecha Entrega":     f_ent.date() if pd.notna(f_ent) else None,
            "Cant. Pedida":      round(qty),
            "Stock":             round(inv),
            "OPs Proceso":       round(ops),
            "Plan hasta Entrega":round(plan_q),
            "Total Disp.":       round(total),
            "Deficit":           round(deficit),
            "Estado":            "Sin cubrir",
        })

    if not sin_cubrir:
        ws.cell(row=3, column=1,
                value="Todos los pedidos quedan cubiertos por el plan de produccion.")
        return

    r = 3
    for row in sin_cubrir:
        bg = C_ALT if r % 2 == 0 else C_BLANCO
        dat(ws, r,  1, row["Pedido"],             bg=bg, align="center")
        dat(ws, r,  2, row["Código PT"],          bg=bg, align="center")
        dat(ws, r,  3, row["Referencia"],         bg=bg)
        dat(ws, r,  4, row["Cliente"],            bg=bg)
        dat(ws, r,  5, row["Fecha Entrega"],      bg=bg, fmt="DD/MM/YYYY", align="center")
        dat(ws, r,  6, row["Cant. Pedida"],       bg=bg, fmt="#,##0", align="right")
        dat(ws, r,  7, row["Stock"],              bg=bg, fmt="#,##0", align="right")
        dat(ws, r,  8, row["OPs Proceso"],        bg=bg, fmt="#,##0", align="right")
        dat(ws, r,  9, row["Plan hasta Entrega"], bg=bg, fmt="#,##0", align="right")
        dat(ws, r, 10, row["Total Disp."],        bg=bg, fmt="#,##0", align="right")
        c = ws.cell(row=r, column=11, value=row["Deficit"])
        c.font          = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill          = _fill(C_ROJO)
        c.border        = _border
        c.number_format = "#,##0"
        c.alignment     = Alignment(horizontal="right", vertical="center")
        dat(ws, r, 12, row["Estado"], bg=C_ROJO, fc="FFFFFF", bold=True, align="center")
        r += 1

    ws.freeze_panes = "A3"
    if r > 3:
        ws.auto_filter.ref = f"A2:L{r - 1}"


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACION DE RUTAS (llamar antes de run() cuando se usan rutas externas)
# ══════════════════════════════════════════════════════════════════════════════

def configurar_rutas(arch_ventas, arch_ops_hist, arch_pedidos,
                     arch_ops_proc, arch_bodega, arch_lote_min,
                     arch_entradas=None, arch_plan_comercial=None):
    """Sobreescribe las rutas de archivos de entrada antes de ejecutar run()."""
    global ARCH_VENTAS, ARCH_OPS_HIST, ARCH_PEDIDOS, ARCH_OPS_PROC, ARCH_BODEGA, ARCH_LOTE_MIN, ARCH_ENTRADAS, ARCH_PLAN_COMERCIAL
    ARCH_VENTAS          = arch_ventas
    ARCH_OPS_HIST        = arch_ops_hist
    ARCH_PEDIDOS         = arch_pedidos
    ARCH_OPS_PROC        = arch_ops_proc
    ARCH_BODEGA          = arch_bodega
    ARCH_LOTE_MIN        = arch_lote_min
    ARCH_ENTRADAS        = arch_entradas
    ARCH_PLAN_COMERCIAL  = arch_plan_comercial


def _hoja_comparativa_comercial(wb, mrp_df, pc_df, pc_name_map=None):
    """
    Hoja Comparativa_Comercial: proyección ML vs plan comercial por producto y mes.
    Semáforo Δ%: verde ≤10%, amarillo ≤30%, rojo >30%.
    "Sin ML"  = comercial planifica pero ML no proyecta ese mes (naranja).
    "Sin Com." = ML proyecta pero comercial no planifica (gris).
    Columnas extra al final: Total ML, Total Com., Δ% Total.
    """
    ws = wb.create_sheet("Comparativa_Comercial")

    if pc_df.empty or mrp_df.empty:
        ws["A1"] = "Sin datos suficientes para la comparativa"
        return

    # ── Proyección ML agregada por codigo + año + mes ────────────────────────
    ml = (mrp_df.groupby(["codigo_pt", "año", "mes"])["demanda_proyectada"]
                .sum().reset_index()
                .rename(columns={"codigo_pt": "codigo", "demanda_proyectada": "ml"}))

    # Mapeo codigo -> nombre: mrp_df tiene prioridad, luego pc_name_map, luego vacío
    ref_map = mrp_df.drop_duplicates("codigo_pt").set_index("codigo_pt")["producto"].to_dict()
    if pc_name_map:
        for cod, nom in pc_name_map.items():
            if cod not in ref_map:
                ref_map[cod] = nom

    # ── Meses disponibles en el plan comercial (orden cronológico) ───────────
    meses_pc = (pc_df[["año", "mes"]].drop_duplicates()
                                      .sort_values(["año", "mes"])
                                      .itertuples(index=False))
    meses = [(r.año, r.mes) for r in meses_pc]

    if not meses:
        ws["A1"] = "El plan comercial no contiene columnas de meses reconocibles"
        return

    # ── Unión de productos ───────────────────────────────────────────────────
    codigos = sorted(set(pc_df["codigo"]) | set(ml["codigo"]))

    # ── Dimensiones: 2 fijas + N*2 meses + 2 columnas de total ─────────────
    N          = len(meses)
    col_tot_ml  = 2 + N * 2 + 1   # Total ML
    col_tot_com = col_tot_ml + 1   # Total Comercial
    total_col   = col_tot_com

    # Fila 1: título
    ws.merge_cells(f"A1:{get_column_letter(total_col)}1")
    ws["A1"] = "SEEDPACK — COMPARATIVA PLAN COMERCIAL vs PROYECCIÓN ML"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill      = _fill(C_AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fila 2: encabezados de mes fusionados en 2 columnas + encabezado totales
    hdr(ws, 2, 1, "Codigo PT",  bg=C_AZUL_OSC); ws.column_dimensions["A"].width = 18
    hdr(ws, 2, 2, "Referencia", bg=C_AZUL_OSC); ws.column_dimensions["B"].width = 40
    for i, (yr, mes) in enumerate(meses):
        col_ini = 3 + i * 2
        col_fin = col_ini + 1
        lbl = f"{_MESES_ES[mes]} {yr}"
        ws.merge_cells(f"{get_column_letter(col_ini)}2:{get_column_letter(col_fin)}2")
        c = ws.cell(row=2, column=col_ini, value=lbl)
        c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        c.fill      = _fill(C_AZUL_MED)
        c.border    = _border
        c.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(col_ini, col_fin + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.merge_cells(f"{get_column_letter(col_tot_ml)}2:{get_column_letter(col_tot_com)}2")
    c_tot = ws.cell(row=2, column=col_tot_ml, value="TOTAL PERIODO")
    c_tot.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c_tot.fill      = _fill(C_MORADO)
    c_tot.border    = _border
    c_tot.alignment = Alignment(horizontal="center", vertical="center")
    for ci in (col_tot_ml, col_tot_com):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[2].height = 24

    # Fila 3: sub-encabezados
    hdr(ws, 3, 1, "Codigo",     bg=C_AZUL_OSC)
    hdr(ws, 3, 2, "Referencia", bg=C_AZUL_OSC)
    for i in range(N):
        col_ini = 3 + i * 2
        yr_i, mes_i = meses[i]
        m_lbl = f"{_MESES_ES[mes_i]}{str(yr_i)[2:]}"
        hdr(ws, 3, col_ini,     f"{m_lbl} ML",  bg=C_AZUL_MED)
        hdr(ws, 3, col_ini + 1, f"{m_lbl} Com.", bg=C_AZUL_MED)
    hdr(ws, 3, col_tot_ml,  "Total ML",   bg=C_MORADO)
    hdr(ws, 3, col_tot_com, "Total Com.", bg=C_MORADO)
    ws.row_dimensions[3].height = 20

    # ── Índices para lookup rápido ───────────────────────────────────────────
    ml_idx = ml.set_index(["codigo", "año", "mes"])["ml"].to_dict()
    pc_idx = pc_df.set_index(["codigo", "año", "mes"])["cantidad_comercial"].to_dict()

    # ── Filas de datos ───────────────────────────────────────────────────────
    r = 4
    tot_ml  = {(yr, mes): 0.0 for yr, mes in meses}
    tot_com = {(yr, mes): 0.0 for yr, mes in meses}

    for codigo in codigos:
        bg  = C_ALT if r % 2 == 0 else C_BLANCO
        ref = ref_map.get(codigo, "")
        dat(ws, r, 1, codigo, bg=bg, align="center")
        dat(ws, r, 2, ref,    bg=bg)

        sum_ml  = 0.0
        sum_com = 0.0

        for i, (yr, mes) in enumerate(meses):
            col_ini = 3 + i * 2
            v_ml    = float(ml_idx.get((codigo, yr, mes), 0.0))
            v_com   = float(pc_idx.get((codigo, yr, mes), 0.0))
            tot_ml [(yr, mes)] += v_ml
            tot_com[(yr, mes)] += v_com
            sum_ml  += v_ml
            sum_com += v_com

            dat(ws, r, col_ini,     round(v_ml)  if v_ml  else "-", bg=bg, fmt="#,##0", align="right")
            dat(ws, r, col_ini + 1, round(v_com) if v_com else "-", bg=bg, fmt="#,##0", align="right")

        dat(ws, r, col_tot_ml,  round(sum_ml)  if sum_ml  else "-",
            bg=C_MORADO_CLR, fmt="#,##0", align="right", bold=True, fc=C_MORADO)
        dat(ws, r, col_tot_com, round(sum_com) if sum_com else "-",
            bg=C_MORADO_CLR, fmt="#,##0", align="right", bold=True, fc=C_MORADO)
        r += 1

    # ── Fila totales generales ────────────────────────────────────────────────
    dat(ws, r, 1, "TOTAL", bg=C_AZUL_OSC, bold=True, fc="FFFFFF", align="center")
    dat(ws, r, 2, "",      bg=C_AZUL_OSC)
    grand_ml = grand_com = 0.0
    for i, (yr, mes) in enumerate(meses):
        col_ini = 3 + i * 2
        t_ml  = tot_ml[(yr, mes)]
        t_com = tot_com[(yr, mes)]
        grand_ml  += t_ml
        grand_com += t_com
        for col, val in ((col_ini, t_ml), (col_ini + 1, t_com)):
            c = ws.cell(row=r, column=col, value=round(val))
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = _fill(C_AZUL_MED); c.border = _border
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right", vertical="center")
    for col, val in ((col_tot_ml, grand_ml), (col_tot_com, grand_com)):
        c = ws.cell(row=r, column=col, value=round(val))
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = _fill(C_MORADO); c.border = _border
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A4"
    if r > 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(total_col)}{r}"


# ══════════════════════════════════════════════════════════════════════════════
# FUNCION PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def run(FI=None, FF=None, MOD="Random Forest", REF=None):
    """
    Genera el plan de produccion diario y guarda en Excel.

    Parametros
    ----------
    FI  : str   fecha inicio  ej. "2026-06-01"  (por defecto: primer dia del mes actual)
    FF  : str   fecha fin     ej. "2026-12-31"  (por defecto: ultimo dia del año actual)
    MOD : str   modelo de proyeccion: "Random Forest" | "XGBoost" |
                "Reg. Regularizada" | "Red Neuronal"
    REF : str   nombre exacto de producto, o None para todos
    """
    # Resolver FI/FF dinamicamente si no se proporcionan
    if FI is None or FF is None:
        from datetime import datetime as _dt
        _hoy = _dt.today()
        if FI is None:
            FI = _hoy.replace(day=1).strftime("%Y-%m-%d")
        if FF is None:
            FF = _dt(_hoy.year, 12, 31).strftime("%Y-%m-%d")

    # Asegurar que paso4_proyeccion puede importarse desde el mismo directorio
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from paso4_proyeccion import proyectar

    SEP = "=" * 62
    print(f"\n{SEP}")
    print(f"  PLAN DE PRODUCCION DIARIO  |  {FI} -> {FF}")
    print(f"  Modelo: {MOD}   |   Referencia: {REF or 'todas'}")
    print(SEP)

    # 1. Proyeccion mensual de demanda
    print("\n[1/6] Generando proyeccion de demanda (paso 4)...")
    proj_df = proyectar(FI=FI, FF=FF, MOD=MOD, REF=REF)

    # Filtrar productos no fabricables (servicios, fletes, obsoletos)
    antes = len(proj_df["producto"].unique())
    pattern = "|".join(SKIP_NOMBRE_KW)
    proj_df = proj_df[~proj_df["producto"].str.contains(pattern, case=False, na=False)]
    filtrados = antes - len(proj_df["producto"].unique())
    if filtrados:
        print(f"      Filtrados {filtrados} productos no fabricables (servicios/obsoletos)")

    # 2. Inventario actual
    print("[2/6] Cargando inventario (Bodega.xlsx)...")
    stock     = _cargar_bodega()
    bodega_df = _cargar_bodega_df()
    print(f"      {len(stock):,} codigos en stock activo  |  "
          f"total unidades: {int(sum(stock.values())):,}")

    # 3. OPs en proceso
    print("[3/6] Cargando ordenes de produccion en proceso...")
    ops_df, ops_supply = _cargar_ops_proceso()
    sin_cant = int((ops_df["Cant. Aprobada"] == 0).sum())
    print(f"      {len(ops_df)} OPs en proceso"
          + (f"  ({sin_cant} sin cantidad definida, se tratan como 0)" if sin_cant else ""))

    # 4. Mapeos codigo <-> nombre, lotes historicos y lotes minimos
    print("[4/6] Construyendo mapeos de productos, tamanos de lote y lotes minimos...")
    name_to_code, _ = _build_name_code_map()
    batch_sizes      = _infer_batch_sizes()
    lote_min         = _cargar_lote_minimo()
    codigos_stock    = _cargar_codigos_stock()
    print(f"      {len(lote_min)} referencias con lote minimo definido  |  {len(codigos_stock)} en lista de stock")

    # 5. MRP mensual + ajuste por lote minimo (integrado en el rollforward)
    print("[5/6] Calculando neto a producir (MRP) y ajustando a lotes minimos...")
    mrp_df       = _mrp_mensual(proj_df, stock, ops_supply, name_to_code, lote_min)
    mrp_df       = _aplicar_lote_minimo(mrp_df, lote_min)  # garantiza columnas
    proj_codigos = set(mrp_df["codigo_pt"].unique())
    sin_codigo   = proj_df["producto"].nunique() - len(proj_codigos)
    if sin_codigo:
        print(f"      Advertencia: {sin_codigo} producto(s) sin codigo mapeado (omitidos)")
    ajustados = int((mrp_df["neto_producir"] != mrp_df["neto_sin_ajuste"]).sum())
    print(f"      {len(proj_codigos)} referencias proyectadas  |  "
          f"{ajustados} registros ajustados por lote minimo")

    # 6. Distribucion diaria + pedidos ocasionales
    print("[6/6] Distribuyendo en dias habiles y procesando pedidos ocasionales...")
    plan_proy = _distribuir_diario(mrp_df, FI, FF, batch_sizes)
    plan_proy = plan_proy[plan_proy["codigo_pt"].isin(codigos_stock)].copy()

    pedidos_df = _cargar_pedidos(FI, FF)
    # Solo los Genericos (codigos_stock) estan cubiertos por el plan ML.
    # Todo pedido de otra linea es "ocasional" aunque el ML lo haya proyectado.
    n_ocas_raw = len(pedidos_df[~pedidos_df["Código PT"].isin(codigos_stock)])
    print(f"      {len(pedidos_df)} pedidos pendientes  |  {n_ocas_raw} con pedido activo fuera de lista")

    plan_ocas, tabla_ocas = _planificar_ocasionales(
        pedidos_df, stock, ops_supply, codigos_stock, batch_sizes, lote_min, FI, FF
    )

    # Consolidar plan completo
    plan_total = (pd.concat([plan_proy, plan_ocas], ignore_index=True)
                    .sort_values(["fecha", "codigo_pt"])
                    .reset_index(drop=True))

    # Marcar productos fuera de la lista de stock
    plan_total["categoria"] = plan_total["codigo_pt"].apply(
        lambda c: "Stock" if c in codigos_stock else "Pedido Activo"
    )

    # ── Escribir Excel ────────────────────────────────────────────────────────
    os.makedirs(RUTA_RES, exist_ok=True)
    ruta_out = RUTA_RES + "orden_produccion_final.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    # Enriquecer bodega con nombre del producto para la hoja Inventario
    _, code_to_name = _build_name_code_map()
    bodega_df = bodega_df.copy()
    bodega_df["ReferenciaP"] = bodega_df["Cód. PT"].map(code_to_name)
    # Fallback: usar columna Referencia del propio archivo de bodega
    if "Referencia" in bodega_df.columns:
        sin_nombre = bodega_df["ReferenciaP"].isna() | (bodega_df["ReferenciaP"] == "")
        bodega_df.loc[sin_nombre, "ReferenciaP"] = bodega_df.loc[sin_nombre, "Referencia"].fillna("")
    bodega_df["ReferenciaP"] = bodega_df["ReferenciaP"].fillna("")

    # Cargar plan comercial una vez (se usa en Dashboard y Comparativa)
    pc_df = None
    pc_name_map = {}
    if ARCH_PLAN_COMERCIAL and os.path.isfile(ARCH_PLAN_COMERCIAL):
        try:
            pc_df, pc_name_map = _cargar_plan_comercial(ARCH_PLAN_COMERCIAL)
        except Exception as exc:
            print(f"      Advertencia Plan Comercial: {exc}")

    # Cargar entradas para cruzar con OPs (Producida / Pendiente / Estado)
    ent_df = None
    if ARCH_ENTRADAS and os.path.isfile(ARCH_ENTRADAS):
        try:
            ent_df = pd.read_excel(ARCH_ENTRADAS)
            ops_df = _enriquecer_ops_con_entradas(ops_df, ent_df)
            print("      OPs enriquecidas con Cant. Producida/Pendiente desde entradas.")
        except Exception as exc:
            print(f"      Advertencia enriquecimiento OPs: {exc}")

    _hoja_plan_diario(wb, plan_total)
    _hoja_plan_mensual(wb, plan_total)
    _hoja_resumen_mrp(wb, mrp_df)
    _hoja_calendario_semanal(wb, plan_total)
    _hoja_ops_proceso(wb, ops_df)
    _hoja_pedidos_ocasionales(wb, tabla_ocas)
    _hoja_inventario(wb, bodega_df)

    if ent_df is not None:
        try:
            _hoja_saldo_ops(wb, ops_df, ent_df)
            print("      Hoja Saldo_OPs generada.")
        except Exception as exc:
            print(f"      Advertencia Saldo_OPs: {exc}")

    if pc_df is not None and not pc_df.empty:
        try:
            _hoja_comparativa_comercial(wb, mrp_df, pc_df, pc_name_map)
            print("      Hoja Comparativa_Comercial generada.")
        except Exception as exc:
            print(f"      Advertencia Comparativa_Comercial: {exc}")

    _hoja_pedidos_sin_cubrir(wb, pedidos_df, stock, ops_supply, plan_total)
    _hoja_dashboard(wb, plan_total, mrp_df, tabla_ocas,
                    stock, ops_df, pedidos_df, lote_min, codigos_stock, pc_df)

    try:
        wb.save(ruta_out)
    except PermissionError:
        # El archivo esta abierto en Excel; guardar con sello de tiempo
        from datetime import datetime
        ts       = datetime.now().strftime("%H%M%S")
        ruta_out = RUTA_RES + f"orden_produccion_final_{ts}.xlsx"
        wb.save(ruta_out)
        print(f"\n  AVISO: el archivo original estaba abierto en Excel.")
        print(f"         Se guardo una copia en: {os.path.basename(ruta_out)}")
        print(f"         Cierra el archivo en Excel y vuelve a ejecutar para sobreescribirlo.")

    print(f"\n{SEP}")
    print(f"  RESULTADO")
    print(f"    Dias planificados     : {plan_total['fecha'].nunique() if not plan_total.empty else 0}")
    print(f"    Referencias unicas    : {plan_total['codigo_pt'].nunique() if not plan_total.empty else 0}")
    print(f"    Unidades a producir   : {int(plan_total['cantidad'].sum()) if not plan_total.empty else 0:,}")
    print(f"    Pedidos ocasionales   : {len(tabla_ocas)}")
    print(f"    Guardado en           : {ruta_out}")
    print(SEP + "\n")

    return plan_total, ruta_out


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACION - edita aqui los parametros y ejecuta el script
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":

    # Rango de planificacion (formato "YYYY-MM-DD")
    from datetime import datetime as _dt
    _hoy_main = _dt.today()
    FI = _hoy_main.replace(day=1).strftime("%Y-%m-%d")
    FF = _dt(_hoy_main.year, 12, 31).strftime("%Y-%m-%d")

    # Modelo de proyeccion — elige UNA opcion:
    #   "Random Forest"
    #   "XGBoost"
    #   "Reg. Regularizada"
    #   "Red Neuronal"
    MOD = "Random Forest"

    # Producto especifico (nombre exacto) o None para todos
    REF = None

    # ─────────────────────────────────────────────────────────────────────────
    run(FI=FI, FF=FF, MOD=MOD, REF=REF)
