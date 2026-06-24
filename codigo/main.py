"""
SEEDPACK - Pipeline principal
Ejecuta los 5 pasos en secuencia:
  PASO 1 → Limpieza de datos
  PASO 2 → Clusterizacion de productos
  PASO 3 → Modelos ML de prediccion
  PASO 4 → Proyeccion de ventas
  PASO 5 → Ordenes de produccion
"""

import os
import sys
import time

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, "pasos"))

import paso1_limpieza
import paso2_clusterizacion
import paso3_modelos
import paso4_proyeccion
import paso5_orden_produccion

# ── Parametros de proyeccion (PASO 4) ─────────────────────────────────────────
from datetime import datetime as _dt
_hoy = _dt.today()
FI  = _hoy.replace(day=1).strftime("%Y-%m-%d")    # Primer dia del mes actual
FF  = _dt(_hoy.year, 12, 31).strftime("%Y-%m-%d") # Ultimo dia del año actual
MOD = "Random Forest"  # Modelo: "Random Forest" | "XGBoost" | "Reg. Regularizada" | "Red Neuronal"
REF = None             # Producto especifico o None para todos


def separador(titulo):
    print("\n" + "=" * 65)
    print(f"  {titulo}")
    print("=" * 65)


def _gestion_pedidos(arch_pedidos, arch_bodega, arch_ops_proc,
                     arch_lote_min, ruta_output):
    """
    Paso 6: Cruza el listado de pedidos (hoja Maestro) con stock, OPs en proceso
    y la proyeccion ML generada en paso 5. Escribe la hoja 'Gestion_Pedidos'
    en el Excel de resultados.
    """
    import math
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 1. Leer hoja de pedidos (Maestro con header=3 o primera hoja disponible) ─
    print("  Leyendo listado de pedidos...")
    _xl_ped = pd.ExcelFile(arch_pedidos)
    df_raw = None
    if "Maestro" in _xl_ped.sheet_names:
        try:
            _df_m = pd.read_excel(_xl_ped, sheet_name="Maestro", header=3)
            if any(str(c).upper() == "PEDIDO" for c in _df_m.columns):
                df_raw = _df_m
        except Exception:
            pass
    if df_raw is None:
        df_raw = pd.read_excel(_xl_ped)

    try:
        import json as _json
        _mp = os.path.normpath(os.path.join(BASE_DIR, "..", "app", "column_mappings.json"))
        if os.path.isfile(_mp):
            with open(_mp, encoding="utf-8") as _f:
                _m = _json.load(_f).get("arch_pedidos", {})
            _rn = {v: k for k, v in _m.items() if v and v in df_raw.columns and v != k}
            if _rn:
                df_raw = df_raw.rename(columns=_rn)
    except Exception:
        pass

    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    col_up = {c.upper(): c for c in df_raw.columns}

    # Saldo pendiente: columna exacta o alias "Cantidad" (formato historico)
    sp_col = (col_up.get("SALDO PENDIENTE")
              or col_up.get("CANTIDAD")
              or col_up.get("CANT.")
              or "SALDO PENDIENTE")

    # Codigo PT: exacto, con acento, o busqueda parcial (ej: "Código PT")
    cod_col = (col_up.get("CODIGO")
               or col_up.get("CÓDIGO")
               or next((v for k, v in col_up.items()
                        if "CODIGO" in k or "CÓDIGO" in k), None)
               or "CODIGO")

    desc_col = (col_up.get("DESCRIPCION")
                or col_up.get("DESCRIPCIÓN")
                or col_up.get("REFERENCIA")
                or "DESCRIPCION")
    ped_col  = col_up.get("PEDIDO", "PEDIDO")

    df_raw[sp_col] = pd.to_numeric(df_raw[sp_col], errors="coerce").fillna(0)
    needed = [c for c in [ped_col, cod_col, desc_col, sp_col] if c in df_raw.columns]
    df_pend = df_raw[df_raw[sp_col] > 0][needed].copy()
    df_pend = df_pend[df_pend[cod_col].notna()]
    df_pend[cod_col] = df_pend[cod_col].astype(str).str.strip()

    # ── 2. Stock en bodega ────────────────────────────────────────────────────
    stock_dict = {}
    if arch_bodega and os.path.isfile(arch_bodega):
        try:
            df_bod = pd.read_excel(arch_bodega)
            cols_b = {c.lower().strip(): c for c in df_bod.columns}
            if "bodega" not in cols_b:
                for alias in ("bodega sale", "bodega_sale", "almacen"):
                    if alias in cols_b:
                        df_bod = df_bod.rename(columns={cols_b[alias]: "Bodega"}); break
            pt_ok = any("pt" in c.lower() and ("cód" in c.lower() or "cod" in c.lower())
                        for c in df_bod.columns)
            if not pt_ok:
                cols_b2 = {c.lower().strip(): c for c in df_bod.columns}
                for alias in ("prod. terminado", "prod.terminado", "producto terminado"):
                    if alias in cols_b2:
                        df_bod = df_bod.rename(columns={cols_b2[alias]: "Cód. PT"})
                        df_bod["Cód. PT"] = (df_bod["Cód. PT"].astype(str)
                                              .str.split(" - ").str[0].str.strip())
                        break
            else:
                for c in df_bod.columns:
                    if "pt" in c.lower() and ("cód" in c.lower() or "cod" in c.lower()):
                        df_bod = df_bod.rename(columns={c: "Cód. PT"}); break
            cols_b3 = {c.lower().strip(): c for c in df_bod.columns}
            if "saldo" not in cols_b3:
                for alias in ("cantidad", "cant.", "stock"):
                    if alias in cols_b3:
                        df_bod = df_bod.rename(columns={cols_b3[alias]: "Saldo"}); break
            if "Bodega" in df_bod.columns:
                df_bod = df_bod[~df_bod["Bodega"].isin(["Obsoletos"])]
            if "Saldo" in df_bod.columns and "Cód. PT" in df_bod.columns:
                df_bod["Saldo"] = pd.to_numeric(df_bod["Saldo"], errors="coerce").fillna(0)
                stock_dict = df_bod.groupby("Cód. PT")["Saldo"].sum().to_dict()
        except Exception as exc:
            print(f"  Advertencia bodega: {exc}")

    # ── 3. OPs en proceso ─────────────────────────────────────────────────────
    ops_dict = {}
    if arch_ops_proc and os.path.isfile(arch_ops_proc):
        try:
            df_ops = pd.read_excel(arch_ops_proc)
            df_ops["Cant. Aprobada"] = pd.to_numeric(
                df_ops.get("Cant. Aprobada", pd.Series(dtype=float)), errors="coerce").fillna(0)
            cod_ops = next((c for c in df_ops.columns
                            if "producto" in c.lower() and "cód" in c.lower()), None)
            if cod_ops:
                ops_dict = df_ops.groupby(cod_ops)["Cant. Aprobada"].sum().to_dict()
        except Exception as exc:
            print(f"  Advertencia OPs: {exc}")

    # ── 4. Lote minimo ────────────────────────────────────────────────────────
    lmin_dict = {}
    if arch_lote_min and os.path.isfile(arch_lote_min):
        try:
            df_lm = pd.read_excel(arch_lote_min, skiprows=2, header=None)
            df_lm.columns = ["codigo", "referencia", "linea", "cavidades", "lote_minimo"]
            df_lm["lote_minimo"] = pd.to_numeric(
                df_lm["lote_minimo"], errors="coerce").fillna(0).astype(int)
            lmin_dict = df_lm[df_lm["lote_minimo"] > 0].set_index(
                "codigo")["lote_minimo"].to_dict()
        except Exception as exc:
            print(f"  Advertencia lote minimo: {exc}")

    # ── 5. Proyeccion ML desde el Plan_Diario generado en paso 5 ──────────────
    proyeccion_dict = {}
    if os.path.isfile(ruta_output):
        try:
            df_plan = pd.read_excel(ruta_output, sheet_name="Plan_Diario", header=1)
            if "Tipo" in df_plan.columns:
                df_plan = df_plan[df_plan["Tipo"] == "Proyeccion"]
            if "Codigo PT" in df_plan.columns and "Cantidad a Producir" in df_plan.columns:
                df_plan["Cantidad a Producir"] = pd.to_numeric(
                    df_plan["Cantidad a Producir"], errors="coerce").fillna(0)
                proyeccion_dict = (df_plan.groupby("Codigo PT")["Cantidad a Producir"]
                                   .sum().to_dict())
        except Exception as exc:
            print(f"  Advertencia proyeccion: {exc}")

    # ── 6. Calcular gestion ───────────────────────────────────────────────────
    print("  Calculando gestion de pedidos...")
    grouped = df_pend.groupby(cod_col).agg(
        descripcion=(desc_col, "first"),
        n_pedidos=(ped_col, "count"),
        saldo_total=(sp_col, "sum"),
    ).reset_index()

    rows = []
    for _, r in grouped.iterrows():
        cod   = str(r[cod_col]).strip()
        desc  = str(r["descripcion"]).strip()
        n_ped = int(r["n_pedidos"])
        saldo = int(r["saldo_total"])
        stock = int(stock_dict.get(cod, 0))
        ops   = int(ops_dict.get(cod, 0))
        proy  = int(proyeccion_dict.get(cod, 0))
        disp  = stock + ops
        total = saldo + proy
        def_  = max(0, total - disp)
        lmin  = int(lmin_dict.get(cod, 0))

        if def_ > 0:
            a_prod = math.ceil(def_ / lmin) * lmin if lmin > 0 else def_
            n_lot  = a_prod // lmin if lmin > 0 else 1
            estado = "Solicitar OP"
        else:
            a_prod = 0; n_lot = 0; estado = "OK"

        parts = [f"{n_ped} pedido(s), saldo: {saldo:,} uds."]
        if proy > 0:
            parts.append(f"Proy. ML: {proy:,} uds.")
        parts.append(f"Disp: {stock:,} bod + {ops:,} OPs = {disp:,}.")
        if def_ > 0:
            if lmin > 0:
                parts.append(f"Deficit {def_:,} -> lote min {lmin:,} -> producir {a_prod:,} ({n_lot} lotes).")
            else:
                parts.append(f"Deficit {def_:,} uds. Sin lote min. definido.")
        else:
            parts.append(f"Cubierto. Excedente: {disp - total:,} uds.")

        rows.append([cod, desc, n_ped, saldo, stock, ops, proy,
                     disp, total, def_, lmin, a_prod, estado, " ".join(parts)])

    # ── 7. Escribir hoja Gestion_Pedidos en el Excel de resultados ────────────
    print("  Escribiendo hoja Gestion_Pedidos...")
    wb = openpyxl.load_workbook(ruta_output)
    if "Gestion_Pedidos" in wb.sheetnames:
        del wb["Gestion_Pedidos"]
    ws = wb.create_sheet("Gestion_Pedidos")

    C_AZ = "1F4E79"; C_VD = "1A6B3A"; C_AM = "8B5E00"
    bord = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF"),
    )

    HEADERS = ["Codigo", "Descripcion", "Pedidos", "Saldo Pendiente", "Stock Bodega",
               "OPs Proceso", "Proyeccion ML", "Disponible", "Demanda Total",
               "Deficit", "Lote Minimo", "A Producir", "Estado", "Justificacion"]
    WIDTHS  = [18, 40, 10, 15, 15, 15, 15, 15, 15, 12, 14, 14, 16, 80]

    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill(fill_type="solid", fgColor=C_AZ)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bord
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 30

    for ri, row_data in enumerate(rows, 2):
        is_op = (row_data[12] == "Solicitar OP")
        row_bg = "FFF3CD" if is_op else "D4EDDA"
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Calibri", size=10,
                                  color=(C_AM if is_op else C_VD) if ci == 13 else "000000",
                                  bold=(ci == 13))
            cell.fill      = PatternFill(fill_type="solid", fgColor=row_bg)
            cell.alignment = Alignment(horizontal="center" if ci not in (2, 14) else "left",
                                       vertical="center", wrap_text=(ci == 14))
            cell.border    = bord

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    wb.save(ruta_output)
    print(f"  {len(rows)} referencias analizadas en Gestion_Pedidos.")


def run(FI, FF, REF=None,
        arch_ventas=None, arch_ops_hist=None, arch_pedidos=None,
        arch_ops_proc=None, arch_bodega=None, arch_lote_min=None,
        arch_entradas=None, arch_plan_comercial=None):
    """
    Ejecuta el pipeline completo.
    Si se proporcionan las rutas de archivos, se usan en lugar de las rutas
    por defecto definidas en cada script.
    """
    inicio_total = time.time()

    # Configurar rutas externas en paso5 si se proporcionaron
    if all(v is not None for v in [arch_ventas, arch_ops_hist, arch_pedidos,
                                    arch_ops_proc, arch_bodega, arch_lote_min]):
        paso5_orden_produccion.configurar_rutas(
            arch_ventas=arch_ventas,
            arch_ops_hist=arch_ops_hist,
            arch_pedidos=arch_pedidos,
            arch_ops_proc=arch_ops_proc,
            arch_bodega=arch_bodega,
            arch_lote_min=arch_lote_min,
            arch_entradas=arch_entradas,
            arch_plan_comercial=arch_plan_comercial,
        )

    separador("SEEDPACK — PIPELINE COMPLETO")

    separador("PASO 1 — Limpieza de datos")
    t0 = time.time()
    paso1_limpieza.run(arch_ventas=arch_ventas)
    print(f"  ✓ Completado en {time.time() - t0:.1f}s")

    separador("PASO 2 — Clusterizacion de productos")
    t0 = time.time()
    paso2_clusterizacion.run()
    print(f"  ✓ Completado en {time.time() - t0:.1f}s")

    separador("PASO 3 — Modelos ML de prediccion")
    t0 = time.time()
    paso3_modelos.main()
    print(f"  ✓ Completado en {time.time() - t0:.1f}s")

    separador("PASO 4 — Proyeccion de ventas")
    t0 = time.time()
    paso4_proyeccion.proyectar(FI=FI, FF=FF, MOD=MOD, REF=REF)
    print(f"  ✓ Completado en {time.time() - t0:.1f}s")

    separador("PASO 5 — Ordenes de produccion")
    t0 = time.time()
    _, ruta_output = paso5_orden_produccion.run(FI=FI, FF=FF, MOD=MOD, REF=REF)
    print(f"  ✓ Completado en {time.time() - t0:.1f}s")

    if arch_pedidos and os.path.isfile(arch_pedidos):
        separador("PASO 6 — Gestion de pedidos")
        t0 = time.time()
        _gestion_pedidos(arch_pedidos, arch_bodega, arch_ops_proc,
                         arch_lote_min, ruta_output)
        print(f"  ✓ Completado en {time.time() - t0:.1f}s")

    duracion = time.time() - inicio_total
    separador(f"PIPELINE COMPLETADO en {duracion:.0f}s")
    print(f"  Resultado Final en : {os.path.dirname(ruta_output)}")
    print()
    return ruta_output


def main():
    run(FI=FI, FF=FF, REF=REF)


if __name__ == "__main__":
    main()
